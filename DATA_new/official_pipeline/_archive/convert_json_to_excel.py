#!/usr/bin/env python3
"""
从 JSON 文件批量导入到 Excel
测试脚本 - 验证转换的可靠性
"""
import json
import openpyxl
from pathlib import Path
from datetime import datetime
import sys

# Excel 文件路径
EXCEL_PATH = Path("RQ_test.xlsx")

# Sheet 名称
SHEET_BASELINE = "raw_coverage"
SHEET_GENERATED = "question-answer-our"

# 列定义
BASELINE_COLUMNS = [
    "qa_unique_id",
    "scene_id",
    "frame_id",
    "question",
    "answer",
    "q_type",
    "num_hop",
    "l0_nodes",
    "l1_edges",
    "l2_paths",
    "n_l0",
    "n_l1",
    "n_l2",
    "llm_ms",
    "success",
    "timestamp",
]

GENERATED_COLUMNS = [
    "qa_unique_id",
    "scene_id",
    "frame_id",
    "question",
    "answer",
    "q_type",
    "num_hop",
    "l0_nodes",
    "l1_edges",
    "l2_paths",
    "n_l0",
    "n_l1",
    "n_l2",
    "llm_ms",
    "success",
    "timestamp",
]

# JSON 字段名 → Excel 列名映射
FIELD_MAPPING = {
    "question_id": "qa_unique_id",
    "scene_name": "scene_id",
    "frame_idx": "frame_id",
    "template_type": "q_type",
    "total_ms": "llm_ms",
    "timestamp_end": "timestamp",
}


def create_excel():
    """创建新的 Excel 文件"""
    print(f"Creating Excel: {EXCEL_PATH}")

    wb = openpyxl.Workbook()

    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 创建 baseline sheet
    ws_baseline = wb.create_sheet(SHEET_BASELINE)
    ws_baseline.append(BASELINE_COLUMNS)

    # 创建 generated sheet
    ws_generated = wb.create_sheet(SHEET_GENERATED)
    ws_generated.append(GENERATED_COLUMNS)

    wb.save(str(EXCEL_PATH))
    wb.close()
    print(f"  Created with 2 sheets")


def load_json_files(json_dir):
    """加载所有 JSON 文件"""
    json_dir = Path(json_dir)

    if not json_dir.exists():
        print(f"[ERROR] Directory not found: {json_dir}")
        return []

    json_files = list(json_dir.glob("*_qa.json"))
    print(f"\nFound {len(json_files)} JSON files in {json_dir}")

    all_questions = []
    for json_file in sorted(json_files):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 处理 {"questions": [...]} 格式
                if isinstance(data, dict) and "questions" in data:
                    all_questions.extend(data["questions"])
                elif isinstance(data, list):
                    all_questions.extend(data)
                else:
                    all_questions.append(data)
        except Exception as e:
            print(f"  [WARN] Failed to load {json_file.name}: {e}")

    print(f"  Loaded {len(all_questions)} questions total")
    return all_questions


def write_to_excel_batch(questions, sheet_name, columns):
    """批量写入 Excel（一次性）"""
    print(f"\nWriting {len(questions)} rows to sheet '{sheet_name}'...")

    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb[sheet_name]

    # 批量添加行
    rows_added = 0
    for q in questions:
        row = []
        for col in columns:
            # 先尝试直接获取
            val = q.get(col, "")

            # 如果没有，尝试映射
            if val == "" and col in FIELD_MAPPING.values():
                # 反向查找原始字段名
                for json_key, excel_col in FIELD_MAPPING.items():
                    if excel_col == col:
                        val = q.get(json_key, "")
                        break

            # 序列化 list/dict
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            elif val is None:
                val = ""

            row.append(val)

        ws.append(row)
        rows_added += 1

        if rows_added % 1000 == 0:
            print(f"  Progress: {rows_added}/{len(questions)}")

    # 保存
    print(f"  Saving...")
    wb.save(str(EXCEL_PATH))
    wb.close()
    print(f"  Done: {rows_added} rows written")

    return rows_added


def verify_excel():
    """验证 Excel 文件"""
    print(f"\nVerifying Excel file...")

    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH))

        for sheet_name in [SHEET_BASELINE, SHEET_GENERATED]:
            ws = wb[sheet_name]
            row_count = ws.max_row - 1  # 减去表头
            print(f"  [{sheet_name}] {row_count} rows")

        wb.close()
        print(f"  Verification passed!")
        return True

    except Exception as e:
        print(f"  [ERROR] Verification failed: {e}")
        return False


def test_conversion(json_dir, max_files=5):
    """测试转换（只转换前几个文件）"""
    print("="*80)
    print("JSON to Excel Conversion Test")
    print("="*80)

    # 1. 创建 Excel
    create_excel()

    # 2. 加载 JSON（限制数量）
    json_dir = Path(json_dir)
    json_files = sorted(json_dir.glob("*_qa.json"))[:max_files]

    print(f"\nLoading first {len(json_files)} JSON files for testing...")

    all_questions = []
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 处理 {"questions": [...]} 格式
                if isinstance(data, dict) and "questions" in data:
                    questions = data["questions"]
                    all_questions.extend(questions)
                    print(f"  Loaded: {json_file.name} ({len(questions)} questions)")
                elif isinstance(data, list):
                    all_questions.extend(data)
                    print(f"  Loaded: {json_file.name} ({len(data)} questions)")
                else:
                    all_questions.append(data)
                    print(f"  Loaded: {json_file.name} (1 question)")
        except Exception as e:
            print(f"  [WARN] Failed: {json_file.name}: {e}")

    print(f"\nTotal questions: {len(all_questions)}")

    # 3. 写入 Excel
    if all_questions:
        write_to_excel_batch(all_questions, SHEET_GENERATED, GENERATED_COLUMNS)

    # 4. 验证
    verify_excel()

    print("\n" + "="*80)
    print("Test completed!")
    print("="*80)
    print(f"\nTest Excel file: {EXCEL_PATH.absolute()}")
    print("Please open it in Excel to verify:")
    print("  1. All columns are correct")
    print("  2. Data is properly formatted")
    print("  3. JSON arrays (l0_nodes, l1_edges, l2_paths) are readable")
    print("\nIf everything looks good, we can proceed with full conversion.")


def full_conversion(json_dir):
    """完整转换（所有文件）"""
    print("="*80)
    print("Full JSON to Excel Conversion")
    print("="*80)

    start_time = datetime.now()

    # 1. 创建 Excel
    create_excel()

    # 2. 加载所有 JSON
    all_questions = load_json_files(json_dir)

    if not all_questions:
        print("[ERROR] No questions loaded")
        return

    # 3. 写入 Excel
    write_to_excel_batch(all_questions, SHEET_GENERATED, GENERATED_COLUMNS)

    # 4. 验证
    verify_excel()

    elapsed = datetime.now() - start_time
    print("\n" + "="*80)
    print(f"Conversion completed in {elapsed.total_seconds():.1f} seconds")
    print("="*80)
    print(f"\nOutput: {EXCEL_PATH.absolute()}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Convert JSON to Excel")
    parser.add_argument("--json-dir", default="/home/yunyang/ADVTEST/DATA_new/generated_qa",
                        help="Directory containing JSON files")
    parser.add_argument("--test", action="store_true",
                        help="Test mode: only convert first 5 files")
    parser.add_argument("--full", action="store_true",
                        help="Full conversion: convert all files")

    args = parser.parse_args()

    if args.test:
        test_conversion(args.json_dir, max_files=5)
    elif args.full:
        full_conversion(args.json_dir)
    else:
        print("Usage:")
        print("  Test mode:  python convert_json_to_excel.py --test")
        print("  Full mode:  python convert_json_to_excel.py --full")
        print("  Custom dir: python convert_json_to_excel.py --test --json-dir /path/to/json")
