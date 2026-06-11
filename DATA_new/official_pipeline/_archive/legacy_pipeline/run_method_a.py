#!/usr/bin/env python3
"""
run_method_a.py — 方案 A 闭环执行链 (scene-0926 frame-20)

执行前检查清单（必须全部通过才能运行）：
  [OK] RQ.xlsx 已关闭（无 ~$RQ.xlsx 锁定文件）
  [OK] filtered_scene_graphs/scene-0926_frame20_scene_graph.json 存在
  [OK] Neo4j（默认 7687 端口，见 NEO4J_URI）可达

执行链：
  Step 1 : 净化环境（清理脏数据行、删除锁定文件）
  Step 2 : 写入 filter_record（核心宇宙过滤元数据）
  Step 3 : 清空 Neo4j + 从 filtered_scene_graphs/ 导入 scene-0926-20
  Step 4 : Baseline 审计（29 条原题 -> LLM Cypher -> 足迹 -> raw_coverage）
  Step 5 : Gap Detection（未覆盖 L2 路径）
  Step 6 : 增量生成（5 条 L2 题 -> question-answer-our）
  Step 7 : Final save 确认
"""
import os, sys, pathlib, json, time, collections, subprocess
sys.path.insert(0, str(pathlib.Path(__file__).parent))

from advtest_env import load_advtest_env
load_advtest_env()

from advtest_paths import (
    EXCEL_PATH,
    FILTERED_SG_DIR,
    GEN_QA_DIR,
    NEO4J_PASSWORD as NEO4J_PWD,
    NEO4J_URI,
    NEO4J_USER,
    TRAINVAL_META as TRAINVAL,
    VQA_QA_JSON as QA_PATH,
)

from coverage_tracker_patch import install_patch
install_patch()

from gap_pipeline.template_library import get_template_library
import random

FSG_DIR     = pathlib.Path(FILTERED_SG_DIR)
ALLOW_EXCEL_DESTRUCTIVE_CLEAN = os.getenv("VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN", "false").lower() in ("true", "1", "yes")

# ---------------------------------------------------------------------------
#  Plan 接口: 从 JSON 读取帧列表，或使用硬编码默认值
#
#  用法:
#    1. 环境变量: set ADVTEST_FRAME_PLAN_JSON=path/to/plan.json
#    2. 命令行:   python run_method_a.py scene-0103 0
#    3. 直接改:   修改下方 _DEFAULT_FRAMES
# ---------------------------------------------------------------------------
_DEFAULT_FRAMES = [
    ("scene-0103", 0, "scene-0103_frame0_scene_graph.json"),
]

def _load_plan_frames():
    """加载帧计划: 环境变量 > 命令行 > 默认值"""
    plan_path = os.getenv("ADVTEST_FRAME_PLAN_JSON", "").strip()
    if plan_path and os.path.isfile(plan_path):
        data = json.loads(open(plan_path, encoding="utf-8").read())
        frames = []
        for item in data.get("frames", []):
            sid = item["scene_id"]
            fid = item["frame_id"]
            sg = item.get("sg_file", f"{sid}_frame{fid}_scene_graph.json")
            frames.append((sid, fid, sg))
        if frames:
            print(f"  [Plan] Loaded {len(frames)} frames from {plan_path}")
            return frames
    if len(sys.argv) >= 3:
        sid, fid = sys.argv[1], int(sys.argv[2])
        sg = f"{sid}_frame{fid}_scene_graph.json"
        return [(sid, fid, sg)]
    return list(_DEFAULT_FRAMES)

# 当前帧变量（由 __main__ 循环设置）
TARGET_SG   = _DEFAULT_FRAMES[0][2]
SCENE_ID    = _DEFAULT_FRAMES[0][0]
FRAME_ID    = _DEFAULT_FRAMES[0][1]


def _console_safe(s: object) -> str:
    """Strip/replace chars the Windows console (e.g. GBK) cannot print (LLM may emit ✓/❌)."""
    if s is None:
        return ""
    t = str(s)
    enc = getattr(sys.stdout, "encoding", None) or "utf-8"
    try:
        return t.encode(enc, errors="replace").decode(enc, errors="replace")
    except (LookupError, UnicodeError):
        return t.encode("ascii", errors="replace").decode("ascii")


# ─────────────────────────────────────────────────────────────────────────────
# Optional: school VPN (Windows rasdial)
# ─────────────────────────────────────────────────────────────────────────────

def maybe_connect_school_vpn() -> bool:
    """
    若 VQA_AUTO_CONNECT_VPN 为真：在 Windows 上按环境变量尝试拨号（rasdial）。
    账号口令只应写在 advtest_runtime.env（已在 .gitignore），勿提交仓库。
    """
    if os.name != "nt":
        return True
    v = os.getenv("VQA_AUTO_CONNECT_VPN", "").strip().lower()
    if v not in ("1", "true", "yes", "on"):
        return True
    conn = os.getenv("SCHOOL_VPN_CONN_NAME", "").strip()
    user = os.getenv("SCHOOL_VPN_USER", "").strip()
    pwd = os.getenv("SCHOOL_VPN_PASS", "").strip()
    if not conn or not user or not pwd:
        print("[FAIL][VPN] 已开启 VQA_AUTO_CONNECT_VPN，但缺少 SCHOOL_VPN_CONN_NAME / SCHOOL_VPN_USER / SCHOOL_VPN_PASS。")
        return False
    script = pathlib.Path(__file__).resolve().parent / "scripts" / "connect_school_vpn.ps1"
    if not script.is_file():
        print(f"[FAIL][VPN] 脚本不存在: {script}")
        return False
    print(f"\n[VPN] 预连校园网（Windows 连接名: {conn}）…")
    try:
        r = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script),
            ],
            env=os.environ.copy(),
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        print("[FAIL][VPN] 拨号超时（120s）。")
        return False
    except Exception as e:
        print(f"[FAIL][VPN] 执行失败: {e}")
        return False
    if r.returncode != 0:
        print(f"[FAIL][VPN] rasdial 退出码 {r.returncode}。请检查 Windows VPN 条目名、账号口令与学校网关。")
        return False
    print("[VPN] 拨号成功。")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Pre-flight checks
# ─────────────────────────────────────────────────────────────────────────────

def preflight_check() -> bool:
    print("\n[Pre-flight checks]")
    ok = True
    onthefly = os.getenv("VQA_BUILD_SCENE_GRAPH_ONTHEFLY", "").strip().lower() in ("1", "true", "yes", "on")

    # Trainval 元数据 + QA JSON（6019 帧抽自全量 trainval）
    if not TRAINVAL.is_dir():
        print(f"  [FAIL] Trainval metadata missing: {TRAINVAL}")
        print("     设置 NUSCENES_DATAROOT / NUSCENES_VERSION（默认 dataroot 为 $ADVTEST_ROOT/data）")
        ok = False
    else:
        for _fn in ("scene.json", "sample.json"):
            if not (TRAINVAL / _fn).is_file():
                print(f"  [FAIL] Missing {TRAINVAL / _fn}")
                ok = False
    if not QA_PATH.is_file():
        print(f"  [FAIL] NuScenes-QA JSON missing: {QA_PATH}")
        ok = False

    # 1. Kill lock file (OfficeClickToRun keeps recreating it — force-delete and test write)
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        print(f"  [WARN]  Lock file found: {lock} — attempting force-delete...")
        # Kill OfficeClickToRun silently to prevent immediate recreation
        subprocess.run(
            ["powershell", "-Command",
             "Get-Process | Where-Object { $_.Name -match 'EXCEL|OfficeClickToRun' } | "
             "ForEach-Object { Stop-Process -Id $_.Id -Force -ErrorAction SilentlyContinue }"],
            capture_output=True, timeout=10
        )
        import time as _t; _t.sleep(1)
        try:
            lock.unlink(missing_ok=True)
            print(f"  [OK] Lock file deleted")
        except Exception as e:
            print(f"  [WARN]  Could not delete lock: {e}")
    else:
        print(f"  [OK] No lock file")

    # 2. Filtered SG exists (skip in on-the-fly mode)
    sg = FSG_DIR / TARGET_SG
    if onthefly:
        print("  [OK] VQA_BUILD_SCENE_GRAPH_ONTHEFLY=true -> skip disk filtered SG pre-check")
    elif sg.exists():
        data = json.loads(sg.read_text(encoding="utf-8"))
        n = data.get("core_universe_filter", {}).get("filtered_nodes", "?")
        print(f"  [OK] Filtered SG exists: {sg.name} ({n} nodes)")
    else:
        print(f"  [FAIL] Filtered SG missing: {sg}")
        ok = False

    # 3. Neo4j reachable
    try:
        from neo4j import GraphDatabase
        d = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))
        with d.session() as s:
            s.run("RETURN 1").single()
        d.close()
        print(f"  [OK] Neo4j reachable: {NEO4J_URI}")
    except Exception as e:
        print(f"  [FAIL] Neo4j not reachable: {e}")
        ok = False

    # 4. Excel actual write test (the TRUE test — lock file existence alone is not sufficient)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        wb.save(str(EXCEL_PATH))
        wb.close()
        print(f"  [OK] Excel writable: {EXCEL_PATH}")
    except PermissionError:
        print(f"  [FAIL] Excel write test FAILED — file is still locked by another process")
        ok = False
    except Exception as e:
        print(f"  [WARN]  Excel check error: {e}")

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Step 0: Wipe previous run data (DISABLED by default, append-only policy)
# ─────────────────────────────────────────────────────────────────────────────

def step0_wipe_previous_run():
    """Destructive cleanup (opt-in only). Default behavior is append-only."""
    print("\n[Step 0] Destructive Excel wipe")
    if not ALLOW_EXCEL_DESTRUCTIVE_CLEAN:
        print("  [SKIP]  Skipped (append-only mode; VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN is false)")
        return
    print("  [WARN]  Enabled by VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN=true")
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        total_deleted = 0
        for sh in ["raw_coverage", "question-answer-our"]:
            ws = wb[sh]
            max_row = ws.max_row
            # Delete from bottom to preserve row indices
            for r in range(max_row, 1, -1):   # row 1 = header, skip
                ws.delete_rows(r)
                total_deleted += 1
        wb.save(str(EXCEL_PATH))
        wb.close()
        print(f"  [OK] Deleted {total_deleted} data rows (headers preserved)")
    except Exception as e:
        print(f"  [WARN]  Wipe error: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Environment cleanup (non-destructive, lock file only)
# ─────────────────────────────────────────────────────────────────────────────

def step1_cleanup():
    print("\n[Step 1] Environment cleanup (append-only, non-destructive)")

    # Remove lock file if still present
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        try:
            lock.unlink()
            print(f"  [OK] Deleted lock file: {lock}")
        except Exception as e:
            print(f"  [WARN]  Could not delete lock: {e}")
    print("  [OK] Keep all historical rows (no row deletion)")


# ─────────────────────────────────────────────────────────────────────────────
# Step 1.5: Rebuild filtered SG with official policy (30/40/50m + vis + px)
# ─────────────────────────────────────────────────────────────────────────────
def step1_5_refresh_filtered_sg() -> bool:
    print("\n[Step 1.5] Rebuilding filtered SG with official policy")

    # 检查是否使用动态生成模式
    onthefly = os.getenv("VQA_BUILD_SCENE_GRAPH_ONTHEFLY", "").strip().lower() in ("1", "true", "yes", "on")
    if onthefly:
        print("  [SKIP] VQA_BUILD_SCENE_GRAPH_ONTHEFLY=true -> skip filtered SG rebuild")
        return True

    from core_universe_filter import (
        RAW_SG_DIR,
        FILTERED_SG_DIR,
        MIN_VISIBILITY,
        filter_and_save,
    )
    raw_path = RAW_SG_DIR / TARGET_SG
    if not raw_path.exists():
        print(f"  [FAIL] Raw SG missing: {raw_path}")
        return False
    out = filter_and_save(
        raw_path=raw_path,
        write_excel=False,
        distance_mode="official",
        pixel_mode=os.getenv("VQA_PIXEL_MODE", "lenient"),
        min_visibility=float(os.getenv("VQA_FILTER_MIN_VISIBILITY", str(MIN_VISIBILITY))),
        output_dir=FILTERED_SG_DIR,
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    info = data.get("core_universe_filter", {})
    print(
        f"  [OK] {out.name}: nodes {info.get('raw_nodes', '?')} -> {info.get('filtered_nodes', '?')}, "
        f"edges {info.get('raw_edges', '?')} -> {info.get('filtered_edges', '?')}, "
        f"mode={info.get('distance_mode', '?')}, pixel={info.get('pixel_mode', '?')}"
    )
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Write filter_record
# ─────────────────────────────────────────────────────────────────────────────

def step2_filter_record():
    print("\n[Step 2] Writing filter_record")
    from rq_tables import write_filter_record

    sg = FSG_DIR / TARGET_SG
    data = json.loads(sg.read_text(encoding="utf-8"))

    # Check if core_universe_filter exists (from official filter pipeline)
    # If not, use statistics from generated scene graph
    if "core_universe_filter" in data:
        info = data["core_universe_filter"]
        vex_str = ",".join(sorted(info["node_ids_kept"]))
        ratio = info["filtered_nodes"] / max(info["raw_nodes"], 1)
        raw_nodes = info["raw_nodes"]
        filtered_nodes = info["filtered_nodes"]
    else:
        # Scene graph generated directly without filtering
        print("  [INFO] No core_universe_filter found, using direct scene graph statistics")
        stats = data.get("statistics", {})
        nodes = data.get("nodes", [])
        node_ids = [n.get("unique_id") for n in nodes if "unique_id" in n]
        vex_str = ",".join(sorted(node_ids))
        raw_nodes = stats.get("total_objects", len(nodes))
        filtered_nodes = len(nodes)
        ratio = 1.0  # No filtering applied

    ok = write_filter_record(
        scene_id=SCENE_ID, frame_id=FRAME_ID,
        original_num=raw_nodes,
        filtered_num=filtered_nodes,
        filtered_vex=vex_str,
        ratio=ratio,
    )
    print(f"  scene={SCENE_ID} frame={FRAME_ID}")
    print(f"  raw={raw_nodes} -> filtered={filtered_nodes} "
          f"(ratio={ratio:.2%})")
    print(f"  node_ids: {len(node_ids) if 'node_ids' in locals() else 'N/A'}")
    print(f"  write_filter_record: {'[OK]' if ok else '[FAIL]'}")
    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: Neo4j import from filtered_scene_graphs/
# ─────────────────────────────────────────────────────────────────────────────

def step3_import_neo4j():
    print("\n[Step 3] Importing filtered SG to Neo4j")
    from core_universe_filter import import_filtered_sg_to_neo4j
    result = import_filtered_sg_to_neo4j(
        sg_name=TARGET_SG,
        neo4j_uri=NEO4J_URI,
        neo4j_user=NEO4J_USER,
        neo4j_pwd=NEO4J_PWD,
    )
    print(f"  [OK] Neo4j: {result['n_nodes']} nodes, {result['n_edges']} edges")
    print(f"  Source : {result['source']}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Step 4: Baseline audit (29 original questions -> LLM Cypher -> footprint -> raw_coverage)
# ─────────────────────────────────────────────────────────────────────────────

def step4_baseline_audit(driver, llm_client):
    print("\n[Step 4] Baseline audit (original NuScenes-QA questions)")
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    # V15: 改进的 baseline 覆盖率分析（环境变量控制）
    USE_V15_AUDITOR = bool(os.getenv("VQA_USE_V15_AUDITOR", "true").lower() in ("true", "1", "yes"))

    if USE_V15_AUDITOR:
        from semantic_auditor_v15 import audit_baseline_question_v15 as audit_baseline_question
        from semantic_auditor_v15 import build_scene_context
        print("  Using V15 auditor (improved anchor recognition + wider direction matching)")
    else:
        from semantic_auditor import audit_baseline_question, build_scene_context
        print("  Using V14 auditor (legacy)")

    from rq_tables import write_baseline_to_coverage

    # Build sample->scene mapping
    scenes  = json.loads((TRAINVAL/"scene.json").read_text())
    samples = json.loads((TRAINVAL/"sample.json").read_text())
    st2name = {s["token"]: s["name"] for s in scenes}
    s2tok: dict = collections.defaultdict(list)
    tok2info: dict = {}
    for samp in samples:
        sname = st2name.get(samp["scene_token"],"?")
        tok2info[samp["token"]] = {"scene_name": sname, "timestamp": samp["timestamp"]}
        s2tok[sname].append(samp["token"])
    for sname, toks in s2tok.items():
        for idx, tok in enumerate(sorted(toks, key=lambda t: tok2info[t]["timestamp"])):
            tok2info[tok]["frame_idx"] = idx

    # Get scene-0926 val questions  —  keep global index for unique ID generation
    all_val_qs = json.loads(QA_PATH.read_text())["questions"]
    target_qs = [
        (global_i, q)
        for global_i, q in enumerate(all_val_qs)     # global_i = position in val file
        if tok2info.get(q.get("sample_token",""), {}).get("scene_name") == SCENE_ID
        and tok2info.get(q.get("sample_token",""), {}).get("frame_idx") == FRAME_ID
    ]
    print(f"  Found {len(target_qs)} val questions for {SCENE_ID} frame-{FRAME_ID}")
    if target_qs:
        print(f"  Global indices: {target_qs[0][0]} – {target_qs[-1][0]}")

    if not target_qs:
        print("  [WARN]  No questions found — check sample_token mapping")
        return 0

    # Build scene context once (reuse across questions)
    scene_ctx = build_scene_context(driver)
    print(f"  Scene context built ({len(scene_ctx)} chars)")

    from semantic_auditor import make_qa_id, _ms_now as _audit_ms, derive_l2_from_l1
    BASELINE_N_WORKERS = int(os.getenv("VQA_BASELINE_N_WORKERS", "32"))
    BASELINE_L2_BACKFILL = os.getenv("VQA_BASELINE_L2_BACKFILL", "true").lower() in (
        "true",
        "1",
        "yes",
        "on",
    )
    BASELINE_L2_BACKFILL_MAX_PER_Q = max(
        1, int(os.getenv("VQA_BASELINE_L2_BACKFILL_MAX_PER_Q", "1"))
    )
    _tls = threading.local()
    _llm_cls = llm_client.__class__

    # Snapshot real graph entities once; baseline footprints are constrained to this set.
    with driver.session() as _sess:
        _valid_nodes = {
            str(r["id"]).strip()
            for r in _sess.run("MATCH (n:Object) RETURN n.unique_id AS id")
            if r.get("id")
        }
        _valid_edges = set()
        _valid_edge_dir4 = {}
        _valid_edge_dir8 = {}
        for r in _sess.run(
            "MATCH (s:Object)-[rel:RELATES_TO]->(t:Object) "
            "RETURN s.unique_id AS src, t.unique_id AS tgt, "
            "coalesce(rel.direction_4,'') AS dir4, coalesce(rel.direction_8,'') AS dir8"
        ):
            _src = str(r.get("src") or "").strip()
            _tgt = str(r.get("tgt") or "").strip()
            if not _src or not _tgt:
                continue
            _k = (_src, _tgt)
            _valid_edges.add(_k)
            _valid_edge_dir4[_k] = str(r.get("dir4") or "").strip().lower()
            _valid_edge_dir8[_k] = str(r.get("dir8") or "").strip().lower()
        _type_to_ids: dict = {}
        for r in _sess.run("MATCH (n:Object) RETURN n.unique_id AS id, n.type AS t"):
            uid = str(r.get("id") or "").strip()
            typ = str(r.get("t") or "").strip().lower()
            if uid and typ:
                _type_to_ids.setdefault(typ, []).append(uid)
        for _k in _type_to_ids:
            _type_to_ids[_k] = sorted(_type_to_ids[_k])

    def _candidates_for_type_label(lab: str) -> list:
        lab = (lab or "").strip().lower()
        if not lab:
            return []
        ids = list(_type_to_ids.get(lab, []))
        if not ids and len(lab) > 3 and lab.endswith("s"):
            ids = list(_type_to_ids.get(lab[:-1], []))
        return ids

    def _resolve_type_label(label: str) -> list:
        """类型词 -> 若当前图中该 type 唯一，则映射到唯一 unique_id。"""
        lab = (label or "").strip().lower()
        if not lab or lab == "ego" or lab == "any":
            return []
        cand = _candidates_for_type_label(lab)
        if len(cand) != 1:
            return []
        return list(cand)

    def _norm_node_id(x: object) -> str:
        s = str(x or "").strip()
        if not s:
            return ""
        if s in _valid_nodes:
            return s
        cand = _resolve_type_label(s)
        return cand[0] if len(cand) == 1 else ""
    def _edge_rel_match(src_id: str, tgt_id: str, rel_raw: str) -> bool:
        rel = str(rel_raw or "").strip().lower()
        if not rel:
            return True
        k = (src_id, tgt_id)
        d4 = _valid_edge_dir4.get(k, "")
        d8 = _valid_edge_dir8.get(k, "")
        if rel == d4 or rel == d8:
            return True
        if "-" in rel and d4 and rel.split("-", 1)[0] == d4:
            return True
        return False

    def _raw_to_candidates(raw_val: object) -> list:
        raw = str(raw_val or "").strip()
        if not raw:
            return []
        if raw in _valid_nodes:
            return [raw]
        if raw.lower() == "ego" and "ego" in _valid_nodes:
            return ["ego"]
        return _candidates_for_type_label(raw)

    def _norm_edges(raw_edges):
        out = []
        seen = set()
        for e in (raw_edges or []):
            if not isinstance(e, dict):
                continue
            src_raw = e.get("source") or e.get("src") or ""
            tgt_raw = e.get("target") or e.get("dst") or e.get("tgt") or ""
            rel = str(e.get("relation") or e.get("dir") or e.get("direction_4") or "").strip()
            src_cands = _raw_to_candidates(src_raw)
            tgt_cands = _raw_to_candidates(tgt_raw)
            if not src_cands or not tgt_cands:
                continue
            matched_pairs = []
            for src in src_cands:
                for tgt in tgt_cands:
                    if (src, tgt) not in _valid_edges:
                        continue
                    if not _edge_rel_match(src, tgt, rel):
                        continue
                    matched_pairs.append((src, tgt))
            if len(matched_pairs) != 1:
                continue
            key = matched_pairs[0]
            if key in seen:
                continue
            seen.add(key)
            item = {"source": key[0], "target": key[1]}
            if rel:
                item["relation"] = rel
            out.append(item)
        return out

    def _resolve_anchor_from_intent(audit_res: dict) -> str:
        _intent = audit_res.get("intent", {})
        if not isinstance(_intent, dict):
            return ""
        _anchor_hint = _norm_node_id(_intent.get("anchor_id_hint"))
        if _anchor_hint:
            return _anchor_hint
        _anchor_type = str(_intent.get("anchor_type") or "").strip().lower()
        if _anchor_type == "ego" and "ego" in _valid_nodes:
            return "ego"
        _cands = _candidates_for_type_label(_anchor_type)
        if len(_cands) == 1:
            return str(_cands[0]).strip()
        return ""

    def _backfill_l1_from_soft_matches(audit_res: dict) -> list:
        _anchor = _resolve_anchor_from_intent(audit_res)
        if not _anchor:
            return []
        _intent = audit_res.get("intent", {})
        _rel = str((_intent.get("relation_dir") if isinstance(_intent, dict) else "") or "").strip()
        _out = []
        _seen = set()
        for _m in (audit_res.get("soft_matches", []) or []):
            if not isinstance(_m, dict):
                continue
            _tgt = _norm_node_id(_m.get("id"))
            if not _tgt:
                continue
            if (_anchor, _tgt) not in _valid_edges:
                continue
            if _rel and not _edge_rel_match(_anchor, _tgt, _rel):
                continue
            _k = (_anchor, _tgt)
            if _k in _seen:
                continue
            _seen.add(_k)
            _item = {"source": _anchor, "target": _tgt}
            if _rel:
                _item["relation"] = _rel
            _out.append(_item)
        return _out

    def _norm_l2_paths(raw_paths) -> list:
        out = []
        seen = set()
        for p in (raw_paths or []):
            o1 = o2 = o3 = ""
            if isinstance(p, dict):
                o1 = _norm_node_id(p.get("o1"))
                o2 = _norm_node_id(p.get("o2"))
                o3 = _norm_node_id(p.get("o3"))
            elif isinstance(p, (list, tuple)) and len(p) >= 3:
                o1 = _norm_node_id(p[0])
                o2 = _norm_node_id(p[1])
                o3 = _norm_node_id(p[2])
            if not o1 or not o2 or not o3:
                continue
            if o1 == o3:
                continue
            if (o1, o2) not in _valid_edges or (o2, o3) not in _valid_edges:
                continue
            key = (o1, o2, o3)
            if key in seen:
                continue
            seen.add(key)
            out.append({"o1": o1, "o2": o2, "o3": o3})
        return out

    def _derive_l2_backfill_from_l1(l1_norm: list) -> list:
        """Backfill L2 pivots from L1 edges using undirected adjacency (pivot semantics)."""
        if len(l1_norm) < 2:
            return []
        # Build undirected adjacency
        neighbors: dict = {}
        for e in l1_norm:
            s = str(e.get("source") or "").strip()
            t = str(e.get("target") or "").strip()
            if s and t:
                neighbors.setdefault(s, set()).add(t)
                neighbors.setdefault(t, set()).add(s)
        out = []
        seen = set()
        for b, nbrs in neighbors.items():
            nbr_list = sorted(nbrs)
            for i in range(len(nbr_list)):
                for j in range(i + 1, len(nbr_list)):
                    a, c = nbr_list[i], nbr_list[j]
                    lo, hi = (a, c) if a <= c else (c, a)
                    key = (lo, b, hi)
                    if key not in seen:
                        seen.add(key)
                        out.append({"o1": a, "o2": b, "o3": c})
                        if len(out) >= BASELINE_L2_BACKFILL_MAX_PER_Q:
                            return out
        return out

    def _merge_l2_paths(*groups) -> list:
        out = []
        seen = set()
        for g in groups:
            for p in (g or []):
                if not isinstance(p, dict):
                    continue
                o1 = str(p.get("o1") or "").strip()
                o2 = str(p.get("o2") or "").strip()
                o3 = str(p.get("o3") or "").strip()
                if not o1 or not o2 or not o3:
                    continue
                key = (o1, o2, o3)
                if key in seen:
                    continue
                seen.add(key)
                out.append({"o1": o1, "o2": o2, "o3": o3})
        return out
    def _finalize_l0(raw_nodes, l1_norm: list, audit_res: dict) -> list:
        """
        L0 仅包含“有证据”的真实 unique_id（不做兜底补齐）：
          1) 审计结果里的节点ID；
          2) L1 端点；
          3) 审计结果里的类型词，仅当该类型在当前图中唯一时映射到该ID。
        """
        graph_ids: list = []
        gid_set: set = set()

        def _push_gid(x: str) -> None:
            s = str(x or "").strip()
            if not s or s not in _valid_nodes or s in gid_set:
                return
            gid_set.add(s)
            graph_ids.append(s)


        for n in (raw_nodes or []):
            s = str(n or "").strip()
            if not s:
                continue
            if s in _valid_nodes:
                _push_gid(s)
            else:
                for rid in _resolve_type_label(s):
                    _push_gid(rid)
        for e in l1_norm:
            _push_gid(e.get("source", ""))
            _push_gid(e.get("target", ""))

        if graph_ids:
            return graph_ids
        return []

    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            _llm = _llm_cls()
            _tls.llm = _llm
        return _llm

    def _audit_one(global_idx: int, q: dict) -> dict:
        question = q.get("question", "")
        answer = q.get("answer", "")
        qtype = q.get("template_type", "")
        nhop = q.get("num_hop", 0)
        qa_uid = make_qa_id(global_idx, qtype)
        ts0 = _audit_ms()
        try:
            audit_res = audit_baseline_question(
                question=question, q_type=qtype, num_hop=nhop,
                driver=driver, llm_client=_get_worker_llm(),
                scene_context=scene_ctx,
                global_index=global_idx,
            )
        except Exception as exc:
            print(f"    [WARN] [baseline-worker-exc] idx={global_idx} {_console_safe(exc)}")
            audit_res = {"l0_nodes": [], "l1_edges": [], "l2_paths": []}
        ts1 = _audit_ms()
        l1 = _norm_edges(audit_res.get("l1_edges", []) or [])
        if not l1:
            _l1_soft = _backfill_l1_from_soft_matches(audit_res)
            if _l1_soft:
                l1 = _l1_soft
        l0 = _finalize_l0(audit_res.get("l0_nodes", []) or [], l1, audit_res)
        l2_from_l1 = _norm_l2_paths(derive_l2_from_l1(l1))
        l2_from_audit = _norm_l2_paths(audit_res.get("l2_paths", []) or [])
        l2_backfilled = []
        if BASELINE_L2_BACKFILL and not l2_from_l1 and not l2_from_audit and len(l1) >= 2:
            l2_backfilled = _derive_l2_backfill_from_l1(l1)
        l2 = _merge_l2_paths(l2_from_l1, l2_from_audit, l2_backfilled)
        audit_json = json.dumps(
            {"nodes": l0, "edges": l1},
            ensure_ascii=False
        )
        return {
            "qa_uid": qa_uid,
            "question": question,
            "answer": answer,
            "qtype": qtype,
            "l0_nodes": l0,
            "l1_edges": l1,
            "l2_paths": l2,
            "l2_backfilled": bool(l2_backfilled),
            "timestamp_start": ts0,
            "timestamp_end": ts1,
            "audit_cypher": audit_json,
        }

    t_batch0 = time.perf_counter()
    ordered = [None] * len(target_qs)
    with ThreadPoolExecutor(max_workers=BASELINE_N_WORKERS) as _ex:
        fut2idx = {
            _ex.submit(_audit_one, global_idx, q): idx
            for idx, (global_idx, q) in enumerate(target_qs)
        }
        for done_i, _f in enumerate(as_completed(fut2idx), 1):
            idx = fut2idx[_f]
            rec = _f.result()
            ordered[idx] = rec
            if done_i <= 3 or done_i % 10 == 0:
                print(f"  [audit {done_i:2d}/{len(target_qs)}] {rec['qa_uid']:<25} "
                      f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} L2={len(rec['l2_paths'])}")
    batch_ms = int((time.perf_counter() - t_batch0) * 1000)
    avg_ms = int(batch_ms / max(len(target_qs), 1))
    _rows_with_l2 = sum(1 for _r in ordered if _r and _r.get("l2_paths"))
    _rows_l2_backfilled = sum(1 for _r in ordered if _r and _r.get("l2_backfilled"))
    print(f"  [Baseline Batch] questions={len(target_qs)} workers={BASELINE_N_WORKERS} "
          f"elapsed={batch_ms}ms avg={avg_ms}ms/q")
    print(
        "  [Baseline L2] "
        f"rows_with_l2={_rows_with_l2}/{len(target_qs)} "
        f"backfilled={_rows_l2_backfilled} "
        f"(enabled={BASELINE_L2_BACKFILL})"
    )

    success = 0
    for i, rec in enumerate(ordered, 1):
        if rec is None:
            continue

        ok = write_baseline_to_coverage(
            scene_id=SCENE_ID, frame_id=FRAME_ID,
            nuscenes_qa_id=rec["qa_uid"],    # e.g. "val_71051_comparison"
            question=rec["question"], answer=rec["answer"],
            l0_nodes=rec["l0_nodes"],
            l1_edges=rec["l1_edges"],
            l2_paths=rec["l2_paths"],        # 自动推导的 L2
            question_type=rec["qtype"],
            timestamp_start=rec["timestamp_start"], timestamp_end=rec["timestamp_end"],
            audit_cypher=rec["audit_cypher"],  # JSON 子图填入 cypher 列
            global_val_index=target_qs[i - 1][0],
        )
        if ok:
            success += 1
        if i <= 3 or i % 10 == 0:
            _ok_mark = "OK" if ok else "--"
            print(f"  [write {i:2d}/{len(target_qs)}] id={rec['qa_uid']:<25} "
                  f"q_type={rec['qtype']:12s} "
                  f"L0={len(rec['l0_nodes'])} "
                  f"L1={len(rec['l1_edges'])} "
                  f"L2={len(rec['l2_paths'])} "
                  f"{_ok_mark}")

    print(f"\n  Baseline audit done: {success}/{len(target_qs)} written to raw_coverage")
    return success


# ─────────────────────────────────────────────────────────────────────────────
# Step 5+6: Gap detection + generate 5 L2 questions -> question-answer-our
# ─────────────────────────────────────────────────────────────────────────────

# ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――
# V18 题型权重与答案策略（本地模板问句已移除）
# ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――

_Q_TYPE_WEIGHTS = {"exist": 3, "status": 3, "object": 3, "count": 1}


def _answer_for_q_type(q_type: str, cell: dict, verify_n: int = -1) -> str:
    """仅生成答案字段；问题文本必须由 LLM 生成。"""
    if q_type == "exist":
        return "yes" if verify_n != 0 else "no"
    if q_type == "status":
        return cell.get("n3_status", "") or "unknown"
    if q_type == "count":
        return str(verify_n if verify_n >= 0 else 1)
    if q_type == "comparison":
        return "closer"
    return cell.get("n3_id", "")


def step5_6_generate(driver, llm_client):
    """
    V18 真实性硬化协议：
      1) timestamp_start        : 进入每个 gap cell 处理第一行捕获
      2) timestamp_llm          : 来自 llm_client._call() 返回后的物理时刻
      3) timestamp_cypher_return: Neo4j single() 返回后的物理时刻
      4) timestamp_end          : 写 Excel 前最后一行捕获
    严格规则：
      - 自然语言问题默认由 LLM 生成（template 模式或批量异常时可回退模板）
      - 若 (timestamp_end - timestamp_start) < MIN_REAL_MS，则该条作废
      - cypher question / target_gap_cell 为空则作废
    """
    print("\n[Step 5+6] V18 Realness-Hardened Generation (4 physical timestamps locked)")
    import random
    import re
    import json as _json
    from datetime import datetime
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from gap_pipeline.coverage_tracker import CoverageTracker, CoverageRecord
    from gap_pipeline.constraint_methods import CumulativeConstraintChain, L2PivotConstraintChain
    from run_gap_pipeline_v6 import _process_single_cell
    from rq_tables import (
        write_generated_question,
        write_generated_questions_batch,
        make_generated_question_id,
    )
    from coverage_persistence import save_coverage_state, load_coverage_state
    from extract_candidates import expand_l0_with_candidates

    # V25 固定批量大小，批次数可变
    # 每批处理的gap数量（固定）
    GAPS_PER_BATCH = int(os.getenv("VQA_GAPS_PER_BATCH", "500"))
    # 并发参数
    Q_BATCH_SIZE = int(os.getenv("VQA_Q_BATCH_SIZE", "32"))
    Q_N_WORKERS = int(os.getenv("VQA_Q_N_WORKERS", "32"))
    CONTEXT_CYPHER_MODE = str(
        os.getenv("VQA_CONTEXT_CYPHER_MODE", "batch_llm") or "batch_llm"
    ).strip().lower()
    if CONTEXT_CYPHER_MODE not in ("batch_llm", "per_cell_llm", "fallback"):
        CONTEXT_CYPHER_MODE = "batch_llm"
    CTX_BATCH_CHUNK_SIZE = max(1, int(os.getenv("VQA_CTX_BATCH_CHUNK_SIZE", "8")))
    CTX_BATCH_N_WORKERS = max(
        1,
        int(
            os.getenv(
                "VQA_CTX_BATCH_N_WORKERS",
                str(max(1, min(Q_N_WORKERS, 16))),
            )
        ),
    )
    QUESTION_MODE = str(
        os.getenv("VQA_QUESTION_MODE", "template") or "template"
    ).strip().lower()
    if QUESTION_MODE == "llm":
        QUESTION_MODE = "llm_batch"
    if QUESTION_MODE not in ("llm_batch", "llm_strict", "template"):
        QUESTION_MODE = "llm_batch"
    Q_LLM_CHUNK_SIZE = max(1, int(os.getenv("VQA_Q_LLM_CHUNK_SIZE", "32")))
    EXCEL_BATCH_WRITE = os.getenv("VQA_EXCEL_BATCH_WRITE", "true").lower() in (
        "true",
        "1",
        "yes",
        "on",
    )
    # 真实性阈值（可通过 VQA_MIN_REAL_MS 调整；0=关闭）
    MIN_REAL_MS = max(0, int(os.getenv("VQA_MIN_REAL_MS", "0")))
    # V23 质量门控（可通过环境变量覆盖）
    STRICT_UNIQUE_ONLY = os.getenv("VQA_STRICT_UNIQUE_ONLY", "true").lower() in ("true", "1", "yes")
    MIN_ITER_COUNT = int(os.getenv("VQA_MIN_ITER_COUNT", "1"))
    MAX_QTYPE_RETRY = int(os.getenv("VQA_QTYPE_MAX_RETRY", "2"))
    AUTO_DOWNGRADE_NON_UNIQUE = os.getenv(
        "VQA_AUTODOWNGRADE_NON_UNIQUE", "false"
    ).lower() in ("true", "1", "yes")
    AUTO_DOWNGRADE_QTYPE = str(
        os.getenv("VQA_AUTODOWNGRADE_QTYPE", "exist") or "exist"
    ).strip().lower()
    if AUTO_DOWNGRADE_QTYPE not in ("exist", "count"):
        AUTO_DOWNGRADE_QTYPE = "exist"
    AUTO_DOWNGRADE_METHODS = {
        x.strip()
        for x in os.getenv("VQA_AUTODOWNGRADE_METHODS", "*").split(",")
        if x.strip()
    }
    LOW_ITER_STRICT_METHODS = {
        x.strip() for x in os.getenv("VQA_MIN_ITER_STRICT_METHODS", "path").split(",")
        if x.strip()
    }
    MAX_ROUNDS = int(os.getenv("VQA_MAX_ROUNDS", "0"))
    # 连续整批 EMPTY（本批 8 条全失败）时退出本帧 Gap，避免同一帧空转数千轮。0=不启用。
    MAX_EMPTY_STREAK = int(os.getenv("VQA_MAX_EMPTY_STREAK", "0"))
    # 同一 gap cell 连续失败次数超过阈值后本帧内暂时不选；0=关闭。与 EMPTY_STREAK 配合使用。
    GAP_FAIL_COOLDOWN = int(os.getenv("VQA_GAP_FAIL_COOLDOWN", "0"))

    from gap_pipeline.constraint_methods import (
        init_forensics as _init_f,
        _fw as _fwu,
        canonical_cell_key as _canonical_cell_key,
    )
    ENABLE_FORENSICS = os.getenv("VQA_ENABLE_FORENSICS", "true").lower() in ("true", "1", "yes", "on")
    if ENABLE_FORENSICS:
        _f_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        _f_dir = pathlib.Path(__file__).resolve().parent / "output"
        _f_path = str(_f_dir / f"forensics_{_f_ts}.jsonl")
        try:
            _f_dir.mkdir(parents=True, exist_ok=True)
            _init_f(_f_path)
            print(f"  [M0 Forensics] -> {_f_path}")
        except Exception as _f_exc:
            print(
                "  [WARN] Forensics disabled: "
                f"cannot initialize log file at {_f_path} ({_console_safe(_f_exc)})"
            )
            ENABLE_FORENSICS = False
            def _fwu(_rec):
                return None
    else:
        def _fwu(_rec):
            return None

    # ── 连接性预检（fail-fast）──────────────────────────────────────────────
    # Neo4j 连通性
    print("  [Pre-flight] Checking Neo4j connectivity...")
    try:
        with driver.session() as _pf_sess:
            _pf_r = _pf_sess.run("RETURN 1 AS ok").single()
            assert _pf_r is not None, "Neo4j returned None"
        print("  [Pre-flight] Neo4j OK")
    except Exception as _pf_exc:
        raise RuntimeError(
            f"[FATAL] Neo4j connectivity check failed: {_pf_exc}\n"
            "Please ensure Neo4j is running (neo4j console) before starting the pipeline."
        ) from _pf_exc

    # LLM API 连通性
    print("  [Pre-flight] Checking LLM API connectivity...")
    try:
        _pf_llm = llm_client
        _pf_test = _pf_llm._client.chat.completions.create(
            model=_pf_llm._model_audit,
            messages=[{"role": "user", "content": "Say OK"}],
            max_tokens=5,
            timeout=15,
        )
        assert _pf_test.choices, "LLM API returned empty choices"
        print(f"  [Pre-flight] LLM API OK (model={_pf_llm._model_audit})")
    except Exception as _pf_exc:
        raise RuntimeError(
            f"[FATAL] LLM API connectivity check failed: {_pf_exc}\n"
            "Please ensure VPN is connected and LLM server is reachable."
        ) from _pf_exc

    # V2 Pilot 约束方法优先级 P1–P16（与设计表一致）见 gap_pipeline/constraint_methods.py
    # 中「V2 Pilot 设计优先级」注释块；下方 key 多为实现侧 method_used / class.name。

    _CONSTRAINT_DESC_MAP = {
        "type_filter":         "target type unique in this direction",
        "status_anchor":       "target status unique in this direction",
        "type_status_anchor":  "type+status combo unique in this direction",
        "dir8_refine":         "paper 6-direction label narrows candidates to 1",
        "dual_reference":      "two reference objects jointly identify target",
        "dist_ord":            "target is closest/farthest in this direction",
        "two_hop_referent":    "a third node uniquely points to target",
        "dual_hop_referent":   "two third nodes jointly identify target",
        "anchor_intro":        "source object is uniquely described as anchor",
        "type+dir8":           "type + exact direction unique",
        "type+dir8+dist_ord":  "type + direction + distance rank unique",
        "dir8+dist_ord":       "direction + distance rank unique",
        "yesno_fallback":      "path itself uniquely identifies target",
    }

    def _abs_ts() -> str:
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    def _dt_ms(a: str, b: str) -> int:
        fmt = '%Y-%m-%d %H:%M:%S.%f'
        try:
            return int((datetime.strptime(b, fmt) - datetime.strptime(a, fmt)).total_seconds() * 1000)
        except Exception:
            return -1

    def _is_retryable_llm_exc(exc: Exception) -> bool:
        s = str(exc).lower()
        keys = ("timed out", "timeout", "connection", "connect", "readtimeout", "apiconnectionerror")
        return any(k in s for k in keys)
    def _is_auth_error(exc_or_msg: Exception | str) -> bool:
        s = str(exc_or_msg).lower()
        keys = (
            "authenticationerror",
            "unauthorized",
            "invalid token",
            "invalid api key",
            "无效的令牌",
            "401",
            "one_api_error",
        )
        return any(k in s for k in keys)

    def _parse_verify_n(verify_text: str) -> int:
        m = re.search(r"n=(\d+)", str(verify_text or ""))
        return int(m.group(1)) if m else -1

    def _parse_verify_ids(verify_text: str):
        """解析 logic_verification 里的 ids=[...]（与 run_gap_pipeline_v6._run_verify 格式一致）。"""
        t = str(verify_text or "")
        i = t.find("ids=")
        if i < 0:
            return None
        rest = t[i + 4 :].strip()
        if not rest.startswith("["):
            return None
        depth = 0
        end = 0
        for j, ch in enumerate(rest):
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    end = j + 1
                    break
        if end == 0:
            return None
        try:
            import ast

            v = ast.literal_eval(rest[:end])
            if isinstance(v, (list, tuple)):
                return [str(x) for x in v]
        except Exception:
            pass
        return None

    def _is_verify_wrong_target(verify_text: str, verify_n: int, n3_id: str) -> bool:
        """
        去重后只有 1 个目标 id，但该 id 不是 gap 的 n3（约束锁到同类另一个节点）。
        与「真·多候选 non-unique(多个不同对象)」区分，便于日志与 forensics。
        """
        tid = str(n3_id or "").strip()
        if not tid:
            return False
        ids = _parse_verify_ids(verify_text)
        if not ids:
            return False
        unique_ids = set(ids)
        # 只有去重后恰好 1 个目标、且不是 n3 时才算 wrong_target
        if len(unique_ids) == 1 and tid not in unique_ids:
            return True
        return False

    def _is_verified_unique(verify_text: str, verify_n: int) -> bool:
        t = str(verify_text or "").lower()
        if any(bad in t for bad in ("[fail]", "❌", "not unique", "n=0")):
            return False
            
        # V3 修复: 检查去重后的 ids
        ids = _parse_verify_ids(verify_text)
        if ids is not None:
            return len(set(ids)) == 1
            
        return verify_n == 1

    def _infer_q_type_from_question(question: str) -> str:
        """按问句字面归类题型，尽量减少 qtype 审计误杀。"""
        q = " ".join(str(question or "").strip().lower().split())
        if not q:
            return ""
        if q.startswith("how many ") or " number of " in q:
            return "count"
        if q.startswith(
            (
                "is ",
                "are ",
                "does ",
                "do ",
                "did ",
                "can ",
                "could ",
                "would ",
                "will ",
                "was ",
                "were ",
            )
        ):
            return "exist"
        if any(
            w in q
            for w in (
                "closest",
                "farthest",
                "closer",
                "farther",
                "further",
                "nearer",
                "same as",
                "different from",
                "compared to",
            )
        ):
            return "comparison"
        if "status" in q or "state" in q:
            return "status"
        if q.startswith("which "):
            return "status" if ("status" in q or "state" in q) else "object"
        if q.startswith("what "):
            return "status" if ("status" in q or "state" in q) else "object"
        return ""

    def _reconcile_q_type_eff(pool_eff: str, question: str) -> str:
        """将池子题型与问句字面做对齐，优先采用可判定的字面题型。"""
        inferred = _infer_q_type_from_question(question)
        if inferred:
            return inferred
        return pool_eff

    def _qtype_semantic_ok(q_type: str, question: str, method: str = "") -> bool:
        """语义审计：问题文本是否匹配 q_type；fallback 方法不做严格句式匹配。"""
        if method in ("yesno_fallback", "count_fallback", "emergency_fallback"):
            return True
        inferred = _infer_q_type_from_question(question)
        if not inferred:
            return False
        if inferred == q_type:
            return True
        # status/object 在真实问句里容易互相贴边（如“Which moving ...”）
        if {inferred, q_type} == {"status", "object"}:
            q = " ".join(str(question or "").strip().lower().split())
            if any(
                w in q
                for w in (
                    "moving",
                    "stopped",
                    "parked",
                    "standing",
                    "without_rider",
                    "without rider",
                    "with_rider",
                    "with rider",
                )
            ):
                return True
        return False

    def _should_auto_downgrade(method_name: str) -> bool:
        if not AUTO_DOWNGRADE_METHODS:
            return False
        m = str(method_name or "").strip()
        if "*" in AUTO_DOWNGRADE_METHODS:
            return True
        if m in AUTO_DOWNGRADE_METHODS:
            return True
        for p in AUTO_DOWNGRADE_METHODS:
            if p.endswith("*") and m.startswith(p[:-1]):
                return True
        return False

    def _scene_type_distribution() -> dict:
        q = (
            "MATCH (n:Object) "
            "RETURN n.type AS t, count(n) AS c "
            "ORDER BY c DESC LIMIT 12"
        )
        out = {}
        try:
            with driver.session() as sess:
                for rec in sess.run(q):
                    t = rec.get("t", "")
                    c = rec.get("c", 0)
                    if t:
                        out[str(t)] = int(c)
        except Exception:
            pass
        return out

    def _replay_raw_coverage_into_tracker(_tracker) -> dict:
        _enabled = os.getenv("VQA_TRACKER_LOAD_RAW_COVERAGE", "true").lower() in (
            "true",
            "1",
            "yes",
            "on",
        )
        _ret = {
            "enabled": _enabled,
            "rows": 0,
            "l0_touches": 0,
            "l1_touches": 0,
            "l2_touches": 0,
            "error": "",
        }
        if not _enabled:
            return _ret

        # V26: 优先从 CSV 读取（如果启用 CSV 模式）
        _use_csv = os.getenv("VQA_USE_CSV", "false").lower() in ("true", "1", "yes")
        if _use_csv:
            try:
                from csv_writer import _get_csv_baseline
                import csv

                csv_path = _get_csv_baseline()
                if not csv_path.exists():
                    _ret["error"] = f"csv-missing: {csv_path}"
                    return _ret

                def _parse_json_list(_v):
                    if _v is None or _v == "":
                        return []
                    try:
                        _x = _json.loads(_v)
                        return _x if isinstance(_x, list) else []
                    except Exception:
                        return []

                with open(csv_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if row.get('scene_id') != SCENE_ID or str(row.get('frame_id')) != str(FRAME_ID):
                            continue
                        _ret["rows"] += 1

                        # L0
                        _l0_items = _parse_json_list(row.get('l0_nodes', ''))
                        for _nd in _l0_items:
                            _nid = str(_nd or "").strip()
                            if not _nid:
                                continue
                            _tracker.record_from_qa({
                                "topology_level": "L0",
                                "path_pattern": _nid,
                                "template_id": "raw_coverage_replay",
                                "question_id": "raw_coverage_replay",
                            })
                            _ret["l0_touches"] += 1

                        # L1
                        _l1_items = _parse_json_list(row.get('l1_edges', ''))
                        for _it in _l1_items:
                            _s = _t = ""
                            if isinstance(_it, dict):
                                _s = str(_it.get("source") or _it.get("src") or _it.get("o1") or "").strip()
                                _t = str(_it.get("target") or _it.get("tgt") or _it.get("o2") or "").strip()
                            elif isinstance(_it, (list, tuple)) and len(_it) >= 2:
                                _s = str(_it[0] or "").strip()
                                _t = str(_it[1] or "").strip()
                            if not _s or not _t:
                                continue
                            _tracker.record_from_qa({
                                "topology_level": "L1",
                                "path_pattern": f"{_s}→{_t}",
                                "template_id": "raw_coverage_replay",
                                "question_id": "raw_coverage_replay",
                            })
                            _ret["l1_touches"] += 1

                        # L2
                        _l2_items = _parse_json_list(row.get('l2_paths', ''))
                        for _it in _l2_items:
                            _n1 = _n2 = _n3 = ""
                            if isinstance(_it, dict):
                                _n1 = str(_it.get("o1") or "").strip()
                                _n2 = str(_it.get("o2") or "").strip()
                                _n3 = str(_it.get("o3") or "").strip()
                            elif isinstance(_it, (list, tuple)) and len(_it) >= 3:
                                _n1 = str(_it[0] or "").strip()
                                _n2 = str(_it[1] or "").strip()
                                _n3 = str(_it[2] or "").strip()
                            if not _n1 or not _n2 or not _n3:
                                continue
                            _topo = "L2"
                            # 同时记录 legacy 链式和 pivot 枢纽式格式
                            _tracker.record_from_qa({
                                "topology_level": _topo,
                                "path_pattern": f"{_n1}|{_n2}|{_n3}",
                                "template_id": "raw_coverage_replay",
                                "question_id": "raw_coverage_replay",
                            })
                            _ret["l2_touches"] += 1

                return _ret

            except Exception as _exc:
                _ret["error"] = f"csv-read-failed: {_exc}"
                return _ret

        # 原有 Excel 读取逻辑
        try:
            import openpyxl
        except Exception as _exc:
            _ret["error"] = f"openpyxl-unavailable: {_exc}"
            return _ret

        if not EXCEL_PATH.is_file():
            _ret["error"] = f"excel-missing: {EXCEL_PATH}"
            return _ret

        def _norm(_v) -> str:
            return str(_v or "").strip().lower()

        def _parse_json_list(_v):
            if _v is None:
                return []
            if isinstance(_v, list):
                return _v
            _s = str(_v).strip()
            if not _s:
                return []
            try:
                _x = _json.loads(_s)
                return _x if isinstance(_x, list) else []
            except Exception:
                return []

        try:
            _wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
            try:
                if "raw_coverage" not in _wb.sheetnames:
                    _ret["error"] = "sheet-missing:raw_coverage"
                    return _ret
                _ws = _wb["raw_coverage"]
                _hrow = next(_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not _hrow:
                    _ret["error"] = "header-missing:raw_coverage"
                    return _ret
                _hmap = {_norm(h): i for i, h in enumerate(_hrow) if str(h or "").strip()}
                _i_scene = _hmap.get("scene_id")
                _i_frame = _hmap.get("frame_id")
                _i_l0 = _hmap.get("l0")
                _i_l1 = _hmap.get("l1")
                _i_l2 = _hmap.get("l2")
                if _i_scene is None or _i_frame is None:
                    _ret["error"] = "columns-missing:scene_id/frame_id"
                    return _ret

                for _row in _ws.iter_rows(min_row=2, values_only=True):
                    _sv = _row[_i_scene] if _i_scene < len(_row) else None
                    _fv = _row[_i_frame] if _i_frame < len(_row) else None
                    if _sv is None or _fv is None:
                        continue
                    try:
                        _fvi = int(_fv)
                    except Exception:
                        continue
                    if str(_sv).strip() != SCENE_ID or _fvi != FRAME_ID:
                        continue
                    _ret["rows"] += 1

                    _l0_items = _parse_json_list(
                        _row[_i_l0] if (_i_l0 is not None and _i_l0 < len(_row)) else None
                    )
                    for _nd in _l0_items:
                        _nid = str(_nd or "").strip()
                        if not _nid:
                            continue
                        _tracker.record_from_qa(
                            {
                                "topology_level": "L0",
                                "path_pattern": _nid,
                                "template_id": "raw_coverage_replay",
                                "question_id": "raw_coverage_replay",
                            }
                        )
                        _ret["l0_touches"] += 1

                    _l1_items = _parse_json_list(
                        _row[_i_l1] if (_i_l1 is not None and _i_l1 < len(_row)) else None
                    )
                    for _it in _l1_items:
                        _s = _t = ""
                        if isinstance(_it, dict):
                            _s = str(
                                _it.get("source")
                                or _it.get("src")
                                or _it.get("o1")
                                or ""
                            ).strip()
                            _t = str(
                                _it.get("target")
                                or _it.get("tgt")
                                or _it.get("o2")
                                or ""
                            ).strip()
                        elif isinstance(_it, (list, tuple)) and len(_it) >= 2:
                            _s = str(_it[0] or "").strip()
                            _t = str(_it[1] or "").strip()
                        if not _s or not _t:
                            continue
                        _tracker.record_from_qa(
                            {
                                "topology_level": "L1",
                                "path_pattern": f"{_s}→{_t}",
                                "template_id": "raw_coverage_replay",
                                "question_id": "raw_coverage_replay",
                            }
                        )
                        _ret["l1_touches"] += 1

                    _l2_items = _parse_json_list(
                        _row[_i_l2] if (_i_l2 is not None and _i_l2 < len(_row)) else None
                    )
                    for _it in _l2_items:
                        _n1 = _n2 = _n3 = ""
                        if isinstance(_it, dict):
                            _n1 = str(_it.get("o1") or "").strip()
                            _n2 = str(_it.get("o2") or "").strip()
                            _n3 = str(_it.get("o3") or "").strip()
                        elif isinstance(_it, (list, tuple)) and len(_it) >= 3:
                            _n1 = str(_it[0] or "").strip()
                            _n2 = str(_it[1] or "").strip()
                            _n3 = str(_it[2] or "").strip()
                        if not _n1 or not _n2 or not _n3:
                            continue
                        _topo = "L2"
                        # 同时记录 legacy 链式和 pivot 枢纽式格式
                        _tracker.record_from_qa(
                            {
                                "topology_level": _topo,
                                "path_pattern": f"{_n1}|{_n2}|{_n3}",
                                "template_id": "raw_coverage_replay",
                                "question_id": "raw_coverage_replay",
                            }
                        )
                        _ret["l2_touches"] += 1
            finally:
                _wb.close()
        except Exception as _exc:
            _ret["error"] = str(_exc)
        return _ret

    tracker = CoverageTracker()

    # 尝试从持久化文件恢复覆盖状态
    coverage_file = f"coverage_state/{SCENE_ID}_frame{FRAME_ID}.json"
    if os.path.exists(coverage_file):
        load_coverage_state(tracker, coverage_file)
        print(f"  [Coverage] Restored from {coverage_file}")
    else:
        with driver.session() as sess:
            tracker.init_from_session(sess)
        print(f"  [Coverage] Initialized from Neo4j")

    _stats_universe_initial = tracker.stats()

    _scenes = _json.loads((TRAINVAL / "scene.json").read_text())
    _samples = _json.loads((TRAINVAL / "sample.json").read_text())
    _st2name = {s["token"]: s["name"] for s in _scenes}
    _scene_samps = sorted(
        [s for s in _samples if _st2name.get(s["scene_token"], "") == SCENE_ID],
        key=lambda s: s["timestamp"],
    )
    sample_token = (_scene_samps[FRAME_ID]["token"] if len(_scene_samps) > FRAME_ID else "")

    _load_vb = os.getenv("VQA_TRACKER_LOAD_VAL_BASELINE", "true").lower() in ("true", "1", "yes", "on")
    if _load_vb and sample_token and QA_PATH.is_file():
        _vb = tracker.load_nuscenes_qa_baseline(
            str(QA_PATH), SCENE_ID, sample_tokens={sample_token}
        )
        print(
            f"  Val baseline → tracker (sample={sample_token[:16]}...): "
            f"L0+={_vb.get('n_L0', 0)} L1+={_vb.get('n_L1', 0)} L2+={_vb.get('n_L2', 0)}"
        )
    elif _load_vb and not sample_token:
        print("  [WARN] Val baseline skip: no sample_token for this scene/frame")

    _rc = _replay_raw_coverage_into_tracker(tracker)
    if _rc.get("enabled"):
        if _rc.get("error"):
            print(f"  [WARN] raw_coverage replay skip: {_rc.get('error')}")
        else:
            print(
                "  raw_coverage replay → tracker "
                f"(scene={SCENE_ID}, frame={FRAME_ID}): rows={_rc.get('rows', 0)} "
                f"L0_touches={_rc.get('l0_touches', 0)} "
                f"L1_touches={_rc.get('l1_touches', 0)} "
                f"L2_touches={_rc.get('l2_touches', 0)}"
            )
    else:
        print("  raw_coverage replay disabled by VQA_TRACKER_LOAD_RAW_COVERAGE=false")

    _stats_seeded = tracker.stats()
    _seed_delta = {
        _lvl: int(_stats_seeded[_lvl]["covered"] - _stats_universe_initial[_lvl]["covered"])
        for _lvl in ("L0", "L1", "L2")
    }
    print(
        "  tracker seed delta (covered): "
        f"L0={_seed_delta['L0']} L1={_seed_delta['L1']} "
        f"L2={_seed_delta['L2']}"
    )

    try:
        tracker.dump_universe_snapshot(
            scene_id=SCENE_ID,
            frame_id=FRAME_ID,
            out_dir="output/coverage_snapshots",
        )
    except Exception as _snap_exc:
        print(
            "  [WARN] universe snapshot skipped: "
            f"{_console_safe(_snap_exc)}"
        )

    def _print_stats(label: str, s=None):
        if s is None:
            s = tracker.stats()
        print(f"\n  [{label}]")
        for lvl in ("L0", "L1", "L2"):
            v = s[lvl]
            print(f"    {lvl:<4}: {v['gap']:>4} gap / {v['total']:>4} total "
                  f"(covered={v['covered']}, rate={v['rate']:.1f}%)")

    def _print_gap_diag(reason: str):
        s = tracker.stats()
        print(f"\n  [Gap Diagnose] reason={reason}")
        for lvl in ("L0", "L1", "L2"):
            v = s[lvl]
            print(f"    {lvl:<4}: gap={v['gap']} / total={v['total']} (rate={v['rate']:.1f}%)")
        l1_samples = [c.get("path_pattern", "?") for c in tracker.get_gap_cells("L1", limit=8)]
        l0_samples = [c.get("path_pattern", "?") for c in tracker.get_gap_cells("L0", limit=8)]
        if l1_samples:
            print(f"    L1 gap sample: {', '.join(l1_samples[:5])}")
        if l0_samples:
            print(f"    L0 gap sample: {', '.join(l0_samples[:5])}")
        if (
            s["L2"]["gap"] == 0
            and (s["L1"]["gap"] > 0 or s["L0"]["gap"] > 0)
        ):
            print("    [Hint] L2 已清零但 L1/L0 仍有 gap，可优先检查 L2 枚举边界与路径空间截断。")

    _print_stats("Gap Stats — Universe Initial", _stats_universe_initial)
    _print_stats("Gap Stats — Initial (after replay)", _stats_seeded)

    print(f"  sample_token for JSON: {sample_token[:16]}...")
    scene_dist_global = _scene_type_distribution()
    print(f"  scene type distribution: {scene_dist_global}")
    _ctx_perf = (
        f" ctx_chunk={CTX_BATCH_CHUNK_SIZE} ctx_workers={CTX_BATCH_N_WORKERS}"
        if CONTEXT_CYPHER_MODE == "batch_llm"
        else ""
    )
    print(
        f"  Perf mode: context={CONTEXT_CYPHER_MODE} question={QUESTION_MODE} "
        f"workers={Q_N_WORKERS} batch={Q_BATCH_SIZE} min_real_ms={MIN_REAL_MS}{_ctx_perf}"
    )

    _tls = threading.local()
    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            from gap_pipeline.llm_client import LLMClient as _LLMClient
            _llm = _LLMClient()
            _tls.llm = _llm
        return _llm

    def _noun_plural(noun: str) -> str:
        n = str(noun or "object").strip() or "object"
        return n if n.endswith("s") else f"{n}s"

    # 初始化模板库（会话级缓存）
    _template_lib = get_template_library()
    _template_usage_count = collections.defaultdict(int)  # 跟踪模板使用频率，实现均衡选择

    def _template_question(_topology: str, _cell: dict, _qtype: str):
        """
        基于267模板库的问题生成，采用频率反馈均衡策略

        策略：
        1. 根据 coverage_level (L0/L1/L2) 和 question_type 筛选候选模板
        2. 使用 softmax 温度采样：优先选择使用次数少的模板，保持多样性
        3. 温度参数 T=2.0：既保证均衡性，又允许一定随机性

        Returns:
            (question_text, template_entry) 或 (question_text, None) 如果使用 fallback
        """
        # 映射 topology 到 coverage_level
        # L2 统一处理（不再区分L2A/L2B）
        if _topology == "L2":
            coverage_level = "L2"
        elif _topology == "L1":
            coverage_level = "L1"
        else:
            coverage_level = "L0"

        # 获取候选模板
        candidates = _template_lib.get_by_level_type(coverage_level, _qtype)

        if not candidates:
            # 降级：使用简单硬编码模板
            return _fallback_template(_cell, _qtype), None

        # 频率反馈均衡选择（Softmax with Temperature）
        # 使用次数越少的模板，被选中概率越高
        T = 2.0  # 温度参数：越大越均衡，越小越随机
        usage_counts = [_template_usage_count[t.template_id] for t in candidates]
        min_usage = min(usage_counts) if usage_counts else 0

        # 计算权重：使用次数少的权重高
        # weight = exp(-(usage - min_usage) / T)
        import math
        weights = [math.exp(-(cnt - min_usage) / T) for cnt in usage_counts]
        total_weight = sum(weights)

        # 加权随机选择
        r = random.random() * total_weight
        cumsum = 0
        selected_template = candidates[0]
        for i, w in enumerate(weights):
            cumsum += w
            if r <= cumsum:
                selected_template = candidates[i]
                break

        # 更新使用计数
        _template_usage_count[selected_template.template_id] += 1

        # 填充模板参数
        params = _extract_template_params(_cell, selected_template.required_params)

        try:
            question = selected_template.template.format(**params)
            return question, selected_template
        except KeyError as e:
            # 参数缺失，降级到简单模板
            print(f"[Template Warning] Missing param {e} for {selected_template.template_id}, fallback")
            return _fallback_template(_cell, _qtype), None

    def _extract_template_params(_cell: dict, required: list) -> dict:
        """从 cell 提取模板所需参数"""
        params = {}

        # 提取基础字段
        n1_id = str(_cell.get("n1_id", "ego") or "ego")
        n1_type = str(_cell.get("n1_type", "") or "")
        n2_id = str(_cell.get("n2_id", "") or "")
        n2_type = str(_cell.get("n2_type", "object") or "object")
        n3_id = str(_cell.get("n3_id", "") or "")
        n3_type = str(_cell.get("n3_type", "object") or "object")
        n3_status = str(_cell.get("n3_status", "") or "")
        r1_dir = str(_cell.get("r1_dir8") or _cell.get("r1_dir4") or "front")
        r2_dir = str(_cell.get("r2_dir8") or _cell.get("r2_dir4") or "front")

        # 构建标准参数映射
        # L0/L1 风格参数
        if "obj_id" in required:
            params["obj_id"] = n3_id if n3_id else f"{n3_type}1"
        if "ref_id" in required:
            params["ref_id"] = n2_id if n2_id else f"{n2_type}1"
        if "obj_type" in required:
            params["obj_type"] = n3_type
        if "type_plural" in required:
            params["type_plural"] = _noun_plural(n3_type)
        if "ref_type" in required:
            params["ref_type"] = n2_type
        if "status" in required:
            params["status"] = n3_status if n3_status else "moving"
        if "direction" in required:
            params["direction"] = r2_dir
        if "direction1" in required:
            params["direction1"] = r1_dir
        if "direction2" in required:
            params["direction2"] = r2_dir
        if "distance_threshold" in required:
            params["distance_threshold"] = "10"

        # L2 风格参数（三节点链式）
        # ref_id = n1 (起点), mid_id = n2 (中间), target_id = n3 (目标)
        # ref_type = n1_type, mid_type = n2_type, target_type = n3_type
        if "ref1_id" in required:
            params["ref1_id"] = n1_id if n1_id != "ego" else "ego"
        if "ref2_id" in required:
            params["ref2_id"] = n2_id if n2_id else f"{n2_type}1"
        if "mid_id" in required:
            params["mid_id"] = n2_id if n2_id else f"{n2_type}1"
        if "target_id" in required:
            params["target_id"] = n3_id if n3_id else f"{n3_type}1"
        if "mid_type" in required:
            params["mid_type"] = n2_type
        if "target_type" in required:
            params["target_type"] = n3_type
        if "target_status" in required:
            params["target_status"] = n3_status if n3_status else "moving"

        # L2 特殊：构建 anchor 描述（用于降级模板）
        if "anchor" in required or "anchor_desc" in required:
            src_label = (
                "ego"
                if n1_id == "ego"
                else (n1_id if n1_id.lower().startswith(n1_type.lower()) else f"{n1_type} {n1_id}")
            )
            anchor = f"the {n2_type} that is to the {r1_dir} of {src_label}"
            params["anchor"] = anchor
            params["anchor_desc"] = anchor

        # L2 枢纽式 (pivot a|b|c) 参数
        # 从 cell 中读取 pivot 字段（由 coverage_tracker._parse_pivot_key 生成）
        if "a_id" in required:
            params["a_id"] = str(_cell.get("a_id") or n1_id or "ego")
        if "b_id" in required:
            params["b_id"] = str(_cell.get("b_id") or n2_id or "")
        if "c_id" in required:
            params["c_id"] = str(_cell.get("c_id") or n3_id or "")
        if "a_type" in required:
            params["a_type"] = str(_cell.get("a_type") or n1_type or "object")
        if "b_type" in required:
            params["b_type"] = str(_cell.get("b_type") or n2_type or "object")
        if "c_type" in required:
            params["c_type"] = str(_cell.get("c_type") or n3_type or "object")
        if "a_type_plural" in required:
            params["a_type_plural"] = _noun_plural(str(_cell.get("a_type") or n1_type or "object"))
        if "b_type_plural" in required:
            params["b_type_plural"] = _noun_plural(str(_cell.get("b_type") or n2_type or "object"))
        # 方向参数（汇聚型 / 发散型）
        if "dir_ab" in required:
            params["dir_ab"] = str(_cell.get("dir_ab") or r1_dir or "front")
        if "dir_cb" in required:
            params["dir_cb"] = str(_cell.get("dir_cb") or r2_dir or "front")
        if "dir_ba" in required:
            params["dir_ba"] = str(_cell.get("dir_ba") or r1_dir or "front")
        if "dir_bc" in required:
            params["dir_bc"] = str(_cell.get("dir_bc") or r2_dir or "front")
        # 状态参数
        if "a_status" in required:
            params["a_status"] = str(_cell.get("a_status") or "")
        if "b_status" in required:
            params["b_status"] = str(_cell.get("b_status") or n3_status or "")
        if "c_status" in required:
            params["c_status"] = str(_cell.get("c_status") or "")

        return params

    def _fallback_template(_cell: dict, _qtype: str) -> str:
        """降级简单模板（保持向后兼容）"""
        n1_id = str(_cell.get("n1_id", "ego") or "ego")
        n1_type = str(_cell.get("n1_type", "") or "")
        n2_type = str(_cell.get("n2_type", "object") or "object")
        n3_type = str(_cell.get("n3_type", "object") or "object")
        n3_status = str(_cell.get("n3_status", "") or "")
        r1_dir = str(_cell.get("r1_dir8") or _cell.get("r1_dir4") or "front")
        r2_dir = str(_cell.get("r2_dir8") or _cell.get("r2_dir4") or "front")
        src_label = (
            "ego"
            if n1_id == "ego"
            else (n1_id if n1_id.lower().startswith(n1_type.lower()) else f"{n1_type} {n1_id}")
        )
        anchor = f"the {n2_type} that is to the {r1_dir} of {src_label}"
        if _qtype == "exist":
            st = f"{n3_status} " if n3_status else ""
            return f"Is there a {st}{n3_type} to the {r2_dir} of {anchor}?"
        if _qtype == "count":
            return f"How many {_noun_plural(n3_type)} are to the {r2_dir} of {anchor}?"
        if _qtype == "status":
            return f"What is the status of the {n3_type} to the {r2_dir} of {anchor}?"
        if _qtype == "comparison":
            return f"Which is closer to {anchor}, the {n3_type} at {r2_dir} or another one?"
        return f"What {n3_type} is to the {r2_dir} of {anchor}?"

    def _gap_score(cell):
        """
        改进的gap评分函数：优先覆盖L0和L1

        策略：
        - 阶段1（L0<100% 或 L1<80%）：大幅提升包含未覆盖L0/L1的gap评分
        - 阶段2（L0=100%, L1<100%）：平衡覆盖
        - 阶段3（L0=100%, L1=100%）：L2收尾
        """
        # 获取当前覆盖率
        stats = tracker.stats()
        l0_rate = stats["L0"]["covered"] / max(1, stats["L0"]["total"])
        l1_rate = stats["L1"]["covered"] / max(1, stats["L1"]["total"])

        # 基础分：未覆盖的L0节点
        nodes = [cell.get("n1_id",""), cell.get("n2_id",""), cell.get("n3_id","")]
        unc_l0 = sum(1 for n in nodes if n and tracker._L0.get(n, CoverageRecord()).hit_count == 0)

        # L1边覆盖情况
        unc_l1 = 0
        # 检查 n1->n2 边
        if cell.get("n1_id") and cell.get("n2_id"):
            edge_key = f"{cell['n1_id']}→{cell['n2_id']}"
            if tracker._L1.get(edge_key, CoverageRecord()).hit_count == 0:
                unc_l1 += 1
        # 检查 n2->n3 边（如果是L2B）
        if cell.get("n2_id") and cell.get("n3_id"):
            edge_key = f"{cell['n2_id']}→{cell['n3_id']}"
            if tracker._L1.get(edge_key, CoverageRecord()).hit_count == 0:
                unc_l1 += 1

        # ego惩罚（降低ego路径优先级）
        path = cell.get("path_pattern", "")
        ego_penalty = 0.5 if "ego" in path else 1.0

        # 阶段权重
        if l0_rate < 1.0 or l1_rate < 0.8:
            # 阶段1：快速覆盖L0/L1（L0权重=100, L1权重=50）
            score = 100 * unc_l0 + 50 * unc_l1 + 1
        elif l1_rate < 1.0:
            # 阶段2：平衡覆盖（L0权重=50, L1权重=20）
            score = 50 * unc_l0 + 20 * unc_l1 + 1
        else:
            # 阶段3：L2收尾（L0权重=10, L1权重=5）
            score = 10 * unc_l0 + 5 * unc_l1 + 1

        return ego_penalty * score

    def _chunked(_items: list, _size: int) -> list:
        _size = max(1, int(_size))
        return [_items[i : i + _size] for i in range(0, len(_items), _size)]

    q_type_pool = []
    for qt, wt in _Q_TYPE_WEIGHTS.items():
        q_type_pool.extend([qt] * wt)

    generated = 0
    total_att = 0
    dropped_fast = 0
    dropped_empty = 0
    dropped_non_unique = 0
    dropped_wrong_verify_target = 0
    dropped_low_iter = 0
    dropped_qtype_mismatch = 0
    fail_counts: dict = {}
    used_types: list = []
    nusqa_records: list = []
    method_stage_counter: collections.Counter = collections.Counter()
    method_total_counter: collections.Counter = collections.Counter()
    method_keep_counter: collections.Counter = collections.Counter()
    auto_downgraded_total = 0
    auto_downgraded_reason_counter: collections.Counter = collections.Counter()

    def _all_covered() -> bool:
        s = tracker.stats()
        return all(s[lvl]["gap"] == 0 for lvl in ("L0", "L1", "L2"))

    round_idx = 0
    empty_streak = 0
    gap_fail_counts: dict = {}
    while True:
        if _all_covered():
            print(f"\n  [OK] Round {round_idx}: All coverage levels reached 100%.")
            _print_gap_diag("all_covered")
            break
        if MAX_ROUNDS > 0 and round_idx >= MAX_ROUNDS:
            print(f"\n  [M0] Reached MAX_ROUNDS={MAX_ROUNDS}, stopping for forensics.")
            _print_gap_diag("max_rounds")
            break
        round_idx += 1
        raw_l2 = list(tracker.get_gap_cells("L2"))

        def _pass_cooldown(c):
            if GAP_FAIL_COOLDOWN <= 0:
                return True
            return gap_fail_counts.get(c.get("_key", ""), 0) < GAP_FAIL_COOLDOWN

        l2_avail = [c for c in raw_l2 if _pass_cooldown(c)]
        if (
            GAP_FAIL_COOLDOWN > 0
            and not l2_avail
            and raw_l2
        ):
            print(
                f"\n  [INFO] 全部 gap cell 处于失败冷却 (≥{GAP_FAIL_COOLDOWN} 次失败), "
                f"重置本帧失败计数后继续"
            )
            gap_fail_counts.clear()
            l2_avail = raw_l2

        s_l2 = sorted(l2_avail, key=_gap_score, reverse=True)

        # V25: 固定每批gap数量，批次数可变
        total_available = len(s_l2)
        batch_size = min(GAPS_PER_BATCH, total_available)

        round_cells = [("L2", c) for c in s_l2[:batch_size]]
        if not round_cells:
            print(
                f"\n  [WARN] No selectable gap cells while coverage <100% "
                f"(scene={SCENE_ID}, frame={FRAME_ID}). Stopping this frame."
            )
            _print_gap_diag("no_selectable_l2_cells")
            break

        print(f"\n  [Round {round_idx}] batch={len(round_cells)} "
              f"(L2={len(round_cells)}) "
              f"written={generated}")

        # 批次时间分析
        batch_start_time = time.perf_counter()
        gap_selection_ms = 0.0
        context_build_ms = 0.0

        batch_llm_ms = []
        batch_neo_ms = []
        batch_total_ms = []
        round_batch_id = f"{SCENE_ID}_f{FRAME_ID}_r{round_idx}"

        # 先分配q_type，避免线程中竞争
        assigned_cells = []
        for topology, cell in round_cells:
            total_att += 1
            rem = [t for t in q_type_pool if t not in used_types]
            if not rem:
                rem = list(_Q_TYPE_WEIGHTS.keys())
                used_types = []
            q_type = random.choice(rem)
            used_types.append(q_type)
            assigned_cells.append((topology, cell, q_type))

        prebuilt_context_by_key: dict = {}
        if assigned_cells and CONTEXT_CYPHER_MODE in ("batch_llm", "fallback"):
            _ctx_t0 = time.perf_counter()
            if CONTEXT_CYPHER_MODE == "fallback":
                from gap_pipeline.llm_client import LLMClient as _LLMClient
                for _t, _c, _ in assigned_cells:
                    _k = str(_c.get("_key", "") or "")
                    if not _k:
                        continue
                    _ts = _abs_ts()
                    _cy = _LLMClient.build_l2a_fallback_cypher(_c)
                    prebuilt_context_by_key[_k] = {
                        "cypher": str(_cy or ""),
                        "ts_start": _ts,
                        "ts_llm": _ts,
                        "llm_ms": 0.0,
                    }
            else:
                from gap_pipeline.llm_client import LLMClient as _LLMClient
                _l2_cells = [_c for _t, _c, _ in assigned_cells if _t == "L2"]
                _ctx_jobs = []
                _ctx_jobs.extend(
                    [("L2", _chunk) for _chunk in _chunked(_l2_cells, CTX_BATCH_CHUNK_SIZE)]
                )

                def _build_ctx_chunk(_topology: str, _chunk_cells: list) -> dict:
                    _ts_start = _abs_ts()
                    _t0 = time.perf_counter()
                    _llm_local = _get_worker_llm()
                    try:
                        _cy_list = _llm_local.generate_context_cypher_batch(
                            _chunk_cells, topology=_topology
                        )
                        _ts_llm = _llm_local.last_call_meta.get("timestamp_llm") or _abs_ts()
                        _total_ms = max(0.0, (time.perf_counter() - _t0) * 1000.0)
                    except Exception as _ctx_exc:
                        print(
                            f"    [WARN] [ctx-batch-exc] topo={_topology} size={len(_chunk_cells)} "
                            f"{_console_safe(_ctx_exc)}"
                        )
                        _cy_list = []
                        _ts_llm = _abs_ts()
                        _total_ms = 0.0
                    _ms_per = _total_ms / max(1, len(_chunk_cells))
                    _items = {}
                    for _idx, _cell_i in enumerate(_chunk_cells):
                        _k = str(_cell_i.get("_key", "") or "")
                        if not _k:
                            continue
                        _cy = str(_cy_list[_idx] or "") if _idx < len(_cy_list) else ""
                        if not _cy:
                            _cy = _LLMClient.build_l2a_fallback_cypher(_cell_i)
                        _items[_k] = {
                            "cypher": _cy,
                            "ts_start": _ts_start,
                            "ts_llm": _ts_llm,
                            "llm_ms": _ms_per,
                        }
                    return _items

                if _ctx_jobs:
                    _ctx_workers = min(CTX_BATCH_N_WORKERS, len(_ctx_jobs))
                    with ThreadPoolExecutor(max_workers=_ctx_workers) as _ctx_ex:
                        _ctx_futs = [
                            _ctx_ex.submit(_build_ctx_chunk, _topo, _chunk)
                            for _topo, _chunk in _ctx_jobs
                        ]
                        for _ctx_f in as_completed(_ctx_futs):
                            try:
                                _ctx_items = _ctx_f.result()
                            except Exception as _ctx_exc:
                                print(f"    [WARN] [ctx-chunk-fail] {_console_safe(_ctx_exc)}")
                                continue
                            if _ctx_items:
                                prebuilt_context_by_key.update(_ctx_items)
            _ctx_ms = int((time.perf_counter() - _ctx_t0) * 1000)
            context_build_ms = _ctx_ms
            print(
                f"  [Context Build] mode={CONTEXT_CYPHER_MODE} "
                f"cells={len(assigned_cells)} ready={len(prebuilt_context_by_key)} elapsed={_ctx_ms}ms"
            )

        def _cell_worker(_topology, _cell, _q_type):
            _llm = _get_worker_llm()
            _gap_key = _cell["_key"]
            _path = _cell.get("path_pattern", "?")
            _cell_key = _canonical_cell_key(
                _cell.get("n1_id", "?"),
                _cell.get("n2_id", "?"),
                _cell.get("n3_id", "?"),
            )

            # [物理采样点1]
            _ts_start = _abs_ts()
            _cypher = ""
            _ts_llm = _ts_start
            _llm_ms = 0.0

            _pre_ctx = prebuilt_context_by_key.get(_gap_key) if _gap_key else None
            if isinstance(_pre_ctx, dict):
                _cypher = str(_pre_ctx.get("cypher", "") or "")
                if _cypher:
                    _ts_start = str(_pre_ctx.get("ts_start") or _ts_start)
                    _ts_llm = str(_pre_ctx.get("ts_llm") or _abs_ts())
                    try:
                        _llm_ms = float(_pre_ctx.get("llm_ms", 0.0) or 0.0)
                    except Exception:
                        _llm_ms = 0.0

            if not _cypher:
                # LLM context (按模式回退)
                if CONTEXT_CYPHER_MODE == "fallback":
                    from gap_pipeline.llm_client import LLMClient as _LLMClient
                    _cypher = _LLMClient.build_l2a_fallback_cypher(_cell)
                    _ts_llm = _abs_ts()
                    _llm_ms = 0.0
                else:
                    _cypher = _llm.generate_l2a_context_cypher(_cell)
                    # [物理采样点2]
                    _ts_llm = _llm.last_call_meta.get("timestamp_llm") or _abs_ts()
                    _llm_ms = float(_llm.last_call_timing.get("total_ms", 0.0))

            _local_chain = CumulativeConstraintChain()
            _qa, _timing = _process_single_cell(
                cell=_cell, topology=_topology, cypher=_cypher,
                driver=driver, chain=_local_chain,
                scene_name=SCENE_ID, frame_idx=FRAME_ID,
                llm_timing=dict(_llm.last_call_timing),
                render_local_question=False,
            )
            _ts_cypher = _timing.get("timestamp_cypher_return", "") or _abs_ts()
            _neo_ms = float(_timing.get("neo4j_ms", 0.0))
            if _qa is None:
                return {"ok": False, "stage": "neo4j", "gap_key": _gap_key, "path": _path}

            _method = _timing.get("method_used", "") or "path"
            _constraint_desc = _CONSTRAINT_DESC_MAP.get(
                _method, f"constraint chain method={_method} identifies target"
            )
            # 约束链兜底方法映射到真实题型（否则上层会把 fallback 结果当成普通 object/status）
            if _method == "count_fallback":
                _q_type_eff = "count"
            elif _method in ("yesno_fallback", "emergency_fallback"):
                _q_type_eff = "exist"
            else:
                _q_type_eff = _q_type

            _verify_cypher = _timing.get("verify_cypher", "") or ""
            if not _verify_cypher.strip() or not _path.strip():
                return {"ok": False, "stage": "empty", "gap_key": _gap_key, "path": _path}
            _verify_text = str(_timing.get("logic_verification", "") or "")
            _verify_n = _parse_verify_n(_verify_text)
            _verify_ids = _parse_verify_ids(_verify_text)
            _verified_unique = _is_verified_unique(_verify_text, _verify_n)
            
            _chain_unique = bool(_timing.get("is_unique", False))
            _target_n3 = str(_cell.get("n3_id") or "").strip()
            _wrong_verify_target = _is_verify_wrong_target(
                _verify_text, _verify_n, _target_n3
            )
            
            # _chain_unique check uses the exact logic as well.
            # is_unique flag in timing is usually true if length of unique targets is 1
            _is_non_unique = not _verified_unique
            
            _fwu({"stage": "upper", "cell_key": _cell_key, "method_name": _method, "method_used": _method, "chain_is_unique": _chain_unique, "verified_unique": _verified_unique, "verify_n": _verify_n, "verify_ids": _verify_ids, "target_n3": _target_n3, "wrong_verify_target": _wrong_verify_target, "verify_cypher": _verify_cypher, "is_non_unique": _is_non_unique, "q_type_eff": _q_type_eff})
            _autodowngraded = False
            _autodowngrade_reason = ""
            _orig_q_type_eff = _q_type_eff
            if STRICT_UNIQUE_ONLY and _q_type_eff not in ("count", "exist"):
                _must_gate = (_wrong_verify_target or _is_non_unique)
                if (
                    _must_gate
                    and AUTO_DOWNGRADE_NON_UNIQUE
                    and _should_auto_downgrade(_method)
                ):
                    _q_type_eff = AUTO_DOWNGRADE_QTYPE
                    _autodowngraded = True
                    _autodowngrade_reason = (
                        "wrong-target" if _wrong_verify_target else "non-unique"
                    )
                    _fwu({
                        "stage": "verdict",
                        "cell_key": _cell_key,
                        "final_verdict": f"downgraded-to-{_q_type_eff}",
                        "reason": _autodowngrade_reason,
                        "verify_n": _verify_n,
                        "q_type_before": _orig_q_type_eff,
                        "q_type_after": _q_type_eff,
                        "method_name": _method,
                        "method_used": _method,
                    })
                elif _wrong_verify_target:
                    _fwu({
                        "stage": "verdict",
                        "cell_key": _cell_key,
                        "final_verdict": "drop-wrong-target",
                        "verify_n": _verify_n,
                        "target_n3": _target_n3,
                        "verify_ids": _verify_ids,
                        "method_name": _method,
                        "method_used": _method,
                    })
                    return {
                        "ok": False,
                        "stage": "wrong_verify_target",
                        "gap_key": _gap_key,
                        "path": _path,
                        "verify": _verify_text,
                        "method": _method,
                        "target_n3": _target_n3,
                        "verify_ids": _verify_ids,
                        "verify_n": _verify_n,
                    }
                elif _is_non_unique:
                    _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "drop-non-unique", "verify_n": _verify_n, "method_name": _method, "method_used": _method})
                    return {
                        "ok": False,
                        "stage": "non_unique",
                        "gap_key": _gap_key,
                        "path": _path,
                        "verify": _verify_text,
                        "method": _method,
                    }

            _work_qt = _q_type_eff
            _answer = _answer_for_q_type(_work_qt, _cell, verify_n=_verify_n)

            def _gen_q_once(qt: str, answer_text: str) -> str:
                try:
                    return _llm.generate_question_nlp_strict(
                        path=_path,
                        q_type=qt,
                        n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                        n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                        n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                        n3_status=_cell.get("n3_status",""),
                        r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                        r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                        constraint_desc=_constraint_desc,
                        answer=str(answer_text),
                        scene_distribution=scene_dist_global,
                    )
                except Exception as _q_exc:
                    if _is_retryable_llm_exc(_q_exc):
                        time.sleep(0.8)  # 单次快速退避
                        return _llm.generate_question_nlp_strict(
                            path=_path,
                            q_type=qt,
                            n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                            n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                            n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                            n3_status=_cell.get("n3_status",""),
                            r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                            r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                            constraint_desc=_constraint_desc,
                            answer=str(answer_text),
                            scene_distribution=scene_dist_global,
                        )
                    raise

            _trace = _timing.get("constraint_trace", "")
            # V24: 优先使用管线内物理计数（CumulativeConstraintChain.trace_log 长度）
            _iter_count = int(_timing.get("constraint_rounds") or 0)
            if _iter_count < 1:
                _parts = [p.strip() for p in _trace.replace("->", "->").split("->")]
                _iter_count = max(1, len([p for p in _parts if p and p != "Path"]))
            if _iter_count < MIN_ITER_COUNT and _method in LOW_ITER_STRICT_METHODS:
                _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "drop-low-iter", "iteration_count": _iter_count, "method_name": _method, "method_used": _method})
                return {
                    "ok": False,
                    "stage": "low_iter",
                    "gap_key": _gap_key,
                    "path": _path,
                    "method": _method,
                    "iteration_count": _iter_count,
                    "trace": _trace,
                }
            _n1 = _cell.get("n1_id",""); _n2 = _cell.get("n2_id",""); _n3 = _cell.get("n3_id","")
            _l0_base = [x for x in [_n1, _n2, _n3] if x]
            _l0 = expand_l0_with_candidates(_l0_base, _verify_cypher, _verify_text)
            _l1 = ([{"source": _n1, "target": _n2}] if _n1 and _n2 else []) + \
                  ([{"source": _n2, "target": _n3}] if _n2 and _n3 else [])
            _l2 = [{"o1": _n1, "o2": _n2, "o3": _n3}] if all([_n1, _n2, _n3]) else []
            _gen_id = make_generated_question_id(SCENE_ID, FRAME_ID)

            if QUESTION_MODE == "llm_batch":
                return {
                    "ok": True,
                    "need_question_batch": True,
                    "gen_id": _gen_id,
                    "gap_key": _gap_key,
                    "path": _path,
                    "topology": _topology,
                    "q_type": _work_qt,
                    "answer": str(_answer),
                    "verify_cypher": _verify_cypher,
                    "verify_text": _verify_text,
                    "verify_n": _verify_n,
                    "trace": _trace,
                    "iteration_count": _iter_count,
                    "method": _method,
                    "autodowngraded": _autodowngraded,
                    "autodowngrade_reason": _autodowngrade_reason,
                    "is_unique": _timing.get("is_unique", False),
                    "ts_start": _ts_start,
                    "ts_llm": _ts_llm,
                    "ts_cypher_return": _ts_cypher,
                    "llm_ms": _llm_ms,
                    "neo_ms": _neo_ms,
                    "l0": _l0, "l1": _l1, "l2": _l2,
                    "cell": _cell,
                    "constraint_desc": _constraint_desc,
                    "cell_key": _cell_key,
                }

            _question = ""
            _semantic_ok = False
            _answer_logic = ""
            _template_entry = None
            if QUESTION_MODE == "template":
                _question, _template_entry = _template_question(_topology, _cell, _work_qt)
                if _template_entry:
                    _answer_logic = _template_entry.answer_logic
                    # 根据 answer_logic 增强 Cypher 查询
                    if _answer_logic:
                        from gap_pipeline.answer_logic_cypher import augment_cypher_with_answer_logic
                        _template_params = _extract_template_params(_cell, _template_entry.required_params)
                        _verify_cypher = augment_cypher_with_answer_logic(
                            _verify_cypher, _answer_logic, _cell, _template_params
                        )
                _q_type_eff = _work_qt
                _semantic_ok = True
            else:
                for _ in range(max(1, MAX_QTYPE_RETRY)):
                    _answer = _answer_for_q_type(_work_qt, _cell, verify_n=_verify_n)
                    _question = _gen_q_once(_work_qt, _answer)
                    _eff = _reconcile_q_type_eff(_work_qt, _question)
                    if _qtype_semantic_ok(_eff, _question, method=_method):
                        _q_type_eff = _eff
                        _semantic_ok = True
                        break
                    _work_qt = _eff
            if _semantic_ok:
                _answer = _answer_for_q_type(_q_type_eff, _cell, verify_n=_verify_n)
            if not _semantic_ok:
                _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "drop-qtype", "q_type_eff": _q_type_eff, "rendered_question": _question, "method_name": _method, "method_used": _method})
                return {
                    "ok": False,
                    "stage": "qtype_mismatch",
                    "gap_key": _gap_key,
                    "path": _path,
                    "method": _method,
                    "q_type": _q_type_eff,
                    "question": _question,
                }

            # [物理采样点4]
            _ts_end = _abs_ts()
            _total_ms = _dt_ms(_ts_start, _ts_end)
            if _total_ms >= 0 and _total_ms < MIN_REAL_MS:
                return {"ok": False, "stage": "fast", "gap_key": _gap_key, "path": _path, "total_ms": _total_ms}

            _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "kept", "verify_n": _verify_n, "method_name": _method, "method_used": _method, "rendered_question": _question, "q_type_eff": _q_type_eff})
            return {
                "ok": True,
                "gen_id": _gen_id,
                "gap_key": _gap_key,
                "path": _path,
                "topology": _topology,
                "q_type": _q_type_eff,
                "answer": str(_answer),
                "question": _question,
                "verify_cypher": _verify_cypher,
                "verify_text": _verify_text,
                "verify_n": _verify_n,
                "trace": _trace,
                "iteration_count": _iter_count,
                "method": _method,
                "autodowngraded": _autodowngraded,
                "autodowngrade_reason": _autodowngrade_reason,
                "is_unique": _timing.get("is_unique", False),
                "ts_start": _ts_start,
                "ts_llm": _ts_llm,
                "ts_cypher_return": _ts_cypher,
                "ts_end": _ts_end,
                "total_ms": _total_ms,
                "llm_ms": _llm_ms,
                "neo_ms": _neo_ms,
                "l0": _l0, "l1": _l1, "l2": _l2,
                "answer_logic": _answer_logic,  # 新增：传递 answer_logic
            }

        pending_question_batch: list = []
        def _to_sheet_b_payload(_res: dict) -> dict:
            return {
                "scene_id": SCENE_ID,
                "frame_id": FRAME_ID,
                "question_id": _res["gen_id"],
                "timestamp_start": _res["ts_start"],
                "timestamp_llm": _res["ts_llm"],
                "timestamp_cypher_return": _res["ts_cypher_return"],
                "timestamp_end": _res["ts_end"],
                "iteration_count": _res["iteration_count"],
                "question_type": _res["q_type"],
                "complexity": "L2",
                "natural language question": _res["question"],
                "cypher question": _res["verify_cypher"],
                "answer": _res["answer"],
                "L0": _res["l0"],
                "L1": _res["l1"],
                "L2": _res["l2"],
                "gap_cell": _res["path"],
                "target_gap_cell": _res["path"],
                "batch_id": round_batch_id,
                "Batch_ID": round_batch_id,
            }

        def _record_kept(_res: dict, _write_ok: bool | None = None):
            nonlocal generated, auto_downgraded_total
            if GAP_FAIL_COOLDOWN > 0 and _res.get("gap_key"):
                gap_fail_counts.pop(_res["gap_key"], None)
            _method_ok = _res.get("method", "?")
            method_stage_counter[(_method_ok, "kept")] += 1
            method_total_counter[_method_ok] += 1
            method_keep_counter[_method_ok] += 1
            if _res.get("autodowngraded"):
                auto_downgraded_total += 1
                auto_downgraded_reason_counter[
                    str(_res.get("autodowngrade_reason") or "unknown")
                ] += 1

            if _write_ok is None:
                ok = write_generated_question(
                    scene_id=SCENE_ID, frame_id=FRAME_ID, question_id=_res["gen_id"],
                    timestamp_start=_res["ts_start"],
                    timestamp_llm=_res["ts_llm"],
                    timestamp_cypher_return=_res["ts_cypher_return"],
                    timestamp_end=_res["ts_end"],
                    iteration_count=_res["iteration_count"],
                    question_type=_res["q_type"],
                    complexity="L2",
                    natural_language_question=_res["question"],
                    cypher_question=_res["verify_cypher"],
                    answer=_res["answer"],
                    l0_nodes=_res["l0"], l1_edges=_res["l1"], l2_paths=_res["l2"],
                    target_gap_cell=_res["path"],
                    batch_id=round_batch_id,
                )
            else:
                ok = bool(_write_ok)
            if ok:
                generated += 1
            nusqa_records.append({
                "sample_token": sample_token,
                "question": _res["question"],
                "answer": _res["answer"],
                "template_type": _res["q_type"],
                "num_hop": 2,
                "question_id": _res["gen_id"],
                "scene_name": SCENE_ID,
                "frame_idx": FRAME_ID,
                "topology_level": _res["topology"],
                "path_pattern": _res["path"],
                "constraint_trace": _res["trace"],
                "iteration_count": _res["iteration_count"],
                "is_unique": _res["is_unique"],
                "method_used": _res["method"],
                "verify_cypher": _res["verify_cypher"],
                "verify_result": _res.get("verify_text", ""),
                "verify_n": _res.get("verify_n", -1),
                "timestamp_start": _res["ts_start"],
                "timestamp_llm": _res["ts_llm"],
                "timestamp_cypher_return": _res["ts_cypher_return"],
                "timestamp_end": _res["ts_end"],
                "total_ms": _res["total_ms"],
                "batch_id": round_batch_id,
                "answer_logic": _res.get("answer_logic", ""),  # 新增：传递 answer_logic
            })
            # 覆盖率记录：只有 verify 确认目标匹配时才标记 covered
            _vn = _res.get("verify_n", -1)
            _vt = _res.get("verify_text", "")
            _qt_res = _res.get("q_type", "")
            _target = str(_res.get("path", "")).split("|")[-1] if "|" in str(_res.get("path", "")) else ""
            # 严格覆盖记录：所有题型都必须由 verify 明确返回目标节点。
            # 旧逻辑把 exist/count 的非唯一结果也记为 covered，会造成“覆盖率虚高但题不准”。
            _verify_ok = (
                _vn >= 1
                and (not _target or _target in str(_vt))
            )
            if _verify_ok:
                tracker.record_from_qa({
                    "topology_level": _res["topology"],
                    "path_pattern": _res["path"],
                    "template_id": f"v19_{_res['q_type']}_{_res['method']}",
                    "question_id": _res["gen_id"],
                })
            else:
                print(
                    f"    [WARN] Coverage NOT recorded for {_res.get('path', '?')}: "
                    f"verify_n={_vn} target={_target} not confirmed"
                )
            batch_llm_ms.append(_res["llm_ms"])
            batch_neo_ms.append(_res["neo_ms"])
            batch_total_ms.append(_res["total_ms"] if _res["total_ms"] >= 0 else 0.0)
            print(f"    #{generated:03d} [{_res['topology']}] iter={_res['iteration_count']} "
                  f"q={_res['q_type']:<10} v={_res.get('verify_n',-1)} "
                  f"dt={_res['total_ms']}ms batch={round_batch_id}")

        with ThreadPoolExecutor(max_workers=Q_N_WORKERS) as _ex:
            _futs = {}
            for _t, _c, _q in assigned_cells:
                _fu = _ex.submit(_cell_worker, _t, _c, _q)
                _futs[_fu] = (_c.get("_key", ""), _c.get("path_pattern", "?"))
            for _f in as_completed(_futs):
                _gap_key, _path = _futs.get(_f, ("", "?"))
                try:
                    _res = _f.result()
                except Exception as exc:
                    _msg = str(exc)
                    if _is_auth_error(_msg):
                        raise RuntimeError(
                            f"LLM authentication failed ({_msg}). "
                            f"Please set a valid school API key (see VQA_MODEL_NAME, default Qwen3.5-35B-A3B)."
                        ) from exc
                    if _gap_key:
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if GAP_FAIL_COOLDOWN > 0:
                            gap_fail_counts[_gap_key] = gap_fail_counts.get(_gap_key, 0) + 1
                    print(f"    [FAIL] [worker-exc] {_path} {_console_safe(_msg)}")
                    continue
                _path = _res.get("path", _path)
                _gap_key = _res.get("gap_key", _gap_key)
                if not _res.get("ok"):
                    if GAP_FAIL_COOLDOWN > 0 and _gap_key:
                        gap_fail_counts[_gap_key] = gap_fail_counts.get(_gap_key, 0) + 1
                    _stage = _res.get("stage", "unknown")
                    _method = _res.get("method", "?")
                    method_stage_counter[(_method, _stage)] += 1
                    method_total_counter[_method] += 1
                    if _stage == "fast":
                        dropped_fast += 1
                        print(f"    [WARN] [drop-fast] {_path} total={_res.get('total_ms',-1)}ms < {MIN_REAL_MS}ms")
                    elif _stage == "empty":
                        dropped_empty += 1
                        print(f"    [WARN] [drop-empty] path={_path}")
                    elif _stage == "wrong_verify_target":
                        dropped_wrong_verify_target += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(
                            f"    [WARN] [drop-wrong-target] {_path} "
                            f"expected={_res.get('target_n3', '?')} "
                            f"verify_n={_res.get('verify_n', -1)} "
                            f"ids={_res.get('verify_ids', '?')} "
                            f"raw={_console_safe(_res.get('verify', ''))}"
                        )
                    elif _stage == "non_unique":
                        dropped_non_unique += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(f"    [WARN] [drop-non-unique] {_path} verify={_console_safe(_res.get('verify', ''))}")
                    elif _stage == "low_iter":
                        dropped_low_iter += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(f"    [WARN] [drop-low-iter] {_path} iter={_res.get('iteration_count',-1)} < {MIN_ITER_COUNT}")
                    elif _stage == "qtype_mismatch":
                        dropped_qtype_mismatch += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(
                            f"    [WARN] [drop-qtype] {_path} q_type={_res.get('q_type', '?')} "
                            f"q={_console_safe(_res.get('question', ''))}"
                        )
                    else:
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(f"    [FAIL] [{_stage}] {_path}")
                    continue

                if _res.get("need_question_batch"):
                    pending_question_batch.append(_res)
                    continue
                _record_kept(_res)

        if pending_question_batch:
            _kept_after_qbuild: list = []
            _q_inputs = []
            for _res in pending_question_batch:
                _cell = _res.get("cell") or {}
                _qt = str(_res.get("q_type", "object") or "object")
                _ans = _answer_for_q_type(_qt, _cell, verify_n=_res.get("verify_n", -1))
                _res["answer"] = str(_ans)
                _q_inputs.append({
                    "q_type": _qt,
                    "n1_id": _cell.get("n1_id", ""),
                    "n1_type": _cell.get("n1_type", ""),
                    "n2_id": _cell.get("n2_id", ""),
                    "n2_type": _cell.get("n2_type", ""),
                    "n3_id": _cell.get("n3_id", ""),
                    "n3_type": _cell.get("n3_type", ""),
                    "n3_status": _cell.get("n3_status", ""),
                    "r1_dir": _cell.get("r1_dir8") or _cell.get("r1_dir4", "front"),
                    "r2_dir": _cell.get("r2_dir8") or _cell.get("r2_dir4", "front"),
                    "answer": str(_ans),
                    "fallback": _template_question(
                        _res.get("topology", "L2"),
                        _cell,
                        _qt,
                    ),
                })

            _q_t0 = time.perf_counter()
            _q_texts = []
            _q_batch_ms = 0.0
            _n_chunks = (len(_q_inputs) + Q_LLM_CHUNK_SIZE - 1) // Q_LLM_CHUNK_SIZE
            for _ci in range(_n_chunks):
                _s = _ci * Q_LLM_CHUNK_SIZE
                _e = min(len(_q_inputs), _s + Q_LLM_CHUNK_SIZE)
                _chunk = _q_inputs[_s:_e]
                _chunk_out = llm_client.generate_questions_batch(_chunk)
                _q_texts.extend(_chunk_out)
                _chunk_ms = float(llm_client.last_call_timing.get("total_ms", 0.0) or 0.0)
                if _chunk_ms > 0:
                    _q_batch_ms += _chunk_ms
            _q_elapsed_ms = max(0.0, (time.perf_counter() - _q_t0) * 1000.0)
            if _q_batch_ms <= 0:
                _q_batch_ms = _q_elapsed_ms
            _q_batch_per = _q_batch_ms / max(1, len(pending_question_batch))
            print(
                f"  [Question Build] mode=llm_batch "
                f"cells={len(pending_question_batch)} chunks={_n_chunks} "
                f"chunk_size={Q_LLM_CHUNK_SIZE} elapsed={int(_q_elapsed_ms)}ms"
            )

            for _idx, _res in enumerate(pending_question_batch):
                _path = _res.get("path", "?")
                _gap_key = _res.get("gap_key", "")
                _method = _res.get("method", "?")
                _cell = _res.get("cell") or {}
                _cell_key = str(_res.get("cell_key") or "")
                _work_qt = str(_res.get("q_type", "object") or "object")
                _question = str(
                    (_q_texts[_idx] if _idx < len(_q_texts) else _q_inputs[_idx].get("fallback", "")) or ""
                ).strip()
                _semantic_ok = False
                _strict_added_ms = 0.0

                if _question:
                    _eff = _reconcile_q_type_eff(_work_qt, _question)
                    if _qtype_semantic_ok(_eff, _question, method=_method):
                        _work_qt = _eff
                        _semantic_ok = True
                    else:
                        _work_qt = _eff

                if not _semantic_ok:
                    _strict_llm = _get_worker_llm()
                    for _ in range(max(1, MAX_QTYPE_RETRY)):
                        _ans_try = _answer_for_q_type(_work_qt, _cell, verify_n=_res.get("verify_n", -1))
                        try:
                            _question = _strict_llm.generate_question_nlp_strict(
                                path=_path,
                                q_type=_work_qt,
                                n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                                n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                                n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                                n3_status=_cell.get("n3_status",""),
                                r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                                r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                                constraint_desc=_res.get("constraint_desc", ""),
                                answer=str(_ans_try),
                                scene_distribution=scene_dist_global,
                            )
                        except Exception as _q_exc:
                            if _is_auth_error(_q_exc):
                                raise RuntimeError(
                                    f"LLM authentication failed ({_q_exc}). "
                                    f"Please set a valid school API key (see VQA_MODEL_NAME, default Qwen3.5-35B-A3B)."
                                ) from _q_exc
                            if _is_retryable_llm_exc(_q_exc):
                                time.sleep(0.8)
                                _question = _strict_llm.generate_question_nlp_strict(
                                    path=_path,
                                    q_type=_work_qt,
                                    n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                                    n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                                    n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                                    n3_status=_cell.get("n3_status",""),
                                    r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                                    r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                                    constraint_desc=_res.get("constraint_desc", ""),
                                    answer=str(_ans_try),
                                    scene_distribution=scene_dist_global,
                                )
                            else:
                                raise
                        _strict_added_ms += float(_strict_llm.last_call_timing.get("total_ms", 0.0) or 0.0)
                        _eff = _reconcile_q_type_eff(_work_qt, _question)
                        if _qtype_semantic_ok(_eff, _question, method=_method):
                            _work_qt = _eff
                            _semantic_ok = True
                            break
                        _work_qt = _eff

                if not _semantic_ok:
                    if GAP_FAIL_COOLDOWN > 0 and _gap_key:
                        gap_fail_counts[_gap_key] = gap_fail_counts.get(_gap_key, 0) + 1
                    method_stage_counter[(_method, "qtype_mismatch")] += 1
                    method_total_counter[_method] += 1
                    dropped_qtype_mismatch += 1
                    fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                    _fwu({
                        "stage": "verdict",
                        "cell_key": _cell_key,
                        "final_verdict": "drop-qtype",
                        "q_type_eff": _work_qt,
                        "rendered_question": _question,
                        "method_name": _method,
                        "method_used": _method,
                    })
                    print(
                        f"    [WARN] [drop-qtype] {_path} q_type={_work_qt} "
                        f"q={_console_safe(_question)}"
                    )
                    continue

                _res["q_type"] = _work_qt
                _res["answer"] = _answer_for_q_type(_work_qt, _cell, verify_n=_res.get("verify_n", -1))
                _res["question"] = _question
                _res["llm_ms"] = float(_res.get("llm_ms", 0.0) or 0.0) + _q_batch_per + _strict_added_ms
                _res["ts_end"] = _abs_ts()
                _res["total_ms"] = _dt_ms(_res.get("ts_start", ""), _res["ts_end"])
                if _res["total_ms"] >= 0 and _res["total_ms"] < MIN_REAL_MS:
                    if GAP_FAIL_COOLDOWN > 0 and _gap_key:
                        gap_fail_counts[_gap_key] = gap_fail_counts.get(_gap_key, 0) + 1
                    method_stage_counter[(_method, "fast")] += 1
                    method_total_counter[_method] += 1
                    dropped_fast += 1
                    print(f"    [WARN] [drop-fast] {_path} total={_res.get('total_ms',-1)}ms < {MIN_REAL_MS}ms")
                    continue
                _fwu({
                    "stage": "verdict",
                    "cell_key": _cell_key,
                    "final_verdict": "kept",
                    "verify_n": _res.get("verify_n", -1),
                    "method_name": _method,
                    "method_used": _method,
                    "rendered_question": _question,
                    "q_type_eff": _work_qt,
                })
                _kept_after_qbuild.append(_res)
            if _kept_after_qbuild:
                if EXCEL_BATCH_WRITE:
                    _rows = [_to_sheet_b_payload(_res) for _res in _kept_after_qbuild]
                    _w_t0 = time.perf_counter()
                    _ok_list = write_generated_questions_batch(_rows)
                    _w_ms = int((time.perf_counter() - _w_t0) * 1000)
                    _ok_n = sum(1 for _x in _ok_list if _x)
                    print(
                        f"  [Excel Write] mode=batch rows={len(_rows)} ok={_ok_n} elapsed={_w_ms}ms"
                    )
                    if len(_ok_list) != len(_kept_after_qbuild):
                        _ok_list = [False] * len(_kept_after_qbuild)
                    for _res, _ok in zip(_kept_after_qbuild, _ok_list):
                        _record_kept(_res, _write_ok=_ok)
                else:
                    for _res in _kept_after_qbuild:
                        _record_kept(_res)

        if batch_total_ms:
            avg_llm = int(sum(batch_llm_ms) / len(batch_llm_ms))
            avg_neo = int(sum(batch_neo_ms) / len(batch_neo_ms))
            avg_total = int(sum(batch_total_ms) / len(batch_total_ms))
            batch_elapsed_ms = int((time.perf_counter() - batch_start_time) * 1000)
            status = "REAL" if avg_total >= MIN_REAL_MS else "SUSPECT"
            print(f"[Batch Verify] Round {round_idx} | "
                  f"LLM Latency: {avg_llm}ms | Neo4j Latency: {avg_neo}ms | "
                  f"Total: {avg_total}ms | Status: {status}")
            print(f"[Batch Time] Round {round_idx} | "
                  f"Context: {context_build_ms}ms | "
                  f"LLM Total: {int(sum(batch_llm_ms))}ms | "
                  f"Neo4j Total: {int(sum(batch_neo_ms))}ms | "
                  f"Batch Total: {batch_elapsed_ms}ms")
        else:
            print(f"[Batch Verify] Round {round_idx} | Status: EMPTY")

        if batch_total_ms:
            empty_streak = 0
            # 保存覆盖状态
            save_coverage_state(tracker, coverage_file)
        else:
            empty_streak += 1
            if MAX_EMPTY_STREAK > 0 and empty_streak >= MAX_EMPTY_STREAK:
                print(
                    f"\n  [WARN] 连续 {empty_streak} 轮 Batch EMPTY，"
                    f"达到 VQA_MAX_EMPTY_STREAK={MAX_EMPTY_STREAK}，结束本帧 Gap。"
                )
                _print_gap_diag("max_empty_streak")
                break

    # 最终保存覆盖状态
    save_coverage_state(tracker, coverage_file)
    print(f"  [Coverage] Final state saved to {coverage_file}")

    qa_dir = pathlib.Path(GEN_QA_DIR)
    qa_dir.mkdir(parents=True, exist_ok=True)
    qa_path = qa_dir / f"{SCENE_ID}_frame{FRAME_ID}_qa.json"
    qa_path.write_text(
        _json.dumps({"questions": nusqa_records}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n  NuScenes-QA JSON -> {qa_path} ({len(nusqa_records)} records)")

    _print_stats("Gap Stats — Final")
    n_l2_gen = len(nusqa_records)
    print(f"\n  V18 Generated: {generated}/{total_att} attempts written")
    print(f"  L2: {n_l2_gen}")
    print(f"  Dropped (<{MIN_REAL_MS}ms): {dropped_fast}")
    print(f"  Dropped (empty fields): {dropped_empty}")
    print(f"  Dropped (verify n=1 but wrong target id): {dropped_wrong_verify_target}")
    print(f"  Dropped (non-unique / verify fail): {dropped_non_unique}")
    print(f"  Dropped (iteration<{MIN_ITER_COUNT}): {dropped_low_iter}")
    print(f"  Dropped (q_type semantic mismatch): {dropped_qtype_mismatch}")
    print(f"  Auto-downgraded (strict gate rescue): {auto_downgraded_total}")
    print(
        f"  Quality gates: strict_unique={STRICT_UNIQUE_ONLY} "
        f"min_iter={MIN_ITER_COUNT} low_iter_methods={sorted(LOW_ITER_STRICT_METHODS)} "
        f"min_real_ms={MIN_REAL_MS} "
        f"autodowngrade={AUTO_DOWNGRADE_NON_UNIQUE} "
        f"autodowngrade_qtype={AUTO_DOWNGRADE_QTYPE} "
        f"autodowngrade_methods={sorted(AUTO_DOWNGRADE_METHODS)} "
        f"context_mode={CONTEXT_CYPHER_MODE} question_mode={QUESTION_MODE}"
    )
    print(f"  Question types used: {list(dict.fromkeys(used_types))}")
    if auto_downgraded_total > 0:
        print(
            f"  Auto-downgrade reasons: {dict(auto_downgraded_reason_counter)}"
        )
    if method_total_counter:
        print("\n  [Quality Diagnose — method kept rate]")
        for _m, _tot in method_total_counter.most_common():
            if _m == "?":
                continue
            _k = method_keep_counter.get(_m, 0)
            _r = (_k / max(_tot, 1)) * 100.0
            print(f"    {_m:<22} kept={_k:>3}/{_tot:<3}  rate={_r:5.1f}%")
    if method_stage_counter:
        print("\n  [Quality Diagnose — top method×stage]")
        for (_m, _s), _c in method_stage_counter.most_common(16):
            print(f"    {_m:<22} {_s:<18} {_c}")
    return generated


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  Method A — Full Execution Chain (scene-0926 frame-20)")
    print("=" * 65)

    if not maybe_connect_school_vpn():
        print("\n[FAIL] VPN 预连失败或未配置。关闭 VQA_AUTO_CONNECT_VPN 可跳过此步。")
        return

    # Pre-flight
    if not preflight_check():
        print("\n[FAIL] Pre-flight failed. Fix issues above and re-run.")
        return

    if os.getenv("VQA_CLEAR_RUN_DATA", "").lower() in ("1", "true", "yes"):
        from advtest_cleanup import clear_excel_data_rows
        n = clear_excel_data_rows(EXCEL_PATH)
        print(f"\n[VQA_CLEAR_RUN_DATA] Excel cleared (removed ~{n} row-slots across sheets).")

    print("\n[OK] All pre-flight checks passed. Starting execution...\n")

    from neo4j_bootstrap import ensure_neo4j_listening
    if not ensure_neo4j_listening(NEO4J_URI):
        print("\n[FAIL] Neo4j 不可用。请启动数据库或配置 NEO4J_DOCKER_NAMES / Docker。")
        return

    from neo4j import GraphDatabase
    from gap_pipeline.llm_client import LLMClient

    llm    = LLMClient()
    try:
        _ = llm._call(
            "ping",
            system_prompt="Return only: pong",
            max_tokens=8,
            call_tag="audit_healthcheck",
        )
    except Exception as exc:
        print(f"\n[FAIL] LLM API health check failed: {exc}")
        print("   当前是鉴权/账号问题（不是本地流程逻辑问题）。")
        print("   请先配置有效 VQA_API_KEY，并确认 VQA_MODEL_NAME 与网关一致（默认 Qwen3.5-35B-A3B）。")
        return
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))

    try:
        # append-only policy: never wipe historical evidence unless explicitly opted in
        step0_wipe_previous_run()
        step1_cleanup()
        if not step1_5_refresh_filtered_sg():
            print("\n[FAIL] Cannot rebuild filtered SG with official policy.")
            return
        step2_filter_record()
        step3_import_neo4j()
        step4_baseline_audit(driver, llm)
        step5_6_generate(driver, llm)

        print("\n" + "=" * 65)
        print("  [OK] Method A complete")
        print("  Check RQ.xlsx:")
        print("    filter_record          ← upsert 1 row / frame (overwrite if exists)")
        print("    raw_coverage           ← ~29 new rows (baseline)")
        print("    question-answer-our    ← 5 new rows (generated)")
        print("=" * 65)

    finally:
        driver.close()


if __name__ == "__main__":
    _plan_frames = _load_plan_frames()
    for _pi, (_sid, _fid, _sg) in enumerate(_plan_frames, 1):
        # 设置全局变量供 main() 使用
        SCENE_ID = _sid
        FRAME_ID = _fid
        TARGET_SG = _sg
        print(f"\n{'='*70}")
        print(f"  Plan [{_pi}/{len(_plan_frames)}]: {SCENE_ID} frame-{FRAME_ID}")
        print(f"{'='*70}")
        try:
            main()
        except Exception as _exc:
            print(f"  [ERROR] {SCENE_ID} frame-{FRAME_ID}: {_exc}")
            import traceback; traceback.print_exc()
            continue
    print(f"\n  [Plan Complete] {len(_plan_frames)} frames processed.")
