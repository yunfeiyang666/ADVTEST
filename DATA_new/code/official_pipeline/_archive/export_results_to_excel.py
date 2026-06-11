"""
将 pipeline 生成的 QA JSON 导出为 Excel。
每帧独立一个文件夹，包含 JSON、CSV、Excel。
问题编号按 1,2,3 顺序递增。

用法:
  python export_results_to_excel.py
"""

import json, os, sys, csv, shutil
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── 路径配置 ──────────────────────────────────────────────
QA_DIR = Path(r"E:\Project\ADVTEST\DATA_new\generated_qa")
COVERAGE_DIR = Path(r"E:\Project\ADVTEST\DATA_new\code\official_pipeline\coverage_state")
SCENE_GRAPH_DIR = Path(r"E:\Project\ADVTEST\filtered_scene_graphs_official")
RESULTS_BASE = Path(r"E:\Project\ADVTEST\DATA_new\results")

# ── 样式 ─────────────────────────────────────────────────
HEADER_FONT = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply_header_style(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def _load_qa(scene_id, frame_id):
    path = QA_DIR / f"{scene_id}_frame{frame_id}_qa.json"
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("questions", [])


def _load_scene_graph(scene_id, frame_id):
    """获取场景图完整数据"""
    # 尝试多种路径
    candidates = [
        SCENE_GRAPH_DIR / f"{scene_id}_frame{frame_id}_scene_graph.json",
        SCENE_GRAPH_DIR / f"{scene_id}" / f"frame{frame_id}.json",
        SCENE_GRAPH_DIR / f"{scene_id}_f{frame_id}.json",
    ]
    for p in candidates:
        if p.exists():
            data = json.loads(p.read_text(encoding="utf-8"))
            # 兼容两种 key 命名
            data.setdefault("objects", data.get("nodes", []))
            data.setdefault("relations", data.get("edges", []))
            return data
    return {"objects": [], "relations": []}


def _load_coverage_state(scene_id, frame_id):
    path = COVERAGE_DIR / f"{scene_id}_frame{frame_id}.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


# ── Sheet 构造 ────────────────────────────────────────────

def _build_filter_record(wb, scene_id, frame_id):
    """filter_record: 场景图节点/边信息"""
    ws = wb.create_sheet("filter_record")
    headers = [
        "scene_id", "frame_id", "original_num", "filtered_num",
        "filtered_vex(node_ids)", "ratio", "timestamp",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    sg = _load_scene_graph(scene_id, frame_id)
    objects = sg.get("objects", [])
    node_ids = sorted([o.get("unique_id", "") for o in objects])
    n_total = len(objects)

    ws.append([
        scene_id, frame_id,
        n_total, n_total,
        ",".join(node_ids),
        1.0,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])

    # 添加每个节点详情
    ws2 = wb.create_sheet("node_details")
    node_headers = ["node_id", "type", "status", "position_x", "position_y", "position_z"]
    ws2.append(node_headers)
    _apply_header_style(ws2, len(node_headers))
    for obj in objects:
        pos = obj.get("position", {})
        ws2.append([
            obj.get("unique_id", ""),
            obj.get("type", ""),
            obj.get("status", ""),
            pos.get("x", ""),
            pos.get("y", ""),
            pos.get("z", ""),
        ])
    _auto_width(ws2)

    # 添加边详情
    ws3 = wb.create_sheet("edge_details")
    edge_headers = ["source", "target", "direction_6", "distance"]
    ws3.append(edge_headers)
    _apply_header_style(ws3, len(edge_headers))
    for rel in sg.get("relations", []):
        ws3.append([
            rel.get("source", ""),
            rel.get("target", ""),
            rel.get("direction_6", rel.get("direction_official", "")),
            rel.get("distance", ""),
        ])
    _auto_width(ws3)

    _auto_width(ws)


def _build_raw_coverage(wb, scene_id, frame_id):
    """raw_coverage: 原始 NuScenes-QA baseline 覆盖"""
    ws = wb.create_sheet("raw_coverage")
    headers = [
        "scene_id", "frame_id", "nuscenes_qa_id", "question", "answer",
        "L0", "L1", "L2", "question_type", "source",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    # 从 baseline QA 文件加载（如果存在）
    baseline_path = QA_DIR.parent / "baseline_qa" / f"{scene_id}_frame{frame_id}_baseline.json"
    if baseline_path.exists():
        baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
        for q in baseline.get("questions", []):
            ws.append([
                scene_id, frame_id,
                q.get("question_id", ""),
                q.get("question", ""),
                q.get("answer", ""),
                json.dumps(q.get("L0", []), ensure_ascii=False),
                json.dumps(q.get("L1", []), ensure_ascii=False),
                json.dumps(q.get("L2", []), ensure_ascii=False),
                q.get("template_type", ""),
                "nuscenes_qa_baseline",
            ])

    _auto_width(ws)


def _build_qa_sheet(wb, scene_id, frame_id):
    """question-answer-our: 生成的 QA，按 1,2,3 编号"""
    ws = wb.create_sheet("question-answer-our")
    headers = [
        "seq", "question_id", "gap_cell",
        "timestamp_start", "timestamp_llm", "timestamp_cypher_return", "timestamp_end",
        "iteration_count", "method_used", "question_type", "complexity",
        "natural language question", "cypher question", "answer",
        "verify_n", "is_unique",
        "L0_nodes", "L1_edges", "L2_pivot",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    questions = _load_qa(scene_id, frame_id)
    for seq, q in enumerate(questions, start=1):
        path_pattern = q.get("path_pattern", "")
        parts = path_pattern.split("|") if "|" in path_pattern else path_pattern.split("\u2192")
        n1 = parts[0] if len(parts) > 0 else ""
        n2 = parts[1] if len(parts) > 1 else ""
        n3 = parts[2] if len(parts) > 2 else ""

        l0_nodes = sorted(set(p for p in parts if p))
        l1_edges = []
        if len(parts) >= 2:
            l1_edges.append(f"{n1}-{n2}")
        if len(parts) >= 3:
            l1_edges.append(f"{n2}-{n3}")

        ws.append([
            seq,
            q.get("question_id", ""),
            path_pattern,
            q.get("timestamp_start", ""),
            q.get("timestamp_llm", ""),
            q.get("timestamp_cypher_return", ""),
            q.get("timestamp_end", ""),
            q.get("iteration_count", 0),
            q.get("method_used", ""),
            q.get("template_type", ""),
            q.get("topology_level", ""),
            q.get("question", ""),
            q.get("verify_cypher", ""),
            q.get("answer", ""),
            q.get("verify_n", -1),
            q.get("is_unique", False),
            json.dumps(l0_nodes, ensure_ascii=False),
            json.dumps(l1_edges, ensure_ascii=False),
            path_pattern,
        ])

    _auto_width(ws)
    return len(questions)


def _build_coverage_summary(wb, scene_id, frame_id, n_qa):
    """coverage_summary: 覆盖率统计 (从实际 hit_count 计算)"""
    ws = wb.create_sheet("coverage_summary")
    headers = [
        "level", "total", "covered", "uncovered", "rate",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    cov = _load_coverage_state(scene_id, frame_id)
    for level in ("L0", "L1", "L2"):
        level_data = cov.get(level, {})
        total = len(level_data)
        covered = sum(1 for v in level_data.values() if isinstance(v, dict) and v.get("hit_count", 0) > 0)
        uncovered = total - covered
        rate = (covered / total * 100) if total > 0 else 0.0
        ws.append([level, total, covered, uncovered, f"{rate:.1f}%"])

    ws.append([])
    ws.append(["Total QA generated", n_qa])
    ws.append(["Scene", scene_id])
    ws.append(["Frame", frame_id])
    ws.append(["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    _auto_width(ws)


def _build_model_performance(wb):
    """model_performance_raw_our: 空表头"""
    ws = wb.create_sheet("model_performance_raw_our")
    headers = [
        "seq", "question_id", "model_name",
        "model_answer", "correct_answer", "question_type",
        "pass", "timestamp",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))
    _auto_width(ws)


def _write_csv(questions, csv_path, scene_id, frame_id):
    """写 CSV 文件"""
    fieldnames = [
        "seq", "question_id", "gap_cell", "question_type", "method_used",
        "question", "answer", "verify_cypher", "verify_n", "is_unique",
        "iteration_count", "timestamp_start", "timestamp_end", "total_ms",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for seq, q in enumerate(questions, start=1):
            writer.writerow({
                "seq": seq,
                "question_id": q.get("question_id", ""),
                "gap_cell": q.get("path_pattern", ""),
                "question_type": q.get("template_type", ""),
                "method_used": q.get("method_used", ""),
                "question": q.get("question", ""),
                "answer": q.get("answer", ""),
                "verify_cypher": q.get("verify_cypher", ""),
                "verify_n": q.get("verify_n", -1),
                "is_unique": q.get("is_unique", False),
                "iteration_count": q.get("iteration_count", 0),
                "timestamp_start": q.get("timestamp_start", ""),
                "timestamp_end": q.get("timestamp_end", ""),
                "total_ms": q.get("total_ms", 0),
            })


def create_frame_output(scene_id, frame_id):
    """为单帧创建独立输出文件夹"""
    frame_dir = RESULTS_BASE / f"{scene_id}_frame{frame_id}"
    frame_dir.mkdir(parents=True, exist_ok=True)

    questions = _load_qa(scene_id, frame_id)
    print(f"\n  {scene_id}/frame{frame_id}: {len(questions)} questions")

    # 1. 复制原始 JSON
    src_json = QA_DIR / f"{scene_id}_frame{frame_id}_qa.json"
    if src_json.exists():
        shutil.copy2(src_json, frame_dir / f"qa.json")
        print(f"    JSON -> {frame_dir / 'qa.json'}")

    # 2. 写 CSV
    csv_path = frame_dir / f"qa.csv"
    _write_csv(questions, csv_path, scene_id, frame_id)
    print(f"    CSV  -> {csv_path}")

    # 3. 写 Excel
    wb = Workbook()
    wb.remove(wb.active)

    _build_filter_record(wb, scene_id, frame_id)
    _build_raw_coverage(wb, scene_id, frame_id)
    n_qa = _build_qa_sheet(wb, scene_id, frame_id)
    _build_coverage_summary(wb, scene_id, frame_id, n_qa)
    _build_model_performance(wb)

    xlsx_path = frame_dir / f"results.xlsx"
    wb.save(str(xlsx_path))
    print(f"    XLSX -> {xlsx_path}")

    # 4. 复制覆盖率状态
    cov_src = COVERAGE_DIR / f"{scene_id}_frame{frame_id}.json"
    if cov_src.exists():
        shutil.copy2(cov_src, frame_dir / "coverage_state.json")

    return len(questions)


def main():
    frames = []
    for p in sorted(QA_DIR.glob("*_qa.json")):
        name = p.stem
        parts = name.replace("_qa", "").split("_frame")
        if len(parts) == 2:
            frames.append((parts[0], int(parts[1])))

    if not frames:
        print("No QA files found!")
        sys.exit(1)

    print(f"Found {len(frames)} frames: {frames}")
    RESULTS_BASE.mkdir(parents=True, exist_ok=True)

    total = 0
    for scene_id, frame_id in frames:
        total += create_frame_output(scene_id, frame_id)

    print(f"\nDone! Total {total} QA across {len(frames)} frames")
    print(f"Results in: {RESULTS_BASE}")


if __name__ == "__main__":
    main()
