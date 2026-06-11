#!/usr/bin/env python3
"""
run_v17_production.py — V17 深夜量产自动化流水线
==================================================
策略：
  - 追加模式：不删除现有 Excel 记录，全部以 append 写入
  - 每帧独立 try/except，某帧失败不影响后续帧
  - V20 批量架构：生成 Batch-15 + 15-Worker；baseline 审计并行批处理
  - 每题立即 wb.save()（rq_tables 底层已保证）
  - 每帧生成独立 JSON: generated_qa/{scene}_{frame}_qa.json

目标帧（顺序执行）：
  1. scene-0553 / frame-8   (24 val qs)
  2. scene-0757 / frame-26  (0  val qs)   ← 无 baseline，仍跑 gap gen
  3. scene-1077 / frame-19  (0  val qs)   ← 无 baseline，仍跑 gap gen
  4. scene-0103 / frame-0   (5  val qs)
  5. scene-0103 / frame-38  (14 val qs)

执行前确认：
  ✅ RQ.xlsx 已关闭（无锁定文件）
  ✅ Neo4j bolt://localhost:7800 可达
  ✅ filtered_scene_graphs/ 中存在上述所有 SG 文件
"""
import os, sys, pathlib, json, time, collections, logging
from datetime import datetime

sys.path.insert(0, str(pathlib.Path(__file__).parent))

# ─── 全局路径 ────────────────────────────────────────────────────────────────
EXCEL_PATH  = pathlib.Path("E:/Project/ADVTEST/RQ(1).xlsx")
FSG_DIR     = pathlib.Path(os.getenv("FILTERED_SG_DIR", "E:/Project/ADVTEST/filtered_scene_graphs"))
TRAINVAL    = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/v1.0-mini")
QA_PATH     = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/qa/NuScenes_val_questions.json")
GEN_QA_DIR  = pathlib.Path("E:/Project/ADVTEST/generated_qa")
NEO4J_URI   = "bolt://localhost:7800"
NEO4J_USER  = "neo4j"
NEO4J_PWD   = "87017563"

# ─── 量产帧计划 ──────────────────────────────────────────────────────────────
# (scene_id, frame_id, sg_filename)
FRAME_PLAN = [
    ("scene-0553",  8, "scene-0553_frame8_scene_graph.json"),
    ("scene-0757", 26, "scene-0757_frame26_scene_graph.json"),
    ("scene-1077", 19, "scene-1077_frame19_scene_graph.json"),
    ("scene-0103",  0, "scene-0103_frame0_scene_graph.json"),
    ("scene-0103", 38, "scene-0103_frame38_scene_graph.json"),
]

# ─── 日志配置 ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("v17_prod")

# ─────────────────────────────────────────────────────────────────────────────
# 预检
# ─────────────────────────────────────────────────────────────────────────────

def preflight() -> bool:
    print("\n[Pre-flight]")
    ok = True

    # Excel writable?
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        print(f"  ❌ Excel 锁定文件存在: {lock}  → 请先关闭 Excel")
        ok = False
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            wb.save(str(EXCEL_PATH)); wb.close()
            print(f"  ✅ Excel 可写")
        except PermissionError:
            print(f"  ❌ Excel 写入测试失败 — 请关闭 Excel 后重试")
            ok = False

    # Neo4j
    try:
        from neo4j import GraphDatabase
        d = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))
        with d.session() as s: s.run("RETURN 1").single()
        d.close()
        print(f"  ✅ Neo4j 可达: {NEO4J_URI}")
    except Exception as e:
        print(f"  ❌ Neo4j 不可达: {e}")
        ok = False

    # SG files
    for sid, fid, sg in FRAME_PLAN:
        p = FSG_DIR / sg
        if p.exists():
            kb = p.stat().st_size // 1024
            print(f"  ✅ {sg} ({kb} KB)")
        else:
            print(f"  ❌ 缺失: {sg}")
            ok = False

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# 帧流水线：步骤 2-6
# ─────────────────────────────────────────────────────────────────────────────

def _abs_ts() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]


def run_step2_filter_record(scene_id: str, frame_id: int, sg_name: str) -> bool:
    """写入 filter_record（追加，不检查重复）。"""
    from rq_tables import write_filter_record
    sg = FSG_DIR / sg_name
    data = json.loads(sg.read_text(encoding="utf-8"))
    info = data["core_universe_filter"]
    vex_str = ",".join(sorted(info["node_ids_kept"]))
    ratio   = info["filtered_nodes"] / max(info["raw_nodes"], 1)
    ok = write_filter_record(
        scene_id=scene_id, frame_id=frame_id,
        original_num=info["raw_nodes"],
        filtered_num=info["filtered_nodes"],
        filtered_vex=vex_str,
        ratio=ratio,
    )
    print(f"  filter_record: raw={info['raw_nodes']} → filtered={info['filtered_nodes']} "
          f"(ratio={ratio:.2%})  {'✅' if ok else '❌'}")
    return ok


def run_step3_import_neo4j(sg_name: str) -> dict:
    """清空 Neo4j 并导入当前帧场景图。"""
    from core_universe_filter import import_filtered_sg_to_neo4j
    result = import_filtered_sg_to_neo4j(
        sg_name=sg_name,
        neo4j_uri=NEO4J_URI, neo4j_user=NEO4J_USER, neo4j_pwd=NEO4J_PWD,
    )
    print(f"  Neo4j: {result['n_nodes']} nodes, {result['n_edges']} edges ✅")
    return result


def run_step4_baseline_audit(
        scene_id: str, frame_id: int,
        driver, llm_client,
) -> int:
    """baseline 审计（val 原题 → 足迹 → raw_coverage）。"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    from semantic_auditor import audit_baseline_question, build_scene_context
    from semantic_auditor import make_qa_id, _ms_now as _audit_ms, derive_l2_from_l1
    from rq_tables import write_baseline_to_coverage

    # 构建 sample_token → (scene_name, frame_idx) 映射
    scenes  = json.loads((TRAINVAL / "scene.json").read_text())
    samples = json.loads((TRAINVAL / "sample.json").read_text())
    st2name = {s["token"]: s["name"] for s in scenes}
    tok2info: dict = {}
    s2tok: dict = collections.defaultdict(list)
    for samp in samples:
        sname = st2name.get(samp["scene_token"], "?")
        tok2info[samp["token"]] = {"scene_name": sname, "timestamp": samp["timestamp"]}
        s2tok[sname].append(samp["token"])
    for sname, toks in s2tok.items():
        for idx, tok in enumerate(sorted(toks, key=lambda t: tok2info[t]["timestamp"])):
            tok2info[tok]["frame_idx"] = idx

    all_val_qs = json.loads(QA_PATH.read_text())["questions"]
    target_qs = [
        (gi, q)
        for gi, q in enumerate(all_val_qs)
        if tok2info.get(q.get("sample_token", ""), {}).get("scene_name") == scene_id
        and tok2info.get(q.get("sample_token", ""), {}).get("frame_idx") == frame_id
    ]
    print(f"  Found {len(target_qs)} val questions for {scene_id}/f{frame_id}")
    if not target_qs:
        print("  (no baseline questions — skipping audit, gap gen will still run)")
        return 0

    scene_ctx = build_scene_context(driver)
    print(f"  Scene context: {len(scene_ctx)} chars")
    BASELINE_N_WORKERS = 15
    _tls = threading.local()
    _llm_cls = llm_client.__class__

    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            _llm = _llm_cls()
            _tls.llm = _llm
        return _llm

    def _audit_one(gi: int, q: dict) -> dict:
        question = q.get("question", "")
        answer = q.get("answer", "")
        qtype = q.get("template_type", "")
        nhop = q.get("num_hop", 0)
        qa_uid = make_qa_id(gi, qtype)
        ts0 = _audit_ms()
        try:
            audit_res = audit_baseline_question(
                question=question, q_type=qtype, num_hop=nhop,
                driver=driver, llm_client=_get_worker_llm(),
                scene_context=scene_ctx, global_index=gi,
            )
        except Exception as exc:
            print(f"    ⚠️ [baseline-worker-exc] idx={gi} {exc}")
            audit_res = {"l0_nodes": [], "l1_edges": []}
        ts1 = _audit_ms()
        l0 = audit_res.get("l0_nodes", []) or []
        l1 = audit_res.get("l1_edges", []) or []
        l2 = derive_l2_from_l1(l1)
        audit_json = json.dumps(
            {"nodes": l0, "edges": l1},
            ensure_ascii=False
        )
        return {
            "qa_uid": qa_uid,
            "qtype": qtype,
            "question": question,
            "answer": answer,
            "l0_nodes": l0,
            "l1_edges": l1,
            "l2_paths": l2,
            "timestamp_start": ts0,
            "timestamp_end": ts1,
            "audit_cypher": audit_json,
        }

    t_batch0 = time.perf_counter()
    ordered = [None] * len(target_qs)
    with ThreadPoolExecutor(max_workers=BASELINE_N_WORKERS) as _ex:
        fut2idx = {
            _ex.submit(_audit_one, gi, q): idx
            for idx, (gi, q) in enumerate(target_qs)
        }
        for done_i, _f in enumerate(as_completed(fut2idx), 1):
            idx = fut2idx[_f]
            rec = _f.result()
            ordered[idx] = rec
            if done_i <= 3 or done_i % 10 == 0:
                print(f"  [audit {done_i:2d}/{len(target_qs)}] {rec['qa_uid']:<28} "
                      f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} L2={len(rec['l2_paths'])}")
    batch_ms = int((time.perf_counter() - t_batch0) * 1000)
    avg_ms = int(batch_ms / max(len(target_qs), 1))
    print(f"  [Baseline Batch] questions={len(target_qs)} workers={BASELINE_N_WORKERS} "
          f"elapsed={batch_ms}ms avg={avg_ms}ms/q")

    success = 0
    for i, rec in enumerate(ordered, 1):
        if rec is None:
            continue
        ok = write_baseline_to_coverage(
            scene_id=scene_id, frame_id=frame_id,
            nuscenes_qa_id=rec["qa_uid"],
            question=rec["question"], answer=rec["answer"],
            l0_nodes=rec["l0_nodes"], l1_edges=rec["l1_edges"],
            l2_paths=rec["l2_paths"], question_type=rec["qtype"],
            timestamp_start=rec["timestamp_start"], timestamp_end=rec["timestamp_end"],
            audit_cypher=rec["audit_cypher"],
        )
        if ok:
            success += 1
        if i <= 3 or i % 10 == 0:
            print(f"  [write {i:2d}/{len(target_qs)}] {rec['qa_uid']:<28} "
                  f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} "
                  f"L2={len(rec['l2_paths'])} {'OK' if ok else '--'}")

    print(f"  Baseline audit: {success}/{len(target_qs)} written ✅")
    return success


def run_step56_generate(
        scene_id: str, frame_id: int,
        driver, llm_client,
) -> dict:
    """
    V16 批量生成（从 run_method_a 借用逻辑，参数化 scene/frame）。
    通过临时修改 run_method_a 模块全局变量来复用已有 V16 实现。
    返回 {"generated": N, "attempts": M, "unresolvable": K, "qa_json_path": ...}
    """
    import run_method_a as _rma

    # 临时替换模块级全局变量
    _old = (_rma.SCENE_ID, _rma.FRAME_ID, _rma.TARGET_SG)
    _rma.SCENE_ID  = scene_id
    _rma.FRAME_ID  = frame_id
    _rma.TARGET_SG = f"{scene_id}_frame{frame_id}_scene_graph.json"
    try:
        n = _rma.step5_6_generate(driver, llm_client)
    finally:
        _rma.SCENE_ID, _rma.FRAME_ID, _rma.TARGET_SG = _old

    qa_path = GEN_QA_DIR / f"{scene_id}_frame{frame_id}_qa.json"
    n_json  = 0
    if qa_path.exists():
        try:
            n_json = len(json.loads(qa_path.read_text())["questions"])
        except Exception:
            pass
    return {
        "generated":   n,
        "qa_json_path": str(qa_path),
        "n_json":       n_json,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 单帧执行入口
# ─────────────────────────────────────────────────────────────────────────────

def run_single_frame(
        scene_id: str, frame_id: int, sg_name: str,
        driver, llm_client,
) -> dict:
    """
    执行单帧完整流水线（步骤 2-6）。
    失败时抛出异常，由外层捕获记录。
    返回统计摘要 dict。
    """
    t0 = time.perf_counter()
    label = f"{scene_id}/f{frame_id}"
    print(f"\n{'─'*65}")
    print(f"  [{label}]  SG={sg_name}")
    print(f"{'─'*65}")

    # Step 2: filter_record
    run_step2_filter_record(scene_id, frame_id, sg_name)

    # Step 3: Neo4j 导入
    run_step3_import_neo4j(sg_name)

    # Step 4: Baseline 审计
    n_baseline = run_step4_baseline_audit(scene_id, frame_id, driver, llm_client)

    # Step 5+6: V16 生成
    gen_result = run_step56_generate(scene_id, frame_id, driver, llm_client)

    elapsed_s = time.perf_counter() - t0
    stats = {
        "scene_id":    scene_id,
        "frame_id":    frame_id,
        "n_baseline":  n_baseline,
        "n_generated": gen_result["generated"],
        "n_json":      gen_result["n_json"],
        "qa_json_path": gen_result["qa_json_path"],
        "elapsed_s":   round(elapsed_s, 1),
        "status":      "OK",
    }
    print(f"\n  [{label}] ✅ DONE  baseline={n_baseline}  generated={gen_result['generated']}  "
          f"elapsed={elapsed_s:.0f}s")
    return stats


# ─────────────────────────────────────────────────────────────────────────────
# 量产汇总报告
# ─────────────────────────────────────────────────────────────────────────────

def print_summary(results: list, total_elapsed_s: float):
    print(f"\n{'='*65}")
    print(f"  V17 量产汇总报告  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*65}")

    ok_frames  = [r for r in results if r["status"] == "OK"]
    err_frames = [r for r in results if r["status"] != "OK"]

    total_baseline  = sum(r.get("n_baseline", 0) for r in ok_frames)
    total_generated = sum(r.get("n_generated", 0) for r in ok_frames)

    print(f"\n  帧执行情况:")
    for r in results:
        mark = "✅" if r["status"] == "OK" else "❌"
        if r["status"] == "OK":
            print(f"    {mark} {r['scene_id']}/f{r['frame_id']:<3}  "
                  f"baseline={r['n_baseline']:>3}  generated={r['n_generated']:>4}  "
                  f"json={r['n_json']:>4}  {r['elapsed_s']:.0f}s")
        else:
            print(f"    {mark} {r['scene_id']}/f{r['frame_id']:<3}  "
                  f"ERROR: {r.get('error','?')[:60]}")

    print(f"\n  汇总统计:")
    print(f"    成功帧数            : {len(ok_frames)} / {len(results)}")
    print(f"    新增 raw_coverage 行 : {total_baseline}")
    print(f"    新增 question-answer 行: {total_generated}")
    print(f"    总耗时              : {total_elapsed_s:.0f}s  ({total_elapsed_s/60:.1f} min)")

    # Excel 当前行数
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
        for sh in ["raw_coverage", "question-answer-our"]:
            ws = wb[sh]
            n_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True)
                         if any(v is not None for v in _))
            print(f"    {sh:<28}: 共 {n_rows} 数据行")
        wb.close()
    except Exception as exc:
        print(f"    (Excel 行数读取失败: {exc})")

    if err_frames:
        print(f"\n  失败帧详情:")
        for r in err_frames:
            print(f"    {r['scene_id']}/f{r['frame_id']}: {r.get('error','?')}")

    print(f"\n  生成 JSON 文件:")
    for r in ok_frames:
        p = pathlib.Path(r["qa_json_path"])
        exists = "✅" if p.exists() else "❌ missing"
        print(f"    {exists}  {r['qa_json_path']}")

    print(f"{'='*65}")


# ─────────────────────────────────────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  run_v17_production.py — 深夜量产自动化流水线")
    print(f"  目标帧: {len(FRAME_PLAN)} 帧")
    print(f"  模式: 追加 (不删除现有 Excel 记录)")
    print("=" * 65)

    if not preflight():
        print("\n❌ Pre-flight 失败，请修复后重试。")
        return

    print("\n✅ Pre-flight 通过，开始量产...\n")

    from neo4j import GraphDatabase
    from gap_pipeline.llm_client import LLMClient

    llm    = LLMClient()
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))

    results = []
    t_total_start = time.perf_counter()

    try:
        for idx, (scene_id, frame_id, sg_name) in enumerate(FRAME_PLAN, 1):
            print(f"\n{'█'*65}")
            print(f"  帧 {idx}/{len(FRAME_PLAN)}: {scene_id}/frame-{frame_id}")
            print(f"{'█'*65}")
            try:
                stats = run_single_frame(scene_id, frame_id, sg_name, driver, llm)
                results.append(stats)
            except KeyboardInterrupt:
                print(f"\n⚠️  用户中断 (Ctrl+C)，保存已完成帧的结果...")
                results.append({
                    "scene_id": scene_id, "frame_id": frame_id,
                    "status": "INTERRUPTED",
                    "error": "KeyboardInterrupt",
                })
                break
            except Exception as exc:
                import traceback
                err_msg = str(exc)
                print(f"\n  ❌ [{scene_id}/f{frame_id}] 帧失败: {err_msg[:120]}")
                print(f"  跳过该帧，继续下一帧...")
                traceback.print_exc()
                results.append({
                    "scene_id": scene_id, "frame_id": frame_id,
                    "status": "ERROR",
                    "error": err_msg,
                    "n_baseline":  0,
                    "n_generated": 0,
                    "n_json":      0,
                    "elapsed_s":   0,
                })

    finally:
        driver.close()
        total_elapsed_s = time.perf_counter() - t_total_start
        print_summary(results, total_elapsed_s)

        # 保存汇总 JSON 到 generated_qa/
        summary_path = GEN_QA_DIR / f"v17_production_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        try:
            GEN_QA_DIR.mkdir(parents=True, exist_ok=True)
            summary_path.write_text(
                json.dumps({
                    "run_time":    datetime.now().isoformat(),
                    "total_frames": len(FRAME_PLAN),
                    "results":     results,
                    "total_elapsed_s": total_elapsed_s,
                }, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"\n  汇总 JSON → {summary_path}")
        except Exception as exc:
            print(f"  ⚠️ 汇总 JSON 保存失败: {exc}")


if __name__ == "__main__":
    main()
