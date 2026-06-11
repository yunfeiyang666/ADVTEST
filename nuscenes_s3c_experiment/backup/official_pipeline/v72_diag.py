#!/usr/bin/env python3
"""V7.2 Diagnostic: Excel audit + write test + low-density frame finder."""
import json, pathlib, sys, time, collections
sys.path.insert(0, str(pathlib.Path(__file__).parent))

import openpyxl

EXCEL_PATH = pathlib.Path("E:/Project/ADVTEST/RQ.xlsx")
TRAINVAL   = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/v1.0-trainval")
QA_PATH    = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/qa/NuScenes_val_questions.json")
SG_DIR     = pathlib.Path("E:/Project/ADVTEST/nuscenes_s3c_experiment/output/coverage_analysis/scene_graphs")

CORE_TYPES  = {"car","truck","bus","pedestrian","bicycle","motorcycle"}
MAX_DIST_M  = 20.0

print("=" * 65)
print(f"  V7.2 Diagnostic")
print(f"  Excel path: {EXCEL_PATH.resolve()}")
print("=" * 65)

# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Re-read ALL sheet headers
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1] Reading Excel headers (all sheets)...")
wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
print(f"    Sheets: {wb.sheetnames}")
all_headers = {}
for shname in wb.sheetnames:
    ws = wb[shname]
    rows = list(ws.iter_rows(values_only=True, max_row=3))
    if rows:
        header_row = [str(v) if v is not None else "" for v in rows[0]]
        fmt_row    = [str(v) if v is not None else "" for v in rows[1]] if len(rows)>1 else []
        all_headers[shname] = {"headers": header_row, "format": fmt_row}
        print(f"\n  Sheet [{shname}]  ({len(header_row)} cols)")
        for i, (h, f) in enumerate(zip(header_row, fmt_row), 1):
            print(f"    col{i:02d}: '{h}'  ← {f[:50] if f else ''}")
wb.close()

# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Compare with rq_tables.py current schema
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2] Schema diff vs rq_tables.py...")
from rq_tables import COLS_A, COLS_B, COLS_C, SHEET_A, SHEET_B, SHEET_C

for sheet_name, code_cols, key in [
    (SHEET_A, COLS_A, "COLS_A"),
    (SHEET_B, COLS_B, "COLS_B"),
    (SHEET_C, COLS_C, "COLS_C"),
]:
    excel_cols = all_headers.get(sheet_name, {}).get("headers", [])
    excel_cols_clean = [c for c in excel_cols if c]
    code_set  = set(code_cols)
    excel_set = set(excel_cols_clean)
    missing_in_code  = excel_set - code_set   # in Excel but NOT in code
    extra_in_code    = code_set  - excel_set  # in code but NOT in Excel
    order_mismatch   = (code_cols != excel_cols_clean) and not missing_in_code and not extra_in_code

    print(f"\n  {sheet_name} ({key}):")
    if missing_in_code:
        print(f"    ⚠️  NEW cols in Excel (missing from code): {sorted(missing_in_code)}")
    if extra_in_code:
        print(f"    ⚠️  Cols in code but NOT in Excel: {sorted(extra_in_code)}")
    if order_mismatch:
        print(f"    ⚠️  Column ORDER differs:")
        print(f"        Excel: {excel_cols_clean}")
        print(f"        Code : {code_cols}")
    if not missing_in_code and not extra_in_code and not order_mismatch:
        print(f"    ✅ Perfectly aligned ({len(code_cols)} cols)")

# ─────────────────────────────────────────────────────────────────────────────
# Step 3: TEST_SYNC write + read-back
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3] TEST_SYNC write to raw_coverage...")
from rq_tables import write_table_a

ts = time.strftime("%Y-%m-%dT%H:%M:%S")
ok = write_table_a(
    scene_id="TEST_SYNC",
    frame_id=0,
    nuscenes_qa_id="diag_001",
    question="Diagnostic test — V7.2",
    answer="ok",
    l0_nodes=["ego","car1"],
    l1_edges=[{"source":"ego","target":"car1","dir":"front"}],
    l2_paths=[],
)
print(f"  write_table_a returned: {ok}")

# Read back last row
wb2 = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
ws2 = wb2[SHEET_A]
all_rows = [r for r in ws2.iter_rows(values_only=True) if any(v is not None for v in r)]
wb2.close()
last = all_rows[-1]
print(f"  Last row in raw_coverage (total rows now={len(all_rows)}):")
for i, (col, val) in enumerate(zip(COLS_A + ["..."], list(last) + [""])):
    if val is not None:
        print(f"    col{i+1:02d} '{col}' = {str(val)[:80]}")

# ─────────────────────────────────────────────────────────────────────────────
# Step 4: Load val questions + sample→scene mapping
# ─────────────────────────────────────────────────────────────────────────────
print("\n[4] Building sample→scene mapping...")
scenes  = json.loads((TRAINVAL/"scene.json").read_text())
samples = json.loads((TRAINVAL/"sample.json").read_text())

scene_token2name = {s["token"]: s["name"] for s in scenes}
sample2info: dict = {}
scene2samples: dict = collections.defaultdict(list)
for samp in samples:
    sname = scene_token2name.get(samp["scene_token"],"?")
    sample2info[samp["token"]] = {
        "scene_name":  sname,
        "scene_token": samp["scene_token"],
        "timestamp":   samp["timestamp"],
    }
    scene2samples[sname].append(samp["token"])

# Assign frame_idx
for sname, toks in scene2samples.items():
    sorted_toks = sorted(toks, key=lambda t: sample2info[t]["timestamp"])
    for idx, tok in enumerate(sorted_toks):
        sample2info[tok]["frame_idx"] = idx

print(f"  Mapped {len(sample2info)} samples across {len(scene2samples)} scenes")

# Load val questions
val_qs = json.loads(QA_PATH.read_text())["questions"]
print(f"  Val questions: {len(val_qs)}")

# Per-(scene,frame) QA count
frame_qa: dict = collections.defaultdict(list)
for q in val_qs:
    info = sample2info.get(q.get("sample_token",""), {})
    key  = (info.get("scene_name","?"), info.get("frame_idx",-1))
    frame_qa[key].append(q)

# ─────────────────────────────────────────────────────────────────────────────
# Step 5: Apply "Core Universe" filter to each scene graph we have
# ─────────────────────────────────────────────────────────────────────────────
print("\n[5] Applying Core Universe filter (20m, types={car,truck,bus,ped,bicycle,motorcycle}; edges unfiltered)")

def load_sg_and_filter(sg_path: pathlib.Path):
    """Load scene graph, apply core-universe filter, return filtered node count."""
    data = json.loads(sg_path.read_text(encoding="utf-8"))
    nodes_raw = data.get("nodes", [])
    edges_raw = data.get("edges", [])

    # Build ego position (first node with type='ego')
    ego_pos = None
    for n in nodes_raw:
        if n.get("type") == "ego" or n.get("unique_id") == "ego":
            t = n.get("translation", {})
            if isinstance(t, dict):
                ego_pos = (t.get("x",0), t.get("y",0))
            elif isinstance(t, list):
                ego_pos = (t[0], t[1])
            break

    # Filter nodes
    keep_ids = set()
    keep_ids.add("ego")
    filtered_nodes = []
    for n in nodes_raw:
        uid  = n.get("unique_id","")
        ntype = n.get("type","")
        if uid == "ego":
            filtered_nodes.append(n)
            continue
        if ntype not in CORE_TYPES:
            continue
        # Distance check using edges (ego→node distance)
        # Or check from translation if ego_pos known
        if ego_pos is not None:
            t = n.get("translation", {})
            if isinstance(t, dict):
                dx = t.get("x",0) - ego_pos[0]
                dy = t.get("y",0) - ego_pos[1]
            elif isinstance(t, list) and len(t) >= 2:
                dx = t[0] - ego_pos[0]
                dy = t[1] - ego_pos[1]
            else:
                dx = dy = 0
            dist = (dx**2 + dy**2)**0.5
            if dist > MAX_DIST_M:
                continue
        filtered_nodes.append(n)
        keep_ids.add(uid)

    filtered_edges = [
        e for e in edges_raw
        if e.get("source","") in keep_ids and e.get("target","") in keep_ids
    ]
    return len(filtered_nodes), filtered_nodes, filtered_edges, len(edges_raw)

# Check existing scene graphs
sg_files = sorted(SG_DIR.glob("*_scene_graph.json"))
print(f"  Found {len(sg_files)} scene graphs")
sg_results = {}
for f in sg_files:
    stem = f.stem.replace("_scene_graph","")
    parts = stem.rsplit("_frame", 1)
    sname = parts[0]
    fidx  = int(parts[1]) if len(parts)>1 else -1
    n_filtered, _, _, n_edges = load_sg_and_filter(f)
    key = (sname, fidx)
    sg_results[key] = n_filtered
    n_qa = len(frame_qa.get(key, []))
    print(f"    {stem:<45} nodes_filtered={n_filtered:2d}  val_qa={n_qa}")

# ─────────────────────────────────────────────────────────────────────────────
# Step 6: Find low-density frames (5-10 val QA, 5-8 filtered nodes)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[6] Low-density frame candidates (5-10 val QA, 5-8 filtered nodes):")
print("  (Searching val frames we have scene graphs for...)")

candidates = []
for (sname, fidx), n_filtered in sg_results.items():
    n_qa = len(frame_qa.get((sname, fidx), []))
    if 5 <= n_qa <= 15 and 5 <= n_filtered <= 8:
        candidates.append((sname, fidx, n_qa, n_filtered))

candidates.sort(key=lambda x: x[2])  # sort by qa count

if candidates:
    print(f"\n  ✅ Found {len(candidates)} candidate(s) in existing SGs:")
    for sname, fidx, nqa, nf in candidates:
        print(f"    {sname} frame-{fidx}  val_qa={nqa}  filtered_nodes={nf}")
else:
    print("  No exact matches in existing SGs. Searching all val frames...")

    # Search all val frames for low-density
    print("\n  Top 15 low-density val frames by node count:")
    print(f"  {'scene:frame':<30} {'val_qa':>8} {'note'}")
    print("  " + "─"*55)

    # We need to know node counts — let's use val question count as proxy
    # and focus on scenes with small QA counts (likely low density)
    frame_counts = [(k, len(v)) for k,v in frame_qa.items()
                    if k[0] != "?" and k[1] >= 0 and 3 <= len(v) <= 12]
    frame_counts.sort(key=lambda x: x[1])

    shown = 0
    for (sname, fidx), nqa in frame_counts[:30]:
        # Check if we have the scene graph
        sg_path = SG_DIR / f"{sname}_frame{fidx}_scene_graph.json"
        has_sg = sg_path.exists()
        if has_sg:
            n_f, _, _, _ = load_sg_and_filter(sg_path)
            tag = f"✅ SG exists, filtered_nodes={n_f}"
        else:
            n_f = -1
            tag = "⬜ no SG yet"
        print(f"  {sname} frame-{fidx:<5} val_qa={nqa:3d}  {tag}")
        shown += 1
        if shown >= 15:
            break

# Also look specifically at scene-0103 which we have SGs for
print(f"\n  scene-0103 frames with val QA:")
for fidx in range(0, 42):
    key = ("scene-0103", fidx)
    nqa = len(frame_qa.get(key, []))
    if nqa > 0:
        sg_path = SG_DIR / f"scene-0103_frame{fidx}_scene_graph.json"
        if sg_path.exists():
            nf, fnodes, _, _ = load_sg_and_filter(sg_path)
            node_ids = [n.get("unique_id","?") for n in fnodes]
            print(f"    frame-{fidx:2d}: val_qa={nqa}  filtered_nodes={nf}  ids={node_ids}")
        else:
            print(f"    frame-{fidx:2d}: val_qa={nqa}  [no SG]")

print("\n✅ Diagnostic complete")
