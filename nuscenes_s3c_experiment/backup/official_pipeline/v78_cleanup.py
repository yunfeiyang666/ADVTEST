#!/usr/bin/env python3
"""V7.8 Environment Cleanup + New Sheet Discovery"""
import os, pathlib, sys, json, time
sys.path.insert(0, str(pathlib.Path(__file__).parent))
import openpyxl

EXCEL_DIR  = pathlib.Path("E:/Project/ADVTEST")
EXCEL_PATH = EXCEL_DIR / "RQ.xlsx"
DIRTY_KEYWORDS = ["DIAGNOSTIC_TEST", "DIAG_SCENE", "smoke_001", "TEST_SYNC",
                  "diag_baseline_001", "diag_gen_001", "diag_001"]

print("=" * 65)
print("  V7.8 Environment Cleanup")
print("=" * 65)

# ── Step 1: Find and kill lock files ──────────────────────────────────────────
print("\n[1] Scanning for ~$ lock files in E:\\Project\\ADVTEST\\...")
lock_found = []
for p in EXCEL_DIR.rglob("~$*.xlsx"):
    lock_found.append(p)
    print(f"  Found lock file: {p}")
    try:
        p.unlink()
        print(f"  ✅ Deleted: {p}")
    except Exception as e:
        print(f"  ⚠️  Could not delete: {e}")

if not lock_found:
    print("  ✅ No ~$ lock files found")

# ── Step 2: Test write access ─────────────────────────────────────────────────
print("\n[2] Testing write access to RQ.xlsx...")
for attempt in range(3):
    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        # Just open and close without modifying
        sheet_names = wb.sheetnames
        wb.close()
        print(f"  ✅ File accessible (attempt {attempt+1})")
        break
    except PermissionError as e:
        print(f"  ⚠️  Permission denied (attempt {attempt+1}): {e}")
        time.sleep(2)
    except Exception as e:
        print(f"  ❌ Other error: {e}")
        break

# ── Step 3: Scan ALL sheets + discover new sheets ─────────────────────────────
print("\n[3] Scanning all sheets in RQ.xlsx...")
try:
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    print(f"  Total sheets: {len(all_sheets)}")
    print(f"  Sheet names : {all_sheets}")
    
    known_sheets = {"raw_coverage", "question-answer-our",
                    "model_performance_raw_our", "RQ2_graph", "RQ3_graph", "RQ3_table"}
    new_sheets = [s for s in all_sheets if s not in known_sheets]
    
    if new_sheets:
        print(f"\n  🆕 NEW sheets found: {new_sheets}")
        for sname in new_sheets:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True, max_row=3))
            if rows and any(v is not None for v in rows[0]):
                header = [str(v) for v in rows[0] if v is not None]
                fmt    = [str(v) for v in rows[1] if v is not None] if len(rows) > 1 else []
                print(f"\n  Sheet [{sname}] — {len(header)} cols:")
                for i, (h, f) in enumerate(zip(header, fmt + [""]*len(header)), 1):
                    print(f"    col{i:02d}: '{h}'  ← {f[:60]}")
    else:
        print("  No new sheets beyond the known set")
    
    # Show all sheet schemas
    print("\n[Full schema of ALL sheets]")
    for sname in all_sheets:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True, max_row=2))
        if rows and any(v is not None for v in rows[0]):
            cols = [str(v) for v in rows[0] if v is not None]
            n_data = ws.max_row - 2  # subtract header rows
            print(f"  [{sname}] {len(cols)} cols, ~{max(0,n_data)} data rows: {cols}")
    
    wb.close()
except Exception as e:
    print(f"  ❌ Cannot read Excel: {e}")
    all_sheets = []
    new_sheets = []

# ── Step 4: Clean dirty rows ─────────────────────────────────────────────────
print("\n[4] Cleaning dirty rows from all sheets...")
dirty_cleaned = 0
try:
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows_to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_text = " ".join(str(v) for v in row if v is not None)
            if any(kw.lower() in row_text.lower() for kw in DIRTY_KEYWORDS):
                rows_to_delete.append(row_idx)
        
        # Delete in reverse order to maintain indices
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
            dirty_cleaned += 1
            print(f"  Deleted row {row_idx} from [{sname}]")
        
        if rows_to_delete:
            print(f"  [{sname}]: removed {len(rows_to_delete)} dirty rows")
        else:
            print(f"  [{sname}]: clean ✅")
    
    if dirty_cleaned > 0:
        wb.save(str(EXCEL_PATH))
        print(f"\n  ✅ Saved: removed {dirty_cleaned} dirty rows total")
    else:
        print(f"\n  ✅ No dirty rows found")
    wb.close()

except PermissionError:
    print(f"  ❌ Permission denied — Excel still locked. Kill Excel processes first.")
except Exception as e:
    print(f"  ❌ Error: {e}")

# ── Step 5: Verify filtered_scene_graphs ─────────────────────────────────────
print("\n[5] Verifying filtered_scene_graphs/...")
fsg_dir = pathlib.Path("E:/Project/ADVTEST/filtered_scene_graphs")
if fsg_dir.exists():
    files = sorted(fsg_dir.glob("*.json"))
    print(f"  Path    : {fsg_dir.resolve()}")
    print(f"  Files   : {len(files)}")
    target = fsg_dir / "scene-0926_frame20_scene_graph.json"
    print(f"  Target  : {'✅ EXISTS' if target.exists() else '❌ MISSING'} — {target.name}")
    for f in files:
        data = json.loads(f.read_text(encoding="utf-8"))
        info = data.get("core_universe_filter", {})
        print(f"    {f.name}: {info.get('filtered_nodes','?')} nodes, "
              f"{info.get('filtered_edges','?')} edges")
else:
    print(f"  ❌ Directory does not exist: {fsg_dir}")

print("\n✅ Cleanup complete")
