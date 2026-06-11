#!/usr/bin/env python3
"""Read RQ.xlsx to extract exact schema, then check data completeness."""
import openpyxl, json, pathlib, sys, os
sys.path.insert(0, str(pathlib.Path(__file__).parent))

# ── 1. Read Excel ─────────────────────────────────────────────────────────────
EXCEL_PATH = pathlib.Path("E:/Project/ADVTEST/RQ.xlsx")
print("=" * 70)
print(f"  Reading: {EXCEL_PATH}")
print("=" * 70)

wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
print(f"\nSheets found: {wb.sheetnames}\n")

schema = {}
for shname in wb.sheetnames:
    ws = wb[shname]
    print(f"{'─'*70}")
    print(f"  Sheet: [{shname}]  (max_row={ws.max_row}, max_col={ws.max_column})")
    print(f"{'─'*70}")
    rows_seen = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(v is not None for v in row):
            rows_seen.append(row)
            vals = [str(v)[:60] if v is not None else "—" for v in row]
            print(f"  row{i+1:03d}: {vals}")
        if i >= 60:
            print("  ... (more rows)")
            break
    if rows_seen:
        schema[shname] = rows_seen
wb.close()

# ── 2. Save schema as JSON ────────────────────────────────────────────────────
out = pathlib.Path("output/rq_excel_schema.json")
out.parent.mkdir(parents=True, exist_ok=True)
# Convert to serializable format
serial = {}
for sh, rows in schema.items():
    serial[sh] = [[str(v) if v is not None else None for v in row] for row in rows]
out.write_text(json.dumps(serial, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"\nSchema saved → {out}")

# ── 3. Check NuScenes data completeness ───────────────────────────────────────
print("\n" + "=" * 70)
print("  Data Completeness Check")
print("=" * 70)

checks = [
    # (description, path, required)
    ("NuScenes val QA",     "E:/Project/ADVTEST/data/nuscenes/qa/NuScenes_val_questions.json",    True),
    ("NuScenes train QA",   "E:/Project/ADVTEST/data/nuscenes/qa/NuScenes_train_questions.json",  True),
    ("trainval scene.json", "E:/Project/ADVTEST/data/nuscenes/v1.0-trainval/scene.json",          True),
    ("trainval sample.json","E:/Project/ADVTEST/data/nuscenes/v1.0-trainval/sample.json",         True),
    ("trainval sensor.json","E:/Project/ADVTEST/data/nuscenes/v1.0-trainval/sensor.json",         False),
    ("mini scene.json",     "E:/Project/ADVTEST/data/nuscenes/v1.0-mini/scene.json",              False),
    ("mini sample.json",    "E:/Project/ADVTEST/data/nuscenes/v1.0-mini/sample.json",             False),
    ("SG scene-0553 f8",    "E:/Project/ADVTEST/nuscenes_s3c_experiment/output/coverage_analysis/scene_graphs/scene-0553_frame8_scene_graph.json", True),
    ("SG scene-0916 f8",    "E:/Project/ADVTEST/nuscenes_s3c_experiment/output/coverage_analysis/scene_graphs/scene-0916_frame8_scene_graph.json", True),
    ("SG scene-0926",       "E:/Project/ADVTEST/nuscenes_s3c_experiment/output/coverage_analysis/scene_graphs/scene-0926_frame20_scene_graph.json", False),
    ("MetaVQA train",       "E:/Project/ADVTEST/MetaVQA/MetaVQA-Train/updated_trainval.json",     False),
    ("NuScenes images/samples", "E:/Project/ADVTEST/data/nuscenes/samples", False),
]

missing_required = []
for desc, path, required in checks:
    p = pathlib.Path(path)
    if p.exists():
        size = p.stat().st_size if p.is_file() else sum(f.stat().st_size for f in p.rglob("*") if f.is_file())
        size_str = f"{size//1024//1024} MB" if size > 1024*1024 else f"{size//1024} KB"
        print(f"  ✅ {'[REQ]' if required else '[OPT]'} {desc:<35} {size_str}")
    else:
        tag = "❌ MISSING [REQUIRED]" if required else "⚠️  missing [optional]"
        print(f"  {tag} {desc:<35} {path}")
        if required:
            missing_required.append((desc, path))

# Check images for top candidate scenes
print("\n  Image directory check:")
samples_dir = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/samples")
if samples_dir.exists():
    cams = [d for d in samples_dir.iterdir() if d.is_dir()]
    total_imgs = sum(len(list(c.glob("*.jpg"))) for c in cams[:3])
    print(f"  ✅ samples/ exists, camera dirs: {[d.name for d in cams[:6]]}")
    print(f"     First 3 cam dirs have ~{total_imgs} images total")
else:
    print("  ❌ samples/ NOT FOUND — no scene images available for VLM evaluation")
    missing_required.append(("NuScenes scene images", str(samples_dir)))

# Check what scenes we can generate SGs for
print("\n  NuScenes-devkit check:")
try:
    from nuscenes import NuScenes
    print("  ✅ nuscenes-devkit importable")
except ImportError:
    print("  ❌ nuscenes-devkit NOT installed (needed to generate new scene graphs)")
    missing_required.append(("nuscenes-devkit package", "pip install nuscenes-devkit"))

print("\n" + "─" * 70)
if missing_required:
    print(f"  ❌ {len(missing_required)} REQUIRED items missing:")
    for desc, path in missing_required:
        print(f"     • {desc}")
        print(f"       → {path}")
else:
    print("  ✅ All required items present")
print("─" * 70)
