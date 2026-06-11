#!/usr/bin/env python3
"""Read back last 3 rows from key Excel sheets and display."""
import openpyxl, pathlib, json

EXCEL = pathlib.Path("E:/Project/ADVTEST/RQ.xlsx")

wb = openpyxl.load_workbook(str(EXCEL), read_only=True, data_only=True)

for shname in ["raw_coverage", "question-answer-our", "filter_record"]:
    if shname not in wb.sheetnames:
        continue
    ws = wb[shname]
    # Row 1 = headers, Row 2 = format desc, Row 3+ = data
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v) if v is not None else "" for v in rows[0]]
    data_rows = [r for r in rows[2:] if any(v is not None for v in r)]

    print(f"\n{'='*70}")
    print(f"  Sheet: [{shname}]  ({len(data_rows)} data rows total)")
    print(f"{'='*70}")
    print(f"  Columns ({len(headers)}): {headers}")
    print(f"\n  ── Last 3 rows ──")
    for r in data_rows[-3:]:
        print()
        for h, v in zip(headers, r):
            if v is not None:
                val_str = str(v)
                # Truncate long JSON strings for display
                if len(val_str) > 80:
                    val_str = val_str[:77] + "..."
                print(f"    {h:<40}: {val_str}")

wb.close()
print("\n✅ Read-back complete")
