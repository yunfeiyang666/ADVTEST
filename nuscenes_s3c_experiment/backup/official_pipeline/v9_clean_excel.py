#!/usr/bin/env python3
"""Step 0: Remove all scene-0926 rows from RQ.xlsx + show final state."""
import openpyxl, pathlib, subprocess, time

EXCEL = pathlib.Path("E:/Project/ADVTEST/RQ.xlsx")
KILL_TARGETS = ["scene-0926"]

print(f"Excel: {EXCEL.resolve()}")

# Kill any Excel lock
subprocess.run(
    ["powershell", "-Command",
     "Get-Process | Where-Object { $_.Name -match 'EXCEL|OfficeClickToRun' } | "
     "ForEach-Object { Stop-Process -Id $_.Id -Force -ErrorAction SilentlyContinue }"],
    capture_output=True, timeout=10
)
time.sleep(1)
lock = EXCEL.parent / f"~${EXCEL.name}"
if lock.exists():
    try:
        lock.unlink()
        print(f"Deleted lock: {lock}")
    except Exception as e:
        print(f"Could not delete lock: {e}")

# Clean
wb = openpyxl.load_workbook(str(EXCEL))
total_removed = 0
sheet_stats = {}

for shname in wb.sheetnames:
    ws = wb[shname]
    rows_before = ws.max_row
    to_del = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2:
            continue  # keep header + format row
        row_text = " ".join(str(v) for v in row if v is not None)
        if any(kw in row_text for kw in KILL_TARGETS):
            to_del.append(i)
    for i in reversed(to_del):
        ws.delete_rows(i)
    sheet_stats[shname] = len(to_del)
    total_removed += len(to_del)
    print(f"  [{shname}]: deleted {len(to_del)} rows")

wb.save(str(EXCEL))
wb.close()
print(f"\nTotal removed: {total_removed} rows → saved")

# Verify
wb2 = openpyxl.load_workbook(str(EXCEL), read_only=True, data_only=True)
for shname in ["raw_coverage", "question-answer-our", "filter_record"]:
    if shname not in wb2.sheetnames:
        continue
    ws2 = wb2[shname]
    data_rows = [r for r in ws2.iter_rows(min_row=3, values_only=True)
                 if any(v is not None for v in r)]
    print(f"  [{shname}] data rows after clean: {len(data_rows)}")
wb2.close()

print("\n✅ Excel cleaned and ready")
