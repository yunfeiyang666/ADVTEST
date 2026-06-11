"""读取RQ.xlsx最新10行并格式化显示，供用户核验"""
import sys, pathlib, json
sys.path.insert(0, str(pathlib.Path(__file__).parent))
import openpyxl
from datetime import datetime

EXCEL = "E:/Project/ADVTEST/RQ.xlsx"
SEP = "=" * 70

def _dt(a, b):
    fmt = '%Y-%m-%d %H:%M:%S.%f'
    try:
        return int((datetime.strptime(str(b), fmt) -
                    datetime.strptime(str(a), fmt)).total_seconds() * 1000)
    except:
        return -1

wb = openpyxl.load_workbook(EXCEL)

# ─── raw_coverage 最后5行 ────────────────────────────────────────────────────
print(SEP)
print("  raw_coverage — 最后5行（Baseline 审计）")
print(SEP)
ws = wb["raw_coverage"]
h = [c.value for c in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))
for i, row in enumerate(rows[-5:], 1):
    d = dict(zip(h, row))
    ts = str(d.get("timestamp_start",""))
    te = str(d.get("timestamp_end",""))
    dt = _dt(ts, te)
    l0 = d.get("L0","")
    l1 = d.get("L1","")
    l2 = d.get("L2","")
    cyph = str(d.get("question_cypher（llm生成的cypher）",""))
    print(f"\n[{i}] {d.get('nuscenes_qa_id','')}  type={d.get('question_type','')}")
    print(f"  Q: {str(d.get('question',''))[:75]}")
    print(f"  A: {d.get('answer','')}")
    print(f"  ts_start = {ts}")
    print(f"  ts_end   = {te}   Δt={dt}ms")
    print(f"  L0 = {str(l0)[:80]}")
    print(f"  L1 = {str(l1)[:80]}")
    print(f"  L2 = {str(l2)[:80]}")
    print(f"  cypher_question len={len(cyph)} chars  preview: {cyph[:60]}")
    # 合格性判断
    l0_ok = bool(l0 and l0 != "[]" and l0 != "")
    dt_ok = dt >= 2000
    print(f"  >>> L0_OK={l0_ok}  Δt_OK(>=2000ms)={dt_ok}")

# ─── question-answer-our 最后5行 ─────────────────────────────────────────────
print(f"\n{SEP}")
print("  question-answer-our — 最后5行（Generated 生成）")
print(SEP)
ws2 = wb["question-answer-our"]
h2 = [c.value for c in ws2[1]]
rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
for i, row in enumerate(rows2[-5:], 1):
    d = dict(zip(h2, row))
    ts  = str(d.get("timestamp_start",""))
    tl  = str(d.get("timestamp_llm",""))
    tc  = str(d.get("timestamp_cypher_return",""))
    te  = str(d.get("timestamp_end",""))
    dt_total = _dt(ts, te)
    dt_llm   = _dt(ts, tl)
    dt_cyph  = _dt(ts, tc)
    itr = d.get("iteration_count", "?")
    cyph = str(d.get("cypher question",""))
    l0 = str(d.get("L0",""))
    l1 = str(d.get("L1",""))
    l2 = str(d.get("L2",""))
    q  = str(d.get("natural language question",""))
    print(f"\n[{i}] {d.get('gap_cell','')}  type={d.get('question_type','')}")
    print(f"  Q: {q[:75]}")
    print(f"  A: {d.get('answer','')}")
    print(f"  ① ts_start         = {ts}")
    print(f"  ② ts_llm           = {tl}   Δ={dt_llm:+d}ms {'← 独立!' if dt_llm > 0 else '← 同时刻'}")
    print(f"  ③ ts_cypher_return = {tc}   Δ={dt_cyph:+d}ms from start")
    print(f"  ④ ts_end           = {te}   Δ={dt_total}ms ← 总耗时")
    print(f"  iteration_count = {itr}")
    print(f"  L0={l0[:60]}")
    print(f"  L1={l1[:60]}")
    print(f"  L2={l2[:60]}")
    print(f"  cypher_question: {cyph[:80].strip()}")
    # 合格性判断
    ts_same = ts == tl
    iter_ok = isinstance(itr, int) and itr >= 1
    dt_ok   = dt_total >= 2000
    cyph_ok = len(cyph) > 10
    l0_ok   = bool(l0 and l0 != "[]")
    print(f"  >>> ts_start==ts_llm={ts_same}  Δt_OK={dt_ok}  iter={itr}  cypher_OK={cyph_ok}  L0_OK={l0_ok}")

print(f"\n{SEP}")
print("  合格标准提醒：Δt>=2000ms | L0非空 | cypher有内容 | iteration_count真实")
print(SEP)
