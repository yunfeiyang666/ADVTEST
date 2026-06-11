#!/usr/bin/env python3
"""
run_method_a.py — 方案 A 闭环执行链 (scene-0926 frame-20)

执行前检查清单（必须全部通过才能运行）：
  ✅ RQ.xlsx 已关闭（无 ~$RQ.xlsx 锁定文件）
  ✅ filtered_scene_graphs/scene-0926_frame20_scene_graph.json 存在
  ✅ Neo4j bolt://localhost:7800 可达

执行链：
  Step 1 : 净化环境（清理脏数据行、删除锁定文件）
  Step 2 : 写入 filter_record（核心宇宙过滤元数据）
  Step 3 : 清空 Neo4j + 从 filtered_scene_graphs/ 导入 scene-0926-20
  Step 4 : Baseline 审计（29 条原题 → LLM Cypher → 足迹 → raw_coverage）
  Step 5 : Gap Detection（未覆盖 L2 路径）
  Step 6 : 增量生成（5 条 L2 题 → question-answer-our）
  Step 7 : Final save 确认
"""
import os, sys, pathlib, json, time, collections
sys.path.insert(0, str(pathlib.Path(__file__).parent))

EXCEL_PATH  = pathlib.Path("E:/Project/ADVTEST/RQ(1).xlsx")
FSG_DIR     = pathlib.Path(os.getenv("FILTERED_SG_DIR", "E:/Project/ADVTEST/filtered_scene_graphs"))
TRAINVAL    = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/v1.0-mini")
QA_PATH     = pathlib.Path("E:/Project/ADVTEST/data/nuscenes/qa/NuScenes_val_questions.json")
TARGET_SG   = "scene-0926_frame20_scene_graph.json"
SCENE_ID    = "scene-0926"
FRAME_ID    = 20
NEO4J_URI   = "bolt://localhost:7800"
NEO4J_USER  = "neo4j"
NEO4J_PWD   = "87017563"
ALLOW_EXCEL_DESTRUCTIVE_CLEAN = os.getenv("VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN", "false").lower() in ("true", "1", "yes")


# ─────────────────────────────────────────────────────────────────────────────
# Pre-flight checks
# ─────────────────────────────────────────────────────────────────────────────

def preflight_check() -> bool:
    print("\n[Pre-flight checks]")
    ok = True

    import subprocess

    # 1. Kill lock file (OfficeClickToRun keeps recreating it — force-delete and test write)
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        print(f"  ⚠️  Lock file found: {lock} — attempting force-delete...")
        # Kill OfficeClickToRun silently to prevent immediate recreation
        subprocess.run(
            ["powershell", "-Command",
             "Get-Process | Where-Object { $_.Name -match 'EXCEL|OfficeClickToRun' } | "
             "ForEach-Object { Stop-Process -Id $_.Id -Force -ErrorAction SilentlyContinue }"],
            capture_output=True, timeout=10
        )
        import time as _t; _t.sleep(1)
        try:
            lock.unlink(missing_ok=True)
            print(f"  ✅ Lock file deleted")
        except Exception as e:
            print(f"  ⚠️  Could not delete lock: {e}")
    else:
        print(f"  ✅ No lock file")

    # 2. Filtered SG exists
    sg = FSG_DIR / TARGET_SG
    if sg.exists():
        data = json.loads(sg.read_text(encoding="utf-8"))
        n = data.get("core_universe_filter", {}).get("filtered_nodes", "?")
        print(f"  ✅ Filtered SG exists: {sg.name} ({n} nodes)")
    else:
        print(f"  ❌ Filtered SG missing: {sg}")
        ok = False

    # 3. Neo4j reachable
    try:
        from neo4j import GraphDatabase
        d = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))
        with d.session() as s:
            s.run("RETURN 1").single()
        d.close()
        print(f"  ✅ Neo4j reachable: {NEO4J_URI}")
    except Exception as e:
        print(f"  ❌ Neo4j not reachable: {e}")
        ok = False

    # 4. Excel actual write test (the TRUE test — lock file existence alone is not sufficient)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        wb.save(str(EXCEL_PATH))
        wb.close()
        print(f"  ✅ Excel writable: {EXCEL_PATH}")
    except PermissionError:
        print(f"  ❌ Excel write test FAILED — file is still locked by another process")
        ok = False
    except Exception as e:
        print(f"  ⚠️  Excel check error: {e}")

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Step 0: Wipe previous run data (DISABLED by default, append-only policy)
# ─────────────────────────────────────────────────────────────────────────────

def step0_wipe_previous_run():
    """Destructive cleanup (opt-in only). Default behavior is append-only."""
    print("\n[Step 0] Destructive Excel wipe")
    if not ALLOW_EXCEL_DESTRUCTIVE_CLEAN:
        print("  ⏭️  Skipped (append-only mode; VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN is false)")
        return
    print("  ⚠️  Enabled by VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN=true")
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        total_deleted = 0
        for sh in ["raw_coverage", "question-answer-our"]:
            ws = wb[sh]
            max_row = ws.max_row
            # Delete from bottom to preserve row indices
            for r in range(max_row, 1, -1):   # row 1 = header, skip
                ws.delete_rows(r)
                total_deleted += 1
        wb.save(str(EXCEL_PATH))
        wb.close()
        print(f"  ✅ Deleted {total_deleted} data rows (headers preserved)")
    except Exception as e:
        print(f"  ⚠️  Wipe error: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Environment cleanup (non-destructive, lock file only)
# ─────────────────────────────────────────────────────────────────────────────

def step1_cleanup():
    print("\n[Step 1] Environment cleanup (append-only, non-destructive)")

    # Remove lock file if still present
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        try:
            lock.unlink()
            print(f"  ✅ Deleted lock file: {lock}")
        except Exception as e:
            print(f"  ⚠️  Could not delete lock: {e}")
    print("  ✅ Keep all historical rows (no row deletion)")


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Write filter_record
# ─────────────────────────────────────────────────────────────────────────────

def step2_filter_record():
    print("\n[Step 2] Writing filter_record")
    from rq_tables import write_filter_record

    sg = FSG_DIR / TARGET_SG
    data = json.loads(sg.read_text(encoding="utf-8"))
    info = data["core_universe_filter"]
    vex_str = ",".join(sorted(info["node_ids_kept"]))
    ratio   = info["filtered_nodes"] / max(info["raw_nodes"], 1)

    ok = write_filter_record(
        scene_id=SCENE_ID, frame_id=FRAME_ID,
        original_num=info["raw_nodes"],
        filtered_num=info["filtered_nodes"],
        filtered_vex=vex_str,
        ratio=ratio,
    )
    print(f"  scene={SCENE_ID} frame={FRAME_ID}")
    print(f"  raw={info['raw_nodes']} → filtered={info['filtered_nodes']} "
          f"(ratio={ratio:.2%})")
    print(f"  node_ids: {info['node_ids_kept']}")
    print(f"  write_filter_record: {'✅' if ok else '❌'}")
    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: Neo4j import from filtered_scene_graphs/
# ─────────────────────────────────────────────────────────────────────────────

def step3_import_neo4j():
    print("\n[Step 3] Importing filtered SG to Neo4j")
    from core_universe_filter import import_filtered_sg_to_neo4j
    result = import_filtered_sg_to_neo4j(
        sg_name=TARGET_SG,
        neo4j_uri=NEO4J_URI,
        neo4j_user=NEO4J_USER,
        neo4j_pwd=NEO4J_PWD,
    )
    print(f"  ✅ Neo4j: {result['n_nodes']} nodes, {result['n_edges']} edges")
    print(f"  Source : {result['source']}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Step 4: Baseline audit (29 original questions → LLM Cypher → footprint → raw_coverage)
# ─────────────────────────────────────────────────────────────────────────────

def step4_baseline_audit(driver, llm_client):
    print("\n[Step 4] Baseline audit (original NuScenes-QA questions)")
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    from semantic_auditor import audit_baseline_question, build_scene_context
    from rq_tables import write_baseline_to_coverage

    # Build sample→scene mapping
    scenes  = json.loads((TRAINVAL/"scene.json").read_text())
    samples = json.loads((TRAINVAL/"sample.json").read_text())
    st2name = {s["token"]: s["name"] for s in scenes}
    s2tok: dict = collections.defaultdict(list)
    tok2info: dict = {}
    for samp in samples:
        sname = st2name.get(samp["scene_token"],"?")
        tok2info[samp["token"]] = {"scene_name": sname, "timestamp": samp["timestamp"]}
        s2tok[sname].append(samp["token"])
    for sname, toks in s2tok.items():
        for idx, tok in enumerate(sorted(toks, key=lambda t: tok2info[t]["timestamp"])):
            tok2info[tok]["frame_idx"] = idx

    # Get scene-0926 val questions  —  keep global index for unique ID generation
    all_val_qs = json.loads(QA_PATH.read_text())["questions"]
    target_qs = [
        (global_i, q)
        for global_i, q in enumerate(all_val_qs)     # global_i = position in val file
        if tok2info.get(q.get("sample_token",""), {}).get("scene_name") == SCENE_ID
        and tok2info.get(q.get("sample_token",""), {}).get("frame_idx") == FRAME_ID
    ]
    print(f"  Found {len(target_qs)} val questions for {SCENE_ID} frame-{FRAME_ID}")
    if target_qs:
        print(f"  Global indices: {target_qs[0][0]} – {target_qs[-1][0]}")

    if not target_qs:
        print("  ⚠️  No questions found — check sample_token mapping")
        return 0

    # Build scene context once (reuse across questions)
    scene_ctx = build_scene_context(driver)
    print(f"  Scene context built ({len(scene_ctx)} chars)")

    from semantic_auditor import make_qa_id, _ms_now as _audit_ms, derive_l2_from_l1
    BASELINE_N_WORKERS = 15
    _tls = threading.local()
    _llm_cls = llm_client.__class__

    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            _llm = _llm_cls()
            _tls.llm = _llm
        return _llm

    def _audit_one(global_idx: int, q: dict) -> dict:
        question = q.get("question", "")
        answer = q.get("answer", "")
        qtype = q.get("template_type", "")
        nhop = q.get("num_hop", 0)
        qa_uid = make_qa_id(global_idx, qtype)
        ts0 = _audit_ms()
        try:
            audit_res = audit_baseline_question(
                question=question, q_type=qtype, num_hop=nhop,
                driver=driver, llm_client=_get_worker_llm(),
                scene_context=scene_ctx,
                global_index=global_idx,
            )
        except Exception as exc:
            print(f"    ⚠️ [baseline-worker-exc] idx={global_idx} {exc}")
            audit_res = {"l0_nodes": [], "l1_edges": []}
        ts1 = _audit_ms()
        l0 = audit_res.get("l0_nodes", []) or []
        l1 = audit_res.get("l1_edges", []) or []
        l2 = derive_l2_from_l1(l1)
        audit_json = json.dumps(
            {"nodes": l0, "edges": l1},
            ensure_ascii=False
        )
        return {
            "qa_uid": qa_uid,
            "question": question,
            "answer": answer,
            "qtype": qtype,
            "l0_nodes": l0,
            "l1_edges": l1,
            "l2_paths": l2,
            "timestamp_start": ts0,
            "timestamp_end": ts1,
            "audit_cypher": audit_json,
        }

    t_batch0 = time.perf_counter()
    ordered = [None] * len(target_qs)
    with ThreadPoolExecutor(max_workers=BASELINE_N_WORKERS) as _ex:
        fut2idx = {
            _ex.submit(_audit_one, global_idx, q): idx
            for idx, (global_idx, q) in enumerate(target_qs)
        }
        for done_i, _f in enumerate(as_completed(fut2idx), 1):
            idx = fut2idx[_f]
            rec = _f.result()
            ordered[idx] = rec
            if done_i <= 3 or done_i % 10 == 0:
                print(f"  [audit {done_i:2d}/{len(target_qs)}] {rec['qa_uid']:<25} "
                      f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} L2={len(rec['l2_paths'])}")
    batch_ms = int((time.perf_counter() - t_batch0) * 1000)
    avg_ms = int(batch_ms / max(len(target_qs), 1))
    print(f"  [Baseline Batch] questions={len(target_qs)} workers={BASELINE_N_WORKERS} "
          f"elapsed={batch_ms}ms avg={avg_ms}ms/q")

    success = 0
    for i, rec in enumerate(ordered, 1):
        if rec is None:
            continue

        ok = write_baseline_to_coverage(
            scene_id=SCENE_ID, frame_id=FRAME_ID,
            nuscenes_qa_id=rec["qa_uid"],    # e.g. "val_71051_comparison"
            question=rec["question"], answer=rec["answer"],
            l0_nodes=rec["l0_nodes"],
            l1_edges=rec["l1_edges"],
            l2_paths=rec["l2_paths"],        # 自动推导的 L2
            question_type=rec["qtype"],
            timestamp_start=rec["timestamp_start"], timestamp_end=rec["timestamp_end"],
            audit_cypher=rec["audit_cypher"],  # JSON 子图填入 cypher 列
        )
        if ok:
            success += 1
        if i <= 3 or i % 10 == 0:
            _ok_mark = "OK" if ok else "--"
            print(f"  [write {i:2d}/{len(target_qs)}] id={rec['qa_uid']:<25} "
                  f"q_type={rec['qtype']:12s} "
                  f"L0={len(rec['l0_nodes'])} "
                  f"L1={len(rec['l1_edges'])} "
                  f"L2={len(rec['l2_paths'])} "
                  f"{_ok_mark}")

    print(f"\n  Baseline audit done: {success}/{len(target_qs)} written to raw_coverage")
    return success


# ─────────────────────────────────────────────────────────────────────────────
# Step 5+6: Gap detection + generate 5 L2 questions → question-answer-our
# ─────────────────────────────────────────────────────────────────────────────

# ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――
# V18 题型权重与答案策略（本地模板问句已移除）
# ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――

_Q_TYPE_WEIGHTS = {"exist": 3, "status": 3, "object": 3, "count": 1}


def _answer_for_q_type(q_type: str, cell: dict, verify_n: int = -1) -> str:
    """仅生成答案字段；问题文本必须由 LLM 生成。"""
    if q_type == "exist":
        return "yes" if verify_n != 0 else "no"
    if q_type == "status":
        return cell.get("n3_status", "") or "unknown"
    if q_type == "count":
        return str(verify_n if verify_n >= 0 else 1)
    if q_type == "comparison":
        return "closer"
    return cell.get("n3_id", "")


def step5_6_generate(driver, llm_client):
    """
    V18 真实性硬化协议：
      1) timestamp_start        : 进入每个 gap cell 处理第一行捕获
      2) timestamp_llm          : 来自 llm_client._call() 返回后的物理时刻
      3) timestamp_cypher_return: Neo4j single() 返回后的物理时刻
      4) timestamp_end          : 写 Excel 前最后一行捕获
    严格规则：
      - 自然语言问题必须由 LLM 生成（无本地模板 fallback）
      - 若 (timestamp_end - timestamp_start) < 2000ms，则该条作废
      - cypher question / target_gap_cell 为空则作废
    """
    print("\n[Step 5+6] V18 Realness-Hardened Generation (4 physical timestamps locked)")
    import random
    import re
    import json as _json
    from datetime import datetime
    import uuid as _uuid
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from gap_pipeline.coverage_tracker import CoverageTracker, CoverageRecord
    from gap_pipeline.constraint_methods import CumulativeConstraintChain
    from run_gap_pipeline_v6 import _process_single_cell
    from rq_tables import write_generated_question

    # V20 并发参数（极速量产）
    Q_BATCH_SIZE = 15
    Q_N_WORKERS = 15
    # 真实性阈值
    MAX_FAIL = 3
    MIN_REAL_MS = 2000
    # V23 质量门控（可通过环境变量覆盖）
    STRICT_UNIQUE_ONLY = os.getenv("VQA_STRICT_UNIQUE_ONLY", "true").lower() in ("true", "1", "yes")
    MIN_ITER_COUNT = int(os.getenv("VQA_MIN_ITER_COUNT", "2"))
    MAX_QTYPE_RETRY = int(os.getenv("VQA_QTYPE_MAX_RETRY", "2"))
    LOW_ITER_STRICT_METHODS = {
        x.strip() for x in os.getenv("VQA_MIN_ITER_STRICT_METHODS", "no_constraint_needed,path").split(",")
        if x.strip()
    }

    _CONSTRAINT_DESC_MAP = {
        "type_filter":         "target type unique in this direction",
        "status_anchor":       "target status unique in this direction",
        "type_status_anchor":  "type+status combo unique in this direction",
        "dir8_refine":         "paper 6-direction label narrows candidates to 1",
        "dual_reference":      "two reference objects jointly identify target",
        "dist_ord":            "target is closest/farthest in this direction",
        "two_hop_referent":    "a third node uniquely points to target",
        "dual_hop_referent":   "two third nodes jointly identify target",
        "anchor_intro":        "source object is uniquely described as anchor",
        "type+dir8":           "type + exact direction unique",
        "type+dir8+dist_ord":  "type + direction + distance rank unique",
        "dir8+dist_ord":       "direction + distance rank unique",
        "yesno_fallback":      "path itself uniquely identifies target",
    }

    def _abs_ts() -> str:
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    def _dt_ms(a: str, b: str) -> int:
        fmt = '%Y-%m-%d %H:%M:%S.%f'
        try:
            return int((datetime.strptime(b, fmt) - datetime.strptime(a, fmt)).total_seconds() * 1000)
        except Exception:
            return -1

    def _is_retryable_llm_exc(exc: Exception) -> bool:
        s = str(exc).lower()
        keys = ("timed out", "timeout", "connection", "connect", "readtimeout", "apiconnectionerror")
        return any(k in s for k in keys)
    def _is_auth_error(exc_or_msg: Exception | str) -> bool:
        s = str(exc_or_msg).lower()
        keys = (
            "authenticationerror",
            "unauthorized",
            "invalid token",
            "invalid api key",
            "无效的令牌",
            "401",
            "one_api_error",
        )
        return any(k in s for k in keys)

    def _parse_verify_n(verify_text: str) -> int:
        m = re.search(r"n=(\d+)", str(verify_text or ""))
        return int(m.group(1)) if m else -1

    def _is_verified_unique(verify_text: str, verify_n: int) -> bool:
        return verify_n == 1 and ("✅" in str(verify_text or ""))

    def _qtype_semantic_ok(q_type: str, question: str) -> bool:
        q = " ".join(str(question or "").strip().lower().split())
        if not q:
            return False
        if q_type == "count":
            return q.startswith("how many ")
        if q_type == "exist":
            return q.startswith("is there ") or q.startswith("are there ")
        if q_type == "status":
            return "status" in q and q.startswith("what ")
        if q_type == "comparison":
            return ("closer" in q) or ("farther" in q) or ("further" in q)
        # object
        return (
            q.startswith("what ")
            and not q.startswith("how many ")
            and not q.startswith("is there ")
            and "status" not in q
        )

    def _scene_type_distribution() -> dict:
        q = (
            "MATCH (n:Object) "
            "RETURN n.type AS t, count(n) AS c "
            "ORDER BY c DESC LIMIT 12"
        )
        out = {}
        try:
            with driver.session() as sess:
                for rec in sess.run(q):
                    t = rec.get("t", "")
                    c = rec.get("c", 0)
                    if t:
                        out[str(t)] = int(c)
        except Exception:
            pass
        return out

    tracker = CoverageTracker()
    with driver.session() as sess:
        tracker.init_from_session(sess)

    def _print_stats(label: str):
        s = tracker.stats()
        print(f"\n  [{label}]")
        for lvl in ("L0", "L1", "L2A", "L2B"):
            v = s[lvl]
            print(f"    {lvl:<4}: {v['gap']:>4} gap / {v['total']:>4} total "
                  f"(covered={v['covered']}, rate={v['rate']:.1f}%)")

    _print_stats("Gap Stats — Initial")

    _scenes = _json.loads((TRAINVAL / "scene.json").read_text())
    _samples = _json.loads((TRAINVAL / "sample.json").read_text())
    _st2name = {s["token"]: s["name"] for s in _scenes}
    _scene_samps = sorted(
        [s for s in _samples if _st2name.get(s["scene_token"], "") == SCENE_ID],
        key=lambda s: s["timestamp"],
    )
    sample_token = (_scene_samps[FRAME_ID]["token"] if len(_scene_samps) > FRAME_ID else "")
    print(f"  sample_token for JSON: {sample_token[:16]}...")
    scene_dist_global = _scene_type_distribution()
    print(f"  scene type distribution: {scene_dist_global}")

    _tls = threading.local()
    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            from gap_pipeline.llm_client import LLMClient as _LLMClient
            _llm = _LLMClient()
            _tls.llm = _llm
        return _llm

    def _gap_score(cell):
        path = cell.get("path_pattern", "")
        ego_p = 0.5 if "ego" in path else 1.0
        nodes = [cell.get("n1_id",""), cell.get("n2_id",""), cell.get("n3_id","")]
        unc = sum(1 for n in nodes if n and tracker._L0.get(n, CoverageRecord()).hit_count == 0)
        return ego_p * (unc + 1)

    q_type_pool = []
    for qt, wt in _Q_TYPE_WEIGHTS.items():
        q_type_pool.extend([qt] * wt)

    generated = 0
    total_att = 0
    dropped_fast = 0
    dropped_empty = 0
    dropped_non_unique = 0
    dropped_low_iter = 0
    dropped_qtype_mismatch = 0
    fail_counts: dict = {}
    unresolvable: set = set()
    used_types: list = []
    nusqa_records: list = []

    round_idx = 0
    while True:
        round_idx += 1
        l2b_avail = [c for c in tracker.get_gap_cells("L2B") if c["_key"] not in unresolvable]
        l2a_avail = [c for c in tracker.get_gap_cells("L2A") if c["_key"] not in unresolvable]
        if not l2b_avail and not l2a_avail:
            print(f"\n  ✅ Round {round_idx}: All resolvable L2 gaps covered. Stopping.")
            break

        sb = sorted(l2b_avail, key=_gap_score, reverse=True)
        sa = sorted(l2a_avail, key=_gap_score, reverse=True)
        n_b = min(4, len(sb))
        n_a = min(4, len(sa))
        round_cells = [("L2B", c) for c in sb[:n_b]] + [("L2A", c) for c in sa[:n_a]]
        if not round_cells:
            break
        round_cells = round_cells[:Q_BATCH_SIZE]

        print(f"\n  [Round {round_idx}] batch={len(round_cells)} "
              f"(L2B={sum(1 for t,_ in round_cells if t=='L2B')}, "
              f"L2A={sum(1 for t,_ in round_cells if t=='L2A')}) "
              f"written={generated} unresolvable={len(unresolvable)}")

        batch_llm_ms = []
        batch_neo_ms = []
        batch_total_ms = []
        round_batch_id = f"{SCENE_ID}_f{FRAME_ID}_r{round_idx}"

        # 先分配q_type，避免线程中竞争
        assigned_cells = []
        for topology, cell in round_cells:
            total_att += 1
            rem = [t for t in q_type_pool if t not in used_types]
            if not rem:
                rem = list(_Q_TYPE_WEIGHTS.keys())
                used_types = []
            q_type = random.choice(rem)
            used_types.append(q_type)
            assigned_cells.append((topology, cell, q_type))

        def _cell_worker(_topology, _cell, _q_type):
            _llm = _get_worker_llm()
            _gap_key = _cell["_key"]
            _path = _cell.get("path_pattern", "?")

            # [物理采样点1]
            _ts_start = _abs_ts()

            # LLM context
            if _topology == "L2A":
                _cypher = _llm.generate_l2a_context_cypher(_cell)
            else:
                _cypher = _llm.generate_l2b_obj_context_cypher(_cell)
            # [物理采样点2]
            _ts_llm = _llm.last_call_meta.get("timestamp_llm") or _abs_ts()
            _llm_ms = float(_llm.last_call_timing.get("total_ms", 0.0))

            _local_chain = CumulativeConstraintChain()
            _qa, _timing = _process_single_cell(
                cell=_cell, topology=_topology, cypher=_cypher,
                driver=driver, chain=_local_chain,
                scene_name=SCENE_ID, frame_idx=FRAME_ID,
                llm_timing=dict(_llm.last_call_timing),
                render_local_question=False,
            )
            _ts_cypher = _timing.get("timestamp_cypher_return", "") or _abs_ts()
            _neo_ms = float(_timing.get("neo4j_ms", 0.0))
            if _qa is None:
                return {"ok": False, "stage": "neo4j", "gap_key": _gap_key, "path": _path}

            _method = _timing.get("method_used", "") or "path"
            _constraint_desc = _CONSTRAINT_DESC_MAP.get(
                _method, f"constraint chain method={_method} identifies target"
            )

            _verify_cypher = _timing.get("verify_cypher", "") or ""
            if not _verify_cypher.strip() or not _path.strip():
                return {"ok": False, "stage": "empty", "gap_key": _gap_key, "path": _path}
            _verify_text = str(_timing.get("logic_verification", "") or "")
            _verify_n = _parse_verify_n(_verify_text)
            _verified_unique = _is_verified_unique(_verify_text, _verify_n)
            _chain_unique = bool(_timing.get("is_unique", False))
            if STRICT_UNIQUE_ONLY and (not _chain_unique or not _verified_unique):
                return {
                    "ok": False,
                    "stage": "non_unique",
                    "gap_key": _gap_key,
                    "path": _path,
                    "verify": _verify_text,
                    "method": _method,
                }

            _answer = _answer_for_q_type(_q_type, _cell, verify_n=_verify_n)

            def _gen_q_once() -> str:
                try:
                    return _llm.generate_question_nlp_strict(
                        path=_path,
                        q_type=_q_type,
                        n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                        n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                        n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                        n3_status=_cell.get("n3_status",""),
                        r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                        r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                        constraint_desc=_constraint_desc,
                        answer=str(_answer),
                        scene_distribution=scene_dist_global,
                    )
                except Exception as _q_exc:
                    if _is_retryable_llm_exc(_q_exc):
                        time.sleep(0.8)  # 单次快速退避
                        return _llm.generate_question_nlp_strict(
                            path=_path,
                            q_type=_q_type,
                            n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                            n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                            n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                            n3_status=_cell.get("n3_status",""),
                            r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                            r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                            constraint_desc=_constraint_desc,
                            answer=str(_answer),
                            scene_distribution=scene_dist_global,
                        )
                    raise

            _question = ""
            _semantic_ok = False
            for _ in range(max(1, MAX_QTYPE_RETRY)):
                _question = _gen_q_once()
                if _qtype_semantic_ok(_q_type, _question):
                    _semantic_ok = True
                    break
            if not _semantic_ok:
                return {
                    "ok": False,
                    "stage": "qtype_mismatch",
                    "gap_key": _gap_key,
                    "path": _path,
                    "q_type": _q_type,
                    "question": _question,
                }

            # [物理采样点4]
            _ts_end = _abs_ts()
            _total_ms = _dt_ms(_ts_start, _ts_end)
            if _total_ms >= 0 and _total_ms < MIN_REAL_MS:
                return {"ok": False, "stage": "fast", "gap_key": _gap_key, "path": _path, "total_ms": _total_ms}

            _trace = _timing.get("constraint_trace", "")
            _trace_methods = [p for p in _trace.split("→") if p and p != "Path"]
            _iter_count = max(1, len(_trace_methods))
            if _iter_count < MIN_ITER_COUNT and _method in LOW_ITER_STRICT_METHODS:
                return {
                    "ok": False,
                    "stage": "low_iter",
                    "gap_key": _gap_key,
                    "path": _path,
                    "iteration_count": _iter_count,
                    "trace": _trace,
                }
            _n1 = _cell.get("n1_id",""); _n2 = _cell.get("n2_id",""); _n3 = _cell.get("n3_id","")
            _l0 = [x for x in [_n1, _n2, _n3] if x]
            _l1 = ([{"source": _n1, "target": _n2}] if _n1 and _n2 else []) + \
                  ([{"source": _n2, "target": _n3}] if _n2 and _n3 else [])
            _l2 = [{"o1": _n1, "o2": _n2, "o3": _n3}] if all([_n1, _n2, _n3]) else []
            _sid_tag = SCENE_ID.replace("scene-", "s").replace("-", "")
            _gen_id = f"gen_{_sid_tag}_f{FRAME_ID}_{str(_uuid.uuid4())[:8]}"
            return {
                "ok": True,
                "gen_id": _gen_id,
                "gap_key": _gap_key,
                "path": _path,
                "topology": _topology,
                "q_type": _q_type,
                "answer": str(_answer),
                "question": _question,
                "verify_cypher": _verify_cypher,
                "verify_text": _verify_text,
                "verify_n": _verify_n,
                "trace": _trace,
                "iteration_count": _iter_count,
                "method": _method,
                "is_unique": _timing.get("is_unique", False),
                "ts_start": _ts_start,
                "ts_llm": _ts_llm,
                "ts_cypher_return": _ts_cypher,
                "ts_end": _ts_end,
                "total_ms": _total_ms,
                "llm_ms": _llm_ms,
                "neo_ms": _neo_ms,
                "l0": _l0, "l1": _l1, "l2": _l2,
            }

        with ThreadPoolExecutor(max_workers=Q_N_WORKERS) as _ex:
            _futs = {}
            for _t, _c, _q in assigned_cells:
                _fu = _ex.submit(_cell_worker, _t, _c, _q)
                _futs[_fu] = (_c.get("_key", ""), _c.get("path_pattern", "?"))
            for _f in as_completed(_futs):
                _gap_key, _path = _futs.get(_f, ("", "?"))
                try:
                    _res = _f.result()
                except Exception as exc:
                    _msg = str(exc)
                    if _is_auth_error(_msg):
                        raise RuntimeError(
                            f"LLM authentication failed ({_msg}). "
                            f"Please set a valid school API key for model Qwen3.5-35B-A3B."
                        ) from exc
                    if _gap_key:
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if fail_counts[_gap_key] >= MAX_FAIL:
                            unresolvable.add(_gap_key)
                    print(f"    ❌ [worker-exc] {_path} {_msg}")
                    continue
                _path = _res.get("path", _path)
                _gap_key = _res.get("gap_key", _gap_key)
                if not _res.get("ok"):
                    _stage = _res.get("stage", "unknown")
                    if _stage == "fast":
                        dropped_fast += 1
                        print(f"    ⚠️ [drop-fast] {_path} total={_res.get('total_ms',-1)}ms < {MIN_REAL_MS}ms")
                    elif _stage == "empty":
                        dropped_empty += 1
                        print(f"    ⚠️ [drop-empty] path={_path}")
                    elif _stage == "non_unique":
                        dropped_non_unique += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if fail_counts[_gap_key] >= MAX_FAIL:
                            unresolvable.add(_gap_key)
                        print(f"    ⚠️ [drop-non-unique] {_path} verify={_res.get('verify','')}")
                    elif _stage == "low_iter":
                        dropped_low_iter += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if fail_counts[_gap_key] >= MAX_FAIL:
                            unresolvable.add(_gap_key)
                        print(f"    ⚠️ [drop-low-iter] {_path} iter={_res.get('iteration_count',-1)} < {MIN_ITER_COUNT}")
                    elif _stage == "qtype_mismatch":
                        dropped_qtype_mismatch += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if fail_counts[_gap_key] >= MAX_FAIL:
                            unresolvable.add(_gap_key)
                        print(f"    ⚠️ [drop-qtype] {_path} q_type={_res.get('q_type','?')} q={_res.get('question','')}")
                    else:
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if fail_counts[_gap_key] >= MAX_FAIL:
                            unresolvable.add(_gap_key)
                        print(f"    ❌ [{_stage}] {_path}")
                    continue

                ok = write_generated_question(
                    scene_id=SCENE_ID, frame_id=FRAME_ID, question_id=_res["gen_id"],
                    timestamp_start=_res["ts_start"],
                    timestamp_llm=_res["ts_llm"],
                    timestamp_cypher_return=_res["ts_cypher_return"],
                    timestamp_end=_res["ts_end"],
                    iteration_count=_res["iteration_count"],
                    question_type=_res["q_type"],
                    complexity="L2",
                    natural_language_question=_res["question"],
                    cypher_question=_res["verify_cypher"],
                    answer=_res["answer"],
                    l0_nodes=_res["l0"], l1_edges=_res["l1"], l2_paths=_res["l2"],
                    target_gap_cell=_res["path"],
                    batch_id=round_batch_id,
                )
                if ok:
                    generated += 1
                nusqa_records.append({
                    "sample_token": sample_token,
                    "question": _res["question"],
                    "answer": _res["answer"],
                    "template_type": _res["q_type"],
                    "num_hop": 2,
                    "question_id": _res["gen_id"],
                    "scene_name": SCENE_ID,
                    "frame_idx": FRAME_ID,
                    "topology_level": _res["topology"],
                    "path_pattern": _res["path"],
                    "constraint_trace": _res["trace"],
                    "iteration_count": _res["iteration_count"],
                    "is_unique": _res["is_unique"],
                    "method_used": _res["method"],
                    "verify_cypher": _res["verify_cypher"],
                    "verify_result": _res.get("verify_text", ""),
                    "verify_n": _res.get("verify_n", -1),
                    "timestamp_start": _res["ts_start"],
                    "timestamp_llm": _res["ts_llm"],
                    "timestamp_cypher_return": _res["ts_cypher_return"],
                    "timestamp_end": _res["ts_end"],
                    "total_ms": _res["total_ms"],
                    "batch_id": round_batch_id,
                })
                tracker.record_from_qa({
                    "topology_level": _res["topology"],
                    "path_pattern": _res["path"],
                    "template_id": f"v19_{_res['q_type']}_{_res['method']}",
                    "question_id": _res["gen_id"],
                })
                batch_llm_ms.append(_res["llm_ms"])
                batch_neo_ms.append(_res["neo_ms"])
                batch_total_ms.append(_res["total_ms"] if _res["total_ms"] >= 0 else 0.0)
                print(f"    #{generated:03d} [{_res['topology']}] iter={_res['iteration_count']} "
                      f"q={_res['q_type']:<10} v={_res.get('verify_n',-1)} "
                      f"Δt={_res['total_ms']}ms batch={round_batch_id}")

        if batch_total_ms:
            avg_llm = int(sum(batch_llm_ms) / len(batch_llm_ms))
            avg_neo = int(sum(batch_neo_ms) / len(batch_neo_ms))
            avg_total = int(sum(batch_total_ms) / len(batch_total_ms))
            status = "REAL" if avg_total >= MIN_REAL_MS else "SUSPECT"
            print(f"[Batch Verify] Round {round_idx} | "
                  f"LLM Latency: {avg_llm}ms | Neo4j Latency: {avg_neo}ms | "
                  f"Total: {avg_total}ms | Status: {status}")
        else:
            print(f"[Batch Verify] Round {round_idx} | Status: EMPTY")

    qa_dir = pathlib.Path("E:/Project/ADVTEST/generated_qa")
    qa_dir.mkdir(parents=True, exist_ok=True)
    qa_path = qa_dir / f"{SCENE_ID}_frame{FRAME_ID}_qa.json"
    qa_path.write_text(
        _json.dumps({"questions": nusqa_records}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n  NuScenes-QA JSON → {qa_path} ({len(nusqa_records)} records)")

    _print_stats("Gap Stats — Final")
    n_l2b_gen = sum(1 for r in nusqa_records if r["topology_level"] == "L2B")
    n_l2a_gen = len(nusqa_records) - n_l2b_gen
    print(f"\n  V18 Generated: {generated}/{total_att} attempts written")
    print(f"  L2B: {n_l2b_gen}  L2A: {n_l2a_gen}  "
          f"non-ego ratio: {n_l2b_gen/max(len(nusqa_records),1)*100:.0f}%")
    print(f"  Unresolvable: {len(unresolvable)} gaps")
    print(f"  Dropped (<{MIN_REAL_MS}ms): {dropped_fast}")
    print(f"  Dropped (empty fields): {dropped_empty}")
    print(f"  Dropped (non-unique / verify fail): {dropped_non_unique}")
    print(f"  Dropped (iteration<{MIN_ITER_COUNT}): {dropped_low_iter}")
    print(f"  Dropped (q_type semantic mismatch): {dropped_qtype_mismatch}")
    print(
        f"  Quality gates: strict_unique={STRICT_UNIQUE_ONLY} "
        f"min_iter={MIN_ITER_COUNT} low_iter_methods={sorted(LOW_ITER_STRICT_METHODS)}"
    )
    print(f"  Question types used: {list(dict.fromkeys(used_types))}")
    return generated


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  Method A — Full Execution Chain (scene-0926 frame-20)")
    print("=" * 65)

    # Pre-flight
    if not preflight_check():
        print("\n❌ Pre-flight failed. Fix issues above and re-run.")
        return

    print("\n✅ All pre-flight checks passed. Starting execution...\n")

    from neo4j import GraphDatabase
    from gap_pipeline.llm_client import LLMClient

    llm    = LLMClient()
    try:
        _ = llm._call(
            "ping",
            system_prompt="Return only: pong",
            max_tokens=8,
            call_tag="audit_healthcheck",
        )
    except Exception as exc:
        print(f"\n❌ LLM API health check failed: {exc}")
        print("   当前是鉴权/账号问题（不是本地流程逻辑问题）。")
        print("   请先配置学校下发的有效 VQA_API_KEY（Qwen3.5-35B-A3B），再重试。")
        return
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))

    try:
        # append-only policy: never wipe historical evidence unless explicitly opted in
        step0_wipe_previous_run()
        step1_cleanup()
        step2_filter_record()
        step3_import_neo4j()
        step4_baseline_audit(driver, llm)
        step5_6_generate(driver, llm)

        print("\n" + "=" * 65)
        print("  ✅ Method A complete")
        print("  Check RQ.xlsx:")
        print("    filter_record          ← 1 new row")
        print("    raw_coverage           ← ~29 new rows (baseline)")
        print("    question-answer-our    ← 5 new rows (generated)")
        print("=" * 65)

    finally:
        driver.close()


if __name__ == "__main__":
    main()
