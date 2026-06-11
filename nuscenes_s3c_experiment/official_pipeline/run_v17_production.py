#!/usr/bin/env python3
"""
run_v17_production.py — V17 深夜量产自动化流水线
==================================================
策略：
  - 追加模式：不删除现有 Excel 记录，全部以 append 写入
  - 每帧独立 try/except，某帧失败不影响后续帧
  - V20 批量架构：生成 Batch-15 + 15-Worker；baseline 审计并行批处理
  - 每题立即 wb.save()（rq_tables 底层已保证）
  - 每帧生成独立 JSON: generated_qa/{scene}_{frame}_qa.json

目标帧（顺序执行）：
  1. scene-0553 / frame-8   (24 val qs)
  2. scene-0757 / frame-26  (0  val qs)   ← 无 baseline，仍跑 gap gen
  3. scene-1077 / frame-19  (0  val qs)   ← 无 baseline，仍跑 gap gen
  4. scene-0103 / frame-0   (5  val qs)
  5. scene-0103 / frame-38  (14 val qs)

执行前确认：
  ✅ 已设置 ADVTEST_ROOT（或 advtest_runtime.env），路径见 unified_site / advtest_paths
  ✅ RQ.xlsx 已关闭（无锁定文件）
  ✅ Neo4j（NEO4J_URI 等由 unified_site 解析）可达
  ✅ filtered_scene_graphs/ 中存在上述所有 SG 文件
     —— 或设置 VQA_BUILD_SCENE_GRAPH_ONTHEFLY=true：每帧从 nuScenes 现场构图再过滤，不读磁盘 SG
"""
import os, sys, pathlib, json, time, collections, logging
from datetime import datetime
from typing import Any, Dict, Optional

sys.path.insert(0, str(pathlib.Path(__file__).parent))

from advtest_env import load_advtest_env
load_advtest_env()

from advtest_paths import (
    EXCEL_PATH,
    FILTERED_SG_DIR,
    GEN_QA_DIR,
    NEO4J_PASSWORD as NEO4J_PWD,
    NEO4J_URI,
    NEO4J_USER,
    TRAINVAL_META as TRAINVAL,
    VQA_QA_JSON as QA_PATH,
)

# ─── 全局路径 / Neo4j（与 unified_site 一致）───────────────────────────────────
FSG_DIR = pathlib.Path(FILTERED_SG_DIR)


def v17_onthefly_enabled() -> bool:
    return os.getenv("VQA_BUILD_SCENE_GRAPH_ONTHEFLY", "").strip().lower() in ("1", "true", "yes", "on")


# ─── 量产帧计划 ──────────────────────────────────────────────────────────────
# (scene_id, frame_id, sg_filename)；可被环境变量 ADVTEST_FRAME_PLAN_JSON 覆盖（见 load_frame_plan）
FRAME_PLAN_DEFAULT = [
    ("scene-0553",  8, "scene-0553_frame8_scene_graph.json"),
    ("scene-0757", 26, "scene-0757_frame26_scene_graph.json"),
    ("scene-1077", 19, "scene-1077_frame19_scene_graph.json"),
    ("scene-0103",  0, "scene-0103_frame0_scene_graph.json"),
    ("scene-0103", 38, "scene-0103_frame38_scene_graph.json"),
]


def load_frame_plan() -> list:
    """
    返回 [(scene_id, frame_id, sg_filename), ...]

    若设置 ADVTEST_FRAME_PLAN_JSON=/path/to/plan.json，则从中读取，格式：
      {"site": "...", "frames": [{"scene_id","frame_id","sg_filename"}, ...]}
    """
    raw = os.getenv("ADVTEST_FRAME_PLAN_JSON", "").strip()
    if not raw:
        return list(FRAME_PLAN_DEFAULT)
    path = pathlib.Path(raw)
    if not path.is_file():
        raise FileNotFoundError(f"ADVTEST_FRAME_PLAN_JSON 不是文件: {path}")
    data = json.loads(path.read_text(encoding="utf-8"))
    frames = data.get("frames") or data
    if not isinstance(frames, list):
        raise ValueError("frame plan JSON 需要 frames 数组")
    out = []
    for f in frames:
        out.append((f["scene_id"], int(f["frame_id"]), f["sg_filename"]))
    return out

# ─── 日志配置 ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("v17_prod")

# ─────────────────────────────────────────────────────────────────────────────
# 预检
# ─────────────────────────────────────────────────────────────────────────────

def preflight() -> bool:
    print("\n[Pre-flight]")
    try:
        from advtest_paths import get_site

        s = get_site()
        print(
            f"  site: root={s.advtest_root}  nuScenes={s.nuscenes_dataroot}  "
            f"version={s.nuscenes_version}"
        )
    except Exception as exc:
        print(f"  ⚠️  site 摘要失败: {exc}")
    ok = True

    # Excel writable?
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        print(f"  ❌ Excel 锁定文件存在: {lock}  → 请先关闭 Excel")
        ok = False
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            wb.save(str(EXCEL_PATH)); wb.close()
            print(f"  ✅ Excel 可写")
        except FileNotFoundError:
            print(f"  ❌ Excel 不存在: {EXCEL_PATH}  → 复制模板 RQ 或创建空簿")
            ok = False
        except PermissionError:
            print(f"  ❌ Excel 写入测试失败 — 请关闭 Excel 后重试")
            ok = False

    # Neo4j（先尝试 docker start，再测 bolt）
    from neo4j_bootstrap import ensure_neo4j_listening
    if not ensure_neo4j_listening(NEO4J_URI):
        ok = False
    else:
        try:
            from neo4j import GraphDatabase
            d = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))
            with d.session() as s:
                s.run("RETURN 1").single()
            d.close()
            print(f"  ✅ Neo4j 可达: {NEO4J_URI}")
        except Exception as e:
            print(f"  ❌ Neo4j 不可达: {e}")
            ok = False

    plan = load_frame_plan()
    if v17_onthefly_enabled():
        print("  ✅ 现场构图模式 VQA_BUILD_SCENE_GRAPH_ONTHEFLY — 不检查磁盘 filtered_scene_graphs")
        try:
            import config as _cfg

            dk_raw = str(getattr(_cfg, "NUSCENES_DEVKIT_PATH", "") or "").strip()
            if dk_raw:
                dk = pathlib.Path(dk_raw)
                if dk.is_dir():
                    if dk_raw not in sys.path:
                        sys.path.insert(0, dk_raw)
                    print(f"  ✅ devkit 路径: {dk}")
                else:
                    # 常见：Windows 默认路径拷到 Linux；可改 .env 清空此项并 pip install nuscenes-devkit
                    print(f"  ⚠️  NUSCENES_DEVKIT_PATH 非本机目录，尝试已安装的 nuscenes 包: {dk_raw[:72]}...")
            from nuscenes.nuscenes import NuScenes  # noqa: F401

            print("  ✅ nuscenes 包可导入")
            print(
                f"  ✅ nuScenes 配置: version={getattr(_cfg, 'NUSCENES_VERSION', '?')}  "
                f"dataroot={str(getattr(_cfg, 'NUSCENES_DATAROOT', ''))[:72]}..."
            )
        except Exception as exc:
            print(f"  ❌ nuscenes 导入失败: {exc}")
            print("     处理: pip install nuscenes-devkit  或在 advtest_runtime.env 设正确的 NUSCENES_DEVKIT_PATH")
            ok = False
        if not TRAINVAL.is_dir():
            print(f"  ❌ Trainval 元数据目录不存在: {TRAINVAL}")
            ok = False
        else:
            print(f"  ✅ Trainval: {TRAINVAL}")
    else:
        for sid, fid, sg in plan:
            p = FSG_DIR / sg
            if p.exists():
                kb = p.stat().st_size // 1024
                print(f"  ✅ {sg} ({kb} KB)")
            else:
                print(f"  ❌ 缺失: {sg}")
                ok = False

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# 帧流水线：步骤 2-6
# ─────────────────────────────────────────────────────────────────────────────

def _abs_ts() -> str:
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]


def run_step2_filter_record(
    scene_id: str, frame_id: int, sg_name: str, filtered_data: Optional[Dict[str, Any]] = None
) -> bool:
    """写入 filter_record（V24：同帧覆盖 upsert，非追加）。"""
    from rq_tables import write_filter_record
    if filtered_data is not None:
        data = filtered_data
    else:
        sg = FSG_DIR / sg_name
        data = json.loads(sg.read_text(encoding="utf-8"))
    info = data["core_universe_filter"]
    vex_str = ",".join(sorted(info["node_ids_kept"]))
    ratio   = info["filtered_nodes"] / max(info["raw_nodes"], 1)
    ok = write_filter_record(
        scene_id=scene_id, frame_id=frame_id,
        original_num=info["raw_nodes"],
        filtered_num=info["filtered_nodes"],
        filtered_vex=vex_str,
        ratio=ratio,
    )
    print(f"  filter_record: raw={info['raw_nodes']} → filtered={info['filtered_nodes']} "
          f"(ratio={ratio:.2%})  {'✅' if ok else '❌'}")
    return ok


def run_step3_import_neo4j(sg_name: str, filtered_data: Optional[Dict[str, Any]] = None) -> dict:
    """清空 Neo4j 并导入当前帧场景图。"""
    if filtered_data is not None:
        from core_universe_filter import import_filtered_sg_data_to_neo4j

        result = import_filtered_sg_data_to_neo4j(
            filtered_data,
            neo4j_uri=NEO4J_URI,
            neo4j_user=NEO4J_USER,
            neo4j_pwd=NEO4J_PWD,
            source_label=f"onthefly:{sg_name}",
        )
    else:
        from core_universe_filter import import_filtered_sg_to_neo4j

        result = import_filtered_sg_to_neo4j(
            sg_name=sg_name,
            neo4j_uri=NEO4J_URI,
            neo4j_user=NEO4J_USER,
            neo4j_pwd=NEO4J_PWD,
        )
    print(f"  Neo4j: {result['n_nodes']} nodes, {result['n_edges']} edges ✅")
    return result


def run_step4_baseline_audit(
        scene_id: str, frame_id: int,
        driver, llm_client,
) -> int:
    """baseline 审计（val 原题 → 足迹 → raw_coverage）。"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    from semantic_auditor import audit_baseline_question, build_scene_context
    from semantic_auditor import make_qa_id, _ms_now as _audit_ms, derive_l2_from_l1
    from rq_tables import write_baseline_to_coverage

    # 构建 sample_token → (scene_name, frame_idx) 映射
    scenes  = json.loads((TRAINVAL / "scene.json").read_text())
    samples = json.loads((TRAINVAL / "sample.json").read_text())
    st2name = {s["token"]: s["name"] for s in scenes}
    tok2info: dict = {}
    s2tok: dict = collections.defaultdict(list)
    for samp in samples:
        sname = st2name.get(samp["scene_token"], "?")
        tok2info[samp["token"]] = {"scene_name": sname, "timestamp": samp["timestamp"]}
        s2tok[sname].append(samp["token"])
    for sname, toks in s2tok.items():
        for idx, tok in enumerate(sorted(toks, key=lambda t: tok2info[t]["timestamp"])):
            tok2info[tok]["frame_idx"] = idx

    all_val_qs = json.loads(QA_PATH.read_text())["questions"]
    target_qs = [
        (gi, q)
        for gi, q in enumerate(all_val_qs)
        if tok2info.get(q.get("sample_token", ""), {}).get("scene_name") == scene_id
        and tok2info.get(q.get("sample_token", ""), {}).get("frame_idx") == frame_id
    ]
    print(f"  Found {len(target_qs)} val questions for {scene_id}/f{frame_id}")
    if not target_qs:
        print("  (no baseline questions — skipping audit, gap gen will still run)")
        return 0

    scene_ctx = build_scene_context(driver)
    print(f"  Scene context: {len(scene_ctx)} chars")
    BASELINE_N_WORKERS = 15
    _tls = threading.local()
    _llm_cls = llm_client.__class__

    with driver.session() as _sess:
        _valid_nodes = {
            str(r["id"]).strip()
            for r in _sess.run("MATCH (n:Object) RETURN n.unique_id AS id")
            if r.get("id")
        }
        _valid_edges = {
            (str(r["src"]).strip(), str(r["tgt"]).strip())
            for r in _sess.run(
                "MATCH (s:Object)-[:RELATES_TO]->(t:Object) "
                "RETURN s.unique_id AS src, t.unique_id AS tgt"
            )
            if r.get("src") and r.get("tgt")
        }
        _type_to_ids: dict = {}
        for r in _sess.run("MATCH (n:Object) RETURN n.unique_id AS id, n.type AS t"):
            uid = str(r.get("id") or "").strip()
            typ = str(r.get("t") or "").strip().lower()
            if uid and typ:
                _type_to_ids.setdefault(typ, []).append(uid)
        for _k in _type_to_ids:
            _type_to_ids[_k] = sorted(_type_to_ids[_k])

    def _norm_edges(raw_edges):
        out, seen = [], set()
        for e in (raw_edges or []):
            if not isinstance(e, dict):
                continue
            src = str(e.get("source") or e.get("src") or "").strip()
            tgt = str(e.get("target") or e.get("dst") or e.get("tgt") or "").strip()
            if not src or not tgt or src not in _valid_nodes or tgt not in _valid_nodes:
                continue
            if (src, tgt) not in _valid_edges:
                continue
            if (src, tgt) in seen:
                continue
            seen.add((src, tgt))
            rel = str(e.get("relation") or e.get("dir") or e.get("direction_4") or "").strip()
            item = {"source": src, "target": tgt}
            if rel:
                item["relation"] = rel
            out.append(item)
        return out

    def _finalize_l0(raw_nodes, l1_norm: list, audit_res: dict) -> list:
        graph_ids: list = []
        gid_set: set = set()

        def _push_gid(x: str) -> None:
            s = str(x or "").strip()
            if not s or s not in _valid_nodes or s in gid_set:
                return
            gid_set.add(s)
            graph_ids.append(s)

        def _candidates_for_type_label(lab: str) -> list:
            lab = (lab or "").strip().lower()
            if not lab:
                return []
            ids = list(_type_to_ids.get(lab, []))
            if not ids and len(lab) > 3 and lab.endswith("s"):
                ids = list(_type_to_ids.get(lab[:-1], []))
            return ids

        def _resolve_type_label(label: str) -> list:
            lab = (label or "").strip().lower()
            if not lab or lab == "ego" or lab == "any":
                return []
            cand = _candidates_for_type_label(lab)
            if len(cand) != 1:
                return []
            return list(cand)

        for n in (raw_nodes or []):
            s = str(n or "").strip()
            if not s:
                continue
            if s in _valid_nodes:
                _push_gid(s)
            else:
                for rid in _resolve_type_label(s):
                    _push_gid(rid)
        for e in l1_norm:
            _push_gid(e.get("source", ""))
            _push_gid(e.get("target", ""))

        if graph_ids:
            return graph_ids

        return []

    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            _llm = _llm_cls()
            _tls.llm = _llm
        return _llm

    def _audit_one(gi: int, q: dict) -> dict:
        question = q.get("question", "")
        answer = q.get("answer", "")
        qtype = q.get("template_type", "")
        nhop = q.get("num_hop", 0)
        qa_uid = make_qa_id(gi, qtype)
        ts0 = _audit_ms()
        try:
            audit_res = audit_baseline_question(
                question=question, q_type=qtype, num_hop=nhop,
                driver=driver, llm_client=_get_worker_llm(),
                scene_context=scene_ctx, global_index=gi,
            )
        except Exception as exc:
            print(f"    ⚠️ [baseline-worker-exc] idx={gi} {exc}")
            audit_res = {"l0_nodes": [], "l1_edges": []}
        ts1 = _audit_ms()
        l1 = _norm_edges(audit_res.get("l1_edges", []) or [])
        l0 = _finalize_l0(audit_res.get("l0_nodes", []) or [], l1, audit_res)
        l2 = derive_l2_from_l1(l1)
        audit_json = json.dumps(
            {"nodes": l0, "edges": l1},
            ensure_ascii=False
        )
        return {
            "qa_uid": qa_uid,
            "qtype": qtype,
            "question": question,
            "answer": answer,
            "l0_nodes": l0,
            "l1_edges": l1,
            "l2_paths": l2,
            "timestamp_start": ts0,
            "timestamp_end": ts1,
            "audit_cypher": audit_json,
        }

    t_batch0 = time.perf_counter()
    ordered = [None] * len(target_qs)
    with ThreadPoolExecutor(max_workers=BASELINE_N_WORKERS) as _ex:
        fut2idx = {
            _ex.submit(_audit_one, gi, q): idx
            for idx, (gi, q) in enumerate(target_qs)
        }
        for done_i, _f in enumerate(as_completed(fut2idx), 1):
            idx = fut2idx[_f]
            rec = _f.result()
            ordered[idx] = rec
            if done_i <= 3 or done_i % 10 == 0:
                print(f"  [audit {done_i:2d}/{len(target_qs)}] {rec['qa_uid']:<28} "
                      f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} L2={len(rec['l2_paths'])}")
    batch_ms = int((time.perf_counter() - t_batch0) * 1000)
    avg_ms = int(batch_ms / max(len(target_qs), 1))
    print(f"  [Baseline Batch] questions={len(target_qs)} workers={BASELINE_N_WORKERS} "
          f"elapsed={batch_ms}ms avg={avg_ms}ms/q")

    success = 0
    for i, rec in enumerate(ordered, 1):
        if rec is None:
            continue
        ok = write_baseline_to_coverage(
            scene_id=scene_id, frame_id=frame_id,
            nuscenes_qa_id=rec["qa_uid"],
            question=rec["question"], answer=rec["answer"],
            l0_nodes=rec["l0_nodes"], l1_edges=rec["l1_edges"],
            l2_paths=rec["l2_paths"], question_type=rec["qtype"],
            timestamp_start=rec["timestamp_start"], timestamp_end=rec["timestamp_end"],
            audit_cypher=rec["audit_cypher"],
            global_val_index=target_qs[i - 1][0],
        )
        if ok:
            success += 1
        if i <= 3 or i % 10 == 0:
            print(f"  [write {i:2d}/{len(target_qs)}] {rec['qa_uid']:<28} "
                  f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} "
                  f"L2={len(rec['l2_paths'])} {'OK' if ok else '--'}")

    print(f"  Baseline audit: {success}/{len(target_qs)} written ✅")
    return success


def run_step56_generate(
        scene_id: str, frame_id: int,
        driver, llm_client,
) -> dict:
    """
    V16 批量生成（从 run_method_a 借用逻辑，参数化 scene/frame）。
    通过临时修改 run_method_a 模块全局变量来复用已有 V16 实现。
    返回 {"generated": N, "attempts": M, "unresolvable": K, "qa_json_path": ...}
    """
    import run_method_a as _rma

    # 临时替换模块级全局变量
    _old = (_rma.SCENE_ID, _rma.FRAME_ID, _rma.TARGET_SG)
    _rma.SCENE_ID  = scene_id
    _rma.FRAME_ID  = frame_id
    _rma.TARGET_SG = f"{scene_id}_frame{frame_id}_scene_graph.json"
    try:
        n = _rma.step5_6_generate(driver, llm_client)
    finally:
        _rma.SCENE_ID, _rma.FRAME_ID, _rma.TARGET_SG = _old

    qa_path = GEN_QA_DIR / f"{scene_id}_frame{frame_id}_qa.json"
    n_json  = 0
    if qa_path.exists():
        try:
            n_json = len(json.loads(qa_path.read_text())["questions"])
        except Exception:
            pass
    return {
        "generated":   n,
        "qa_json_path": str(qa_path),
        "n_json":       n_json,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 单帧执行入口
# ─────────────────────────────────────────────────────────────────────────────

def run_single_frame(
        scene_id: str, frame_id: int, sg_name: str,
        driver, llm_client,
) -> dict:
    """
    执行单帧完整流水线（步骤 2-6）。
    失败时抛出异常，由外层捕获记录。
    返回统计摘要 dict。
    """
    t0 = time.perf_counter()
    label = f"{scene_id}/f{frame_id}"
    print(f"\n{'─'*65}")
    print(f"  [{label}]  SG={sg_name}")
    print(f"{'─'*65}")

    if v17_onthefly_enabled():
        from v17_onthefly_sg import build_filtered_sg_onthefly

        filtered_data = build_filtered_sg_onthefly(scene_id, frame_id)
        run_step2_filter_record(scene_id, frame_id, sg_name, filtered_data=filtered_data)
        run_step3_import_neo4j(sg_name, filtered_data=filtered_data)
    else:
        run_step2_filter_record(scene_id, frame_id, sg_name)
        run_step3_import_neo4j(sg_name)

    # Step 4: Baseline 审计
    n_baseline = run_step4_baseline_audit(scene_id, frame_id, driver, llm_client)

    # Step 5+6: V16 生成
    gen_result = run_step56_generate(scene_id, frame_id, driver, llm_client)

    elapsed_s = time.perf_counter() - t0
    stats = {
        "scene_id":    scene_id,
        "frame_id":    frame_id,
        "n_baseline":  n_baseline,
        "n_generated": gen_result["generated"],
        "n_json":      gen_result["n_json"],
        "qa_json_path": gen_result["qa_json_path"],
        "elapsed_s":   round(elapsed_s, 1),
        "status":      "OK",
    }
    print(f"\n  [{label}] ✅ DONE  baseline={n_baseline}  generated={gen_result['generated']}  "
          f"elapsed={elapsed_s:.0f}s")
    return stats


# ─────────────────────────────────────────────────────────────────────────────
# 量产汇总报告
# ─────────────────────────────────────────────────────────────────────────────

def print_summary(results: list, total_elapsed_s: float):
    print(f"\n{'='*65}")
    print(f"  V17 量产汇总报告  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*65}")

    ok_frames  = [r for r in results if r["status"] == "OK"]
    err_frames = [r for r in results if r["status"] != "OK"]

    total_baseline  = sum(r.get("n_baseline", 0) for r in ok_frames)
    total_generated = sum(r.get("n_generated", 0) for r in ok_frames)

    print(f"\n  帧执行情况:")
    for r in results:
        mark = "✅" if r["status"] == "OK" else "❌"
        if r["status"] == "OK":
            print(f"    {mark} {r['scene_id']}/f{r['frame_id']:<3}  "
                  f"baseline={r['n_baseline']:>3}  generated={r['n_generated']:>4}  "
                  f"json={r['n_json']:>4}  {r['elapsed_s']:.0f}s")
        else:
            print(f"    {mark} {r['scene_id']}/f{r['frame_id']:<3}  "
                  f"ERROR: {r.get('error','?')[:60]}")

    print(f"\n  汇总统计:")
    print(f"    成功帧数            : {len(ok_frames)} / {len(results)}")
    print(f"    新增 raw_coverage 行 : {total_baseline}")
    print(f"    新增 question-answer 行: {total_generated}")
    print(f"    总耗时              : {total_elapsed_s:.0f}s  ({total_elapsed_s/60:.1f} min)")

    # Excel 当前行数
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
        for sh in ["raw_coverage", "question-answer-our"]:
            ws = wb[sh]
            n_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True)
                         if any(v is not None for v in _))
            print(f"    {sh:<28}: 共 {n_rows} 数据行")
        wb.close()
    except Exception as exc:
        print(f"    (Excel 行数读取失败: {exc})")

    if err_frames:
        print(f"\n  失败帧详情:")
        for r in err_frames:
            print(f"    {r['scene_id']}/f{r['frame_id']}: {r.get('error','?')}")

    print(f"\n  生成 JSON 文件:")
    for r in ok_frames:
        p = pathlib.Path(r["qa_json_path"])
        exists = "✅" if p.exists() else "❌ missing"
        print(f"    {exists}  {r['qa_json_path']}")

    print(f"{'='*65}")


# ─────────────────────────────────────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────────────────────────────────────

def main():
    frame_plan = load_frame_plan()
    print("=" * 65)
    print("  run_v17_production.py — 深夜量产自动化流水线")
    print(f"  目标帧: {len(frame_plan)} 帧")
    if os.getenv("ADVTEST_FRAME_PLAN_JSON", "").strip():
        print(f"  帧表: {os.getenv('ADVTEST_FRAME_PLAN_JSON')}")
    print(f"  Excel: {EXCEL_PATH}")
    print(f"  模式: 追加 (不删除现有 Excel 记录)")
    if v17_onthefly_enabled():
        print("  场景图: 现场从 nuScenes 构建 + 过滤 (VQA_BUILD_SCENE_GRAPH_ONTHEFLY)")
    else:
        print(f"  场景图: 磁盘 {FSG_DIR}")
    print("=" * 65)

    if not preflight():
        print("\n❌ Pre-flight 失败，请修复后重试。")
        return

    print("\n✅ Pre-flight 通过，开始量产...\n")

    from neo4j_bootstrap import ensure_neo4j_listening
    if not ensure_neo4j_listening(NEO4J_URI):
        print("\n❌ Neo4j 不可用，中止。")
        return

    from neo4j import GraphDatabase
    from gap_pipeline.llm_client import LLMClient

    llm    = LLMClient()
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))

    results = []
    t_total_start = time.perf_counter()

    try:
        for idx, (scene_id, frame_id, sg_name) in enumerate(frame_plan, 1):
            print(f"\n{'█'*65}")
            print(f"  帧 {idx}/{len(frame_plan)}: {scene_id}/frame-{frame_id}")
            print(f"{'█'*65}")
            try:
                stats = run_single_frame(scene_id, frame_id, sg_name, driver, llm)
                results.append(stats)
            except KeyboardInterrupt:
                print(f"\n⚠️  用户中断 (Ctrl+C)，保存已完成帧的结果...")
                results.append({
                    "scene_id": scene_id, "frame_id": frame_id,
                    "status": "INTERRUPTED",
                    "error": "KeyboardInterrupt",
                })
                break
            except Exception as exc:
                import traceback
                err_msg = str(exc)
                print(f"\n  ❌ [{scene_id}/f{frame_id}] 帧失败: {err_msg[:120]}")
                print(f"  跳过该帧，继续下一帧...")
                traceback.print_exc()
                results.append({
                    "scene_id": scene_id, "frame_id": frame_id,
                    "status": "ERROR",
                    "error": err_msg,
                    "n_baseline":  0,
                    "n_generated": 0,
                    "n_json":      0,
                    "elapsed_s":   0,
                })

    finally:
        driver.close()
        total_elapsed_s = time.perf_counter() - t_total_start
        print_summary(results, total_elapsed_s)

        # 保存汇总 JSON 到 generated_qa/
        summary_path = GEN_QA_DIR / f"v17_production_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        try:
            GEN_QA_DIR.mkdir(parents=True, exist_ok=True)
            summary_path.write_text(
                json.dumps({
                    "run_time":    datetime.now().isoformat(),
                    "total_frames": len(frame_plan),
                    "results":     results,
                    "total_elapsed_s": total_elapsed_s,
                }, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"\n  汇总 JSON → {summary_path}")
        except Exception as exc:
            print(f"  ⚠️ 汇总 JSON 保存失败: {exc}")


if __name__ == "__main__":
    main()
