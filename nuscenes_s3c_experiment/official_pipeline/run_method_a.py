#!/usr/bin/env python3
"""
run_method_a.py — 方案 A 闭环执行链 (scene-0926 frame-20)

执行前检查清单（必须全部通过才能运行）：
  [OK] RQ.xlsx 已关闭（无 ~$RQ.xlsx 锁定文件）
  [OK] filtered_scene_graphs/scene-0926_frame20_scene_graph.json 存在
  [OK] Neo4j（默认 7687 端口，见 NEO4J_URI）可达

执行链：
  Step 1 : 净化环境（清理脏数据行、删除锁定文件）
  Step 2 : 写入 filter_record（核心宇宙过滤元数据）
  Step 3 : 清空 Neo4j + 从 filtered_scene_graphs/ 导入 scene-0926-20
  Step 4 : Baseline 审计（29 条原题 -> LLM Cypher -> 足迹 -> raw_coverage）
  Step 5 : Gap Detection（未覆盖 L2 路径）
  Step 6 : 增量生成（5 条 L2 题 -> question-answer-our）
  Step 7 : Final save 确认
"""
import os, sys, pathlib, json, time, collections, subprocess
sys.path.insert(0, str(pathlib.Path(__file__).parent))

from advtest_env import load_advtest_env
load_advtest_env()

from advtest_paths import (
    EXCEL_PATH,
    FILTERED_SG_DIR,
    GEN_QA_DIR,
    NEO4J_PASSWORD as NEO4J_PWD,
    NEO4J_URI,
    NEO4J_USER,
    TRAINVAL_META as TRAINVAL,
    VQA_QA_JSON as QA_PATH,
)

from coverage_tracker_patch import install_patch
install_patch()

FSG_DIR     = pathlib.Path(FILTERED_SG_DIR)
TARGET_SG   = "scene-0926_frame20_scene_graph.json"
SCENE_ID    = "scene-0926"
FRAME_ID    = 20
ALLOW_EXCEL_DESTRUCTIVE_CLEAN = os.getenv("VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN", "false").lower() in ("true", "1", "yes")


def _console_safe(s: object) -> str:
    """Strip/replace chars the Windows console (e.g. GBK) cannot print (LLM may emit ✓/❌)."""
    if s is None:
        return ""
    t = str(s)
    enc = getattr(sys.stdout, "encoding", None) or "utf-8"
    try:
        return t.encode(enc, errors="replace").decode(enc, errors="replace")
    except (LookupError, UnicodeError):
        return t.encode("ascii", errors="replace").decode("ascii")


# ─────────────────────────────────────────────────────────────────────────────
# Optional: school VPN (Windows rasdial)
# ─────────────────────────────────────────────────────────────────────────────

def maybe_connect_school_vpn() -> bool:
    """
    若 VQA_AUTO_CONNECT_VPN 为真：在 Windows 上按环境变量尝试拨号（rasdial）。
    账号口令只应写在 advtest_runtime.env（已在 .gitignore），勿提交仓库。
    """
    if os.name != "nt":
        return True
    v = os.getenv("VQA_AUTO_CONNECT_VPN", "").strip().lower()
    if v not in ("1", "true", "yes", "on"):
        return True
    conn = os.getenv("SCHOOL_VPN_CONN_NAME", "").strip()
    user = os.getenv("SCHOOL_VPN_USER", "").strip()
    pwd = os.getenv("SCHOOL_VPN_PASS", "").strip()
    if not conn or not user or not pwd:
        print("[FAIL][VPN] 已开启 VQA_AUTO_CONNECT_VPN，但缺少 SCHOOL_VPN_CONN_NAME / SCHOOL_VPN_USER / SCHOOL_VPN_PASS。")
        return False
    script = pathlib.Path(__file__).resolve().parent / "scripts" / "connect_school_vpn.ps1"
    if not script.is_file():
        print(f"[FAIL][VPN] 脚本不存在: {script}")
        return False
    print(f"\n[VPN] 预连校园网（Windows 连接名: {conn}）…")
    try:
        r = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script),
            ],
            env=os.environ.copy(),
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        print("[FAIL][VPN] 拨号超时（120s）。")
        return False
    except Exception as e:
        print(f"[FAIL][VPN] 执行失败: {e}")
        return False
    if r.returncode != 0:
        print(f"[FAIL][VPN] rasdial 退出码 {r.returncode}。请检查 Windows VPN 条目名、账号口令与学校网关。")
        return False
    print("[VPN] 拨号成功。")
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Pre-flight checks
# ─────────────────────────────────────────────────────────────────────────────

def preflight_check() -> bool:
    print("\n[Pre-flight checks]")
    ok = True

    # Trainval 元数据 + QA JSON（6019 帧抽自全量 trainval）
    if not TRAINVAL.is_dir():
        print(f"  [FAIL] Trainval metadata missing: {TRAINVAL}")
        print("     设置 NUSCENES_DATAROOT / NUSCENES_VERSION（默认 dataroot 为 $ADVTEST_ROOT/data）")
        ok = False
    else:
        for _fn in ("scene.json", "sample.json"):
            if not (TRAINVAL / _fn).is_file():
                print(f"  [FAIL] Missing {TRAINVAL / _fn}")
                ok = False
    if not QA_PATH.is_file():
        print(f"  [FAIL] NuScenes-QA JSON missing: {QA_PATH}")
        ok = False

    # 1. Kill lock file (OfficeClickToRun keeps recreating it — force-delete and test write)
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        print(f"  [WARN]  Lock file found: {lock} — attempting force-delete...")
        # Kill OfficeClickToRun silently to prevent immediate recreation
        subprocess.run(
            ["powershell", "-Command",
             "Get-Process | Where-Object { $_.Name -match 'EXCEL|OfficeClickToRun' } | "
             "ForEach-Object { Stop-Process -Id $_.Id -Force -ErrorAction SilentlyContinue }"],
            capture_output=True, timeout=10
        )
        import time as _t; _t.sleep(1)
        try:
            lock.unlink(missing_ok=True)
            print(f"  [OK] Lock file deleted")
        except Exception as e:
            print(f"  [WARN]  Could not delete lock: {e}")
    else:
        print(f"  [OK] No lock file")

    # 2. Filtered SG exists
    sg = FSG_DIR / TARGET_SG
    if sg.exists():
        data = json.loads(sg.read_text(encoding="utf-8"))
        n = data.get("core_universe_filter", {}).get("filtered_nodes", "?")
        print(f"  [OK] Filtered SG exists: {sg.name} ({n} nodes)")
    else:
        print(f"  [FAIL] Filtered SG missing: {sg}")
        ok = False

    # 3. Neo4j reachable
    try:
        from neo4j import GraphDatabase
        d = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))
        with d.session() as s:
            s.run("RETURN 1").single()
        d.close()
        print(f"  [OK] Neo4j reachable: {NEO4J_URI}")
    except Exception as e:
        print(f"  [FAIL] Neo4j not reachable: {e}")
        ok = False

    # 4. Excel actual write test (the TRUE test — lock file existence alone is not sufficient)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        wb.save(str(EXCEL_PATH))
        wb.close()
        print(f"  [OK] Excel writable: {EXCEL_PATH}")
    except PermissionError:
        print(f"  [FAIL] Excel write test FAILED — file is still locked by another process")
        ok = False
    except Exception as e:
        print(f"  [WARN]  Excel check error: {e}")

    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Step 0: Wipe previous run data (DISABLED by default, append-only policy)
# ─────────────────────────────────────────────────────────────────────────────

def step0_wipe_previous_run():
    """Destructive cleanup (opt-in only). Default behavior is append-only."""
    print("\n[Step 0] Destructive Excel wipe")
    if not ALLOW_EXCEL_DESTRUCTIVE_CLEAN:
        print("  [SKIP]  Skipped (append-only mode; VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN is false)")
        return
    print("  [WARN]  Enabled by VQA_ALLOW_EXCEL_DESTRUCTIVE_CLEAN=true")
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH))
        total_deleted = 0
        for sh in ["raw_coverage", "question-answer-our"]:
            ws = wb[sh]
            max_row = ws.max_row
            # Delete from bottom to preserve row indices
            for r in range(max_row, 1, -1):   # row 1 = header, skip
                ws.delete_rows(r)
                total_deleted += 1
        wb.save(str(EXCEL_PATH))
        wb.close()
        print(f"  [OK] Deleted {total_deleted} data rows (headers preserved)")
    except Exception as e:
        print(f"  [WARN]  Wipe error: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Environment cleanup (non-destructive, lock file only)
# ─────────────────────────────────────────────────────────────────────────────

def step1_cleanup():
    print("\n[Step 1] Environment cleanup (append-only, non-destructive)")

    # Remove lock file if still present
    lock = EXCEL_PATH.parent / f"~${EXCEL_PATH.name}"
    if lock.exists():
        try:
            lock.unlink()
            print(f"  [OK] Deleted lock file: {lock}")
        except Exception as e:
            print(f"  [WARN]  Could not delete lock: {e}")
    print("  [OK] Keep all historical rows (no row deletion)")


# ─────────────────────────────────────────────────────────────────────────────
# Step 1.5: Rebuild filtered SG with official policy (30/40/50m + vis + px)
# ─────────────────────────────────────────────────────────────────────────────
def step1_5_refresh_filtered_sg() -> bool:
    print("\n[Step 1.5] Rebuilding filtered SG with official policy")
    from core_universe_filter import (
        RAW_SG_DIR,
        FILTERED_SG_DIR,
        MIN_VISIBILITY,
        filter_and_save,
    )
    raw_path = RAW_SG_DIR / TARGET_SG
    if not raw_path.exists():
        print(f"  [FAIL] Raw SG missing: {raw_path}")
        return False
    out = filter_and_save(
        raw_path=raw_path,
        write_excel=False,
        distance_mode="official",
        pixel_mode=os.getenv("VQA_PIXEL_MODE", "lenient"),
        min_visibility=float(os.getenv("VQA_FILTER_MIN_VISIBILITY", str(MIN_VISIBILITY))),
        output_dir=FILTERED_SG_DIR,
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    info = data.get("core_universe_filter", {})
    print(
        f"  [OK] {out.name}: nodes {info.get('raw_nodes', '?')} -> {info.get('filtered_nodes', '?')}, "
        f"edges {info.get('raw_edges', '?')} -> {info.get('filtered_edges', '?')}, "
        f"mode={info.get('distance_mode', '?')}, pixel={info.get('pixel_mode', '?')}"
    )
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Write filter_record
# ─────────────────────────────────────────────────────────────────────────────

def step2_filter_record():
    print("\n[Step 2] Writing filter_record")
    from rq_tables import write_filter_record

    sg = FSG_DIR / TARGET_SG
    data = json.loads(sg.read_text(encoding="utf-8"))
    info = data["core_universe_filter"]
    vex_str = ",".join(sorted(info["node_ids_kept"]))
    ratio   = info["filtered_nodes"] / max(info["raw_nodes"], 1)

    ok = write_filter_record(
        scene_id=SCENE_ID, frame_id=FRAME_ID,
        original_num=info["raw_nodes"],
        filtered_num=info["filtered_nodes"],
        filtered_vex=vex_str,
        ratio=ratio,
    )
    print(f"  scene={SCENE_ID} frame={FRAME_ID}")
    print(f"  raw={info['raw_nodes']} -> filtered={info['filtered_nodes']} "
          f"(ratio={ratio:.2%})")
    print(f"  node_ids: {info['node_ids_kept']}")
    print(f"  write_filter_record: {'[OK]' if ok else '[FAIL]'}")
    return ok


# ─────────────────────────────────────────────────────────────────────────────
# Step 3: Neo4j import from filtered_scene_graphs/
# ─────────────────────────────────────────────────────────────────────────────

def step3_import_neo4j():
    print("\n[Step 3] Importing filtered SG to Neo4j")
    from core_universe_filter import import_filtered_sg_to_neo4j
    result = import_filtered_sg_to_neo4j(
        sg_name=TARGET_SG,
        neo4j_uri=NEO4J_URI,
        neo4j_user=NEO4J_USER,
        neo4j_pwd=NEO4J_PWD,
    )
    print(f"  [OK] Neo4j: {result['n_nodes']} nodes, {result['n_edges']} edges")
    print(f"  Source : {result['source']}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Step 4: Baseline audit (29 original questions -> LLM Cypher -> footprint -> raw_coverage)
# ─────────────────────────────────────────────────────────────────────────────

def step4_baseline_audit(driver, llm_client):
    print("\n[Step 4] Baseline audit (original NuScenes-QA questions)")
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    from semantic_auditor import audit_baseline_question, build_scene_context
    from rq_tables import write_baseline_to_coverage

    # Build sample->scene mapping
    scenes  = json.loads((TRAINVAL/"scene.json").read_text())
    samples = json.loads((TRAINVAL/"sample.json").read_text())
    st2name = {s["token"]: s["name"] for s in scenes}
    s2tok: dict = collections.defaultdict(list)
    tok2info: dict = {}
    for samp in samples:
        sname = st2name.get(samp["scene_token"],"?")
        tok2info[samp["token"]] = {"scene_name": sname, "timestamp": samp["timestamp"]}
        s2tok[sname].append(samp["token"])
    for sname, toks in s2tok.items():
        for idx, tok in enumerate(sorted(toks, key=lambda t: tok2info[t]["timestamp"])):
            tok2info[tok]["frame_idx"] = idx

    # Get scene-0926 val questions  —  keep global index for unique ID generation
    all_val_qs = json.loads(QA_PATH.read_text())["questions"]
    target_qs = [
        (global_i, q)
        for global_i, q in enumerate(all_val_qs)     # global_i = position in val file
        if tok2info.get(q.get("sample_token",""), {}).get("scene_name") == SCENE_ID
        and tok2info.get(q.get("sample_token",""), {}).get("frame_idx") == FRAME_ID
    ]
    print(f"  Found {len(target_qs)} val questions for {SCENE_ID} frame-{FRAME_ID}")
    if target_qs:
        print(f"  Global indices: {target_qs[0][0]} – {target_qs[-1][0]}")

    if not target_qs:
        print("  [WARN]  No questions found — check sample_token mapping")
        return 0

    # Build scene context once (reuse across questions)
    scene_ctx = build_scene_context(driver)
    print(f"  Scene context built ({len(scene_ctx)} chars)")

    from semantic_auditor import make_qa_id, _ms_now as _audit_ms, derive_l2_from_l1
    BASELINE_N_WORKERS = int(os.getenv("VQA_BASELINE_N_WORKERS", "32"))
    _tls = threading.local()
    _llm_cls = llm_client.__class__

    # Snapshot real graph entities once; baseline footprints are constrained to this set.
    with driver.session() as _sess:
        _valid_nodes = {
            str(r["id"]).strip()
            for r in _sess.run("MATCH (n:Object) RETURN n.unique_id AS id")
            if r.get("id")
        }
        _valid_edges = {
            (str(r["src"]).strip(), str(r["tgt"]).strip())
            for r in _sess.run(
                "MATCH (s:Object)-[:RELATES_TO]->(t:Object) "
                "RETURN s.unique_id AS src, t.unique_id AS tgt"
            )
            if r.get("src") and r.get("tgt")
        }
        _type_to_ids: dict = {}
        for r in _sess.run("MATCH (n:Object) RETURN n.unique_id AS id, n.type AS t"):
            uid = str(r.get("id") or "").strip()
            typ = str(r.get("t") or "").strip().lower()
            if uid and typ:
                _type_to_ids.setdefault(typ, []).append(uid)
        for _k in _type_to_ids:
            _type_to_ids[_k] = sorted(_type_to_ids[_k])

    def _norm_edges(raw_edges):
        out = []
        seen = set()
        for e in (raw_edges or []):
            if not isinstance(e, dict):
                continue
            src = str(e.get("source") or e.get("src") or "").strip()
            tgt = str(e.get("target") or e.get("dst") or e.get("tgt") or "").strip()
            if not src or not tgt:
                continue
            if src not in _valid_nodes or tgt not in _valid_nodes:
                continue
            if (src, tgt) not in _valid_edges:
                continue
            key = (src, tgt)
            if key in seen:
                continue
            seen.add(key)
            rel = str(e.get("relation") or e.get("dir") or e.get("direction_4") or "").strip()
            item = {"source": src, "target": tgt}
            if rel:
                item["relation"] = rel
            out.append(item)
        return out

    def _finalize_l0(raw_nodes, l1_norm: list, audit_res: dict) -> list:
        """
        L0 仅包含“有证据”的真实 unique_id（不做兜底补齐）：
          1) 审计结果里的节点ID；
          2) L1 端点；
          3) 审计结果里的类型词，仅当该类型在当前图中唯一时映射到该ID。
        """
        graph_ids: list = []
        gid_set: set = set()

        def _push_gid(x: str) -> None:
            s = str(x or "").strip()
            if not s or s not in _valid_nodes or s in gid_set:
                return
            gid_set.add(s)
            graph_ids.append(s)

        def _candidates_for_type_label(lab: str) -> list:
            lab = (lab or "").strip().lower()
            if not lab:
                return []
            ids = list(_type_to_ids.get(lab, []))
            if not ids and len(lab) > 3 and lab.endswith("s"):
                ids = list(_type_to_ids.get(lab[:-1], []))
            return ids

        def _resolve_type_label(label: str) -> list:
            """类型词 -> 若当前图中该 type 唯一，则映射到唯一 unique_id。"""
            lab = (label or "").strip().lower()
            if not lab or lab == "ego" or lab == "any":
                return []
            cand = _candidates_for_type_label(lab)
            if len(cand) != 1:
                return []
            return list(cand)

        for n in (raw_nodes or []):
            s = str(n or "").strip()
            if not s:
                continue
            if s in _valid_nodes:
                _push_gid(s)
            else:
                for rid in _resolve_type_label(s):
                    _push_gid(rid)
        for e in l1_norm:
            _push_gid(e.get("source", ""))
            _push_gid(e.get("target", ""))

        if graph_ids:
            return graph_ids
        return []

    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            _llm = _llm_cls()
            _tls.llm = _llm
        return _llm

    def _audit_one(global_idx: int, q: dict) -> dict:
        question = q.get("question", "")
        answer = q.get("answer", "")
        qtype = q.get("template_type", "")
        nhop = q.get("num_hop", 0)
        qa_uid = make_qa_id(global_idx, qtype)
        ts0 = _audit_ms()
        try:
            audit_res = audit_baseline_question(
                question=question, q_type=qtype, num_hop=nhop,
                driver=driver, llm_client=_get_worker_llm(),
                scene_context=scene_ctx,
                global_index=global_idx,
            )
        except Exception as exc:
            print(f"    [WARN] [baseline-worker-exc] idx={global_idx} {_console_safe(exc)}")
            audit_res = {"l0_nodes": [], "l1_edges": []}
        ts1 = _audit_ms()
        l1 = _norm_edges(audit_res.get("l1_edges", []) or [])
        l0 = _finalize_l0(audit_res.get("l0_nodes", []) or [], l1, audit_res)

        l2 = derive_l2_from_l1(l1)
        audit_json = json.dumps(
            {"nodes": l0, "edges": l1},
            ensure_ascii=False
        )
        return {
            "qa_uid": qa_uid,
            "question": question,
            "answer": answer,
            "qtype": qtype,
            "l0_nodes": l0,
            "l1_edges": l1,
            "l2_paths": l2,
            "timestamp_start": ts0,
            "timestamp_end": ts1,
            "audit_cypher": audit_json,
        }

    t_batch0 = time.perf_counter()
    ordered = [None] * len(target_qs)
    with ThreadPoolExecutor(max_workers=BASELINE_N_WORKERS) as _ex:
        fut2idx = {
            _ex.submit(_audit_one, global_idx, q): idx
            for idx, (global_idx, q) in enumerate(target_qs)
        }
        for done_i, _f in enumerate(as_completed(fut2idx), 1):
            idx = fut2idx[_f]
            rec = _f.result()
            ordered[idx] = rec
            if done_i <= 3 or done_i % 10 == 0:
                print(f"  [audit {done_i:2d}/{len(target_qs)}] {rec['qa_uid']:<25} "
                      f"L0={len(rec['l0_nodes'])} L1={len(rec['l1_edges'])} L2={len(rec['l2_paths'])}")
    batch_ms = int((time.perf_counter() - t_batch0) * 1000)
    avg_ms = int(batch_ms / max(len(target_qs), 1))
    print(f"  [Baseline Batch] questions={len(target_qs)} workers={BASELINE_N_WORKERS} "
          f"elapsed={batch_ms}ms avg={avg_ms}ms/q")

    success = 0
    for i, rec in enumerate(ordered, 1):
        if rec is None:
            continue

        ok = write_baseline_to_coverage(
            scene_id=SCENE_ID, frame_id=FRAME_ID,
            nuscenes_qa_id=rec["qa_uid"],    # e.g. "val_71051_comparison"
            question=rec["question"], answer=rec["answer"],
            l0_nodes=rec["l0_nodes"],
            l1_edges=rec["l1_edges"],
            l2_paths=rec["l2_paths"],        # 自动推导的 L2
            question_type=rec["qtype"],
            timestamp_start=rec["timestamp_start"], timestamp_end=rec["timestamp_end"],
            audit_cypher=rec["audit_cypher"],  # JSON 子图填入 cypher 列
            global_val_index=target_qs[i - 1][0],
        )
        if ok:
            success += 1
        if i <= 3 or i % 10 == 0:
            _ok_mark = "OK" if ok else "--"
            print(f"  [write {i:2d}/{len(target_qs)}] id={rec['qa_uid']:<25} "
                  f"q_type={rec['qtype']:12s} "
                  f"L0={len(rec['l0_nodes'])} "
                  f"L1={len(rec['l1_edges'])} "
                  f"L2={len(rec['l2_paths'])} "
                  f"{_ok_mark}")

    print(f"\n  Baseline audit done: {success}/{len(target_qs)} written to raw_coverage")
    return success


# ─────────────────────────────────────────────────────────────────────────────
# Step 5+6: Gap detection + generate 5 L2 questions -> question-answer-our
# ─────────────────────────────────────────────────────────────────────────────

# ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――
# V18 题型权重与答案策略（本地模板问句已移除）
# ―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――

_Q_TYPE_WEIGHTS = {"exist": 3, "status": 3, "object": 3, "count": 1}


def _answer_for_q_type(q_type: str, cell: dict, verify_n: int = -1) -> str:
    """仅生成答案字段；问题文本必须由 LLM 生成。"""
    if q_type == "exist":
        return "yes" if verify_n != 0 else "no"
    if q_type == "status":
        return cell.get("n3_status", "") or "unknown"
    if q_type == "count":
        return str(verify_n if verify_n >= 0 else 1)
    if q_type == "comparison":
        return "closer"
    return cell.get("n3_id", "")


def step5_6_generate(driver, llm_client):
    """
    V18 真实性硬化协议：
      1) timestamp_start        : 进入每个 gap cell 处理第一行捕获
      2) timestamp_llm          : 来自 llm_client._call() 返回后的物理时刻
      3) timestamp_cypher_return: Neo4j single() 返回后的物理时刻
      4) timestamp_end          : 写 Excel 前最后一行捕获
    严格规则：
      - 自然语言问题必须由 LLM 生成（无本地模板 fallback）
      - 若 (timestamp_end - timestamp_start) < 2000ms，则该条作废
      - cypher question / target_gap_cell 为空则作废
    """
    print("\n[Step 5+6] V18 Realness-Hardened Generation (4 physical timestamps locked)")
    import random
    import re
    import json as _json
    from datetime import datetime
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from gap_pipeline.coverage_tracker import CoverageTracker, CoverageRecord
    from gap_pipeline.constraint_methods import CumulativeConstraintChain
    from run_gap_pipeline_v6 import _process_single_cell
    from rq_tables import write_generated_question, make_generated_question_id

    # V20 并发参数：默认拉高以减少 LLM 往返；可用 VQA_Q_BATCH_SIZE / VQA_Q_N_WORKERS 覆盖
    Q_BATCH_SIZE = int(os.getenv("VQA_Q_BATCH_SIZE", "32"))
    Q_N_WORKERS = int(os.getenv("VQA_Q_N_WORKERS", "32"))
    # 真实性阈值
    MIN_REAL_MS = 2000
    # V23 质量门控（可通过环境变量覆盖）
    STRICT_UNIQUE_ONLY = os.getenv("VQA_STRICT_UNIQUE_ONLY", "true").lower() in ("true", "1", "yes")
    MIN_ITER_COUNT = int(os.getenv("VQA_MIN_ITER_COUNT", "1"))
    MAX_QTYPE_RETRY = int(os.getenv("VQA_QTYPE_MAX_RETRY", "2"))
    LOW_ITER_STRICT_METHODS = {
        x.strip() for x in os.getenv("VQA_MIN_ITER_STRICT_METHODS", "path").split(",")
        if x.strip()
    }
    MAX_ROUNDS = int(os.getenv("VQA_MAX_ROUNDS", "0"))
    # 连续整批 EMPTY（本批 8 条全失败）时退出本帧 Gap，避免同一帧空转数千轮。0=不启用。
    MAX_EMPTY_STREAK = int(os.getenv("VQA_MAX_EMPTY_STREAK", "0"))
    GAP_FAIL_COOLDOWN = int(os.getenv("VQA_GAP_FAIL_COOLDOWN", "0"))

    from gap_pipeline.constraint_methods import (
        init_forensics as _init_f,
        _fw as _fwu,
        canonical_cell_key as _canonical_cell_key,
    )
    _f_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _f_path = f"output/forensics_{_f_ts}.jsonl"
    _init_f(_f_path)
    print(f"  [M0 Forensics] -> {_f_path}")

    # V2 Pilot 约束方法优先级 P1–P16（与设计表一致）见 gap_pipeline/constraint_methods.py
    # 中「V2 Pilot 设计优先级」注释块；下方 key 多为实现侧 method_used / class.name。

    _CONSTRAINT_DESC_MAP = {
        "type_filter":         "target type unique in this direction",
        "status_anchor":       "target status unique in this direction",
        "type_status_anchor":  "type+status combo unique in this direction",
        "dir8_refine":         "paper 6-direction label narrows candidates to 1",
        "dual_reference":      "two reference objects jointly identify target",
        "dist_ord":            "target is closest/farthest in this direction",
        "two_hop_referent":    "a third node uniquely points to target",
        "dual_hop_referent":   "two third nodes jointly identify target",
        "anchor_intro":        "source object is uniquely described as anchor",
        "type+dir8":           "type + exact direction unique",
        "type+dir8+dist_ord":  "type + direction + distance rank unique",
        "dir8+dist_ord":       "direction + distance rank unique",
        "yesno_fallback":      "path itself uniquely identifies target",
    }

    def _abs_ts() -> str:
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    def _dt_ms(a: str, b: str) -> int:
        fmt = '%Y-%m-%d %H:%M:%S.%f'
        try:
            return int((datetime.strptime(b, fmt) - datetime.strptime(a, fmt)).total_seconds() * 1000)
        except Exception:
            return -1

    def _is_retryable_llm_exc(exc: Exception) -> bool:
        s = str(exc).lower()
        keys = ("timed out", "timeout", "connection", "connect", "readtimeout", "apiconnectionerror")
        return any(k in s for k in keys)
    def _is_auth_error(exc_or_msg: Exception | str) -> bool:
        s = str(exc_or_msg).lower()
        keys = (
            "authenticationerror",
            "unauthorized",
            "invalid token",
            "invalid api key",
            "无效的令牌",
            "401",
            "one_api_error",
        )
        return any(k in s for k in keys)

    def _parse_verify_n(verify_text: str) -> int:
        m = re.search(r"n=(\d+)", str(verify_text or ""))
        return int(m.group(1)) if m else -1

    def _parse_verify_ids(verify_text: str):
        """解析 logic_verification 里的 ids=[...]（与 run_gap_pipeline_v6._run_verify 格式一致）。"""
        t = str(verify_text or "")
        i = t.find("ids=")
        if i < 0:
            return None
        rest = t[i + 4 :].strip()
        if not rest.startswith("["):
            return None
        depth = 0
        end = 0
        for j, ch in enumerate(rest):
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    end = j + 1
                    break
        if end == 0:
            return None
        try:
            import ast

            v = ast.literal_eval(rest[:end])
            if isinstance(v, (list, tuple)):
                return [str(x) for x in v]
        except Exception:
            pass
        return None

    def _is_verify_wrong_target(verify_text: str, verify_n: int, n3_id: str) -> bool:
        """
        n=1 但 verify 返回的唯一 id 不是 gap 的 n3（约束锁到同类另一个节点）。
        与「真·多候选 non-unique(n>1)」区分，便于日志与 forensics。
        """
        tid = str(n3_id or "").strip()
        if verify_n != 1 or not tid:
            return False
        ids = _parse_verify_ids(verify_text)
        if not ids:
            return False
        return tid not in ids

    def _is_verified_unique(verify_text: str, verify_n: int) -> bool:
        if verify_n != 1:
            return False
        t = str(verify_text or "").lower()
        if any(bad in t for bad in ("[fail]", "❌", "not unique", "n=0")):
            return False
        return True

    def _reconcile_q_type_eff(pool_eff: str, question: str) -> str:
        """LLM 句式常与随机 q_type 不一致：按问句字面收紧有效题型。"""
        q = " ".join(str(question or "").strip().lower().split())
        if not q:
            return pool_eff
        if q.startswith("how many "):
            return "count"
        if any(
            w in q
            for w in ("closest", "farthest", "closer", "farther", "further")
        ):
            return "comparison"
        if q.startswith("is ") or q.startswith("are "):
            return "exist"
        # 池子抽到 status，但 LLM 生成的是「类型辨认」句（无 literal status）
        if pool_eff == "status" and (
            "what type" in q or "what kind" in q or "which type" in q
        ):
            return "object"
        return pool_eff

    def _qtype_semantic_ok(q_type: str, question: str, method: str = "") -> bool:
        """语义审计：问题文本是否匹配 q_type；fallback 方法生成的是降级题，不做严格句式匹配。"""
        if method in ("yesno_fallback", "count_fallback", "emergency_fallback"):
            return True

        q = " ".join(str(question or "").strip().lower().split())
        if not q:
            return False
        if q_type == "count":
            return q.startswith("how many ")
        if q_type == "exist":
            return (
                q.startswith("is there ")
                or q.startswith("are there ")
                or q.startswith("is the ")
                or q.startswith("is a ")
                or q.startswith("are the ")
                or q.startswith("does ")
            )
        if q_type == "status":
            return q.startswith("what ") and (
                "status" in q
                or "moving" in q
                or "stopped" in q
                or "parked" in q
                or "doing" in q
                or "state" in q
            )
        if q_type == "comparison":
            return (
                "closer" in q
                or "farther" in q
                or "further" in q
                or "closest" in q
                or "farthest" in q
                or "same" in q
                or "different" in q
                or "do the" in q
                or "does the" in q
            )
        # object
        return q.startswith("what ")

    def _scene_type_distribution() -> dict:
        q = (
            "MATCH (n:Object) "
            "RETURN n.type AS t, count(n) AS c "
            "ORDER BY c DESC LIMIT 12"
        )
        out = {}
        try:
            with driver.session() as sess:
                for rec in sess.run(q):
                    t = rec.get("t", "")
                    c = rec.get("c", 0)
                    if t:
                        out[str(t)] = int(c)
        except Exception:
            pass
        return out

    tracker = CoverageTracker()
    with driver.session() as sess:
        tracker.init_from_session(sess)

    _scenes = _json.loads((TRAINVAL / "scene.json").read_text())
    _samples = _json.loads((TRAINVAL / "sample.json").read_text())
    _st2name = {s["token"]: s["name"] for s in _scenes}
    _scene_samps = sorted(
        [s for s in _samples if _st2name.get(s["scene_token"], "") == SCENE_ID],
        key=lambda s: s["timestamp"],
    )
    sample_token = (_scene_samps[FRAME_ID]["token"] if len(_scene_samps) > FRAME_ID else "")

    _load_vb = os.getenv("VQA_TRACKER_LOAD_VAL_BASELINE", "true").lower() in ("true", "1", "yes", "on")
    if _load_vb and sample_token and QA_PATH.is_file():
        _vb = tracker.load_nuscenes_qa_baseline(
            str(QA_PATH), SCENE_ID, sample_tokens={sample_token}
        )
        print(
            f"  Val baseline → tracker (sample={sample_token[:16]}...): "
            f"L0+={_vb.get('n_L0', 0)} L1+={_vb.get('n_L1', 0)} L2+={_vb.get('n_L2', 0)}"
        )
    elif _load_vb and not sample_token:
        print("  [WARN] Val baseline skip: no sample_token for this scene/frame")

    tracker.dump_universe_snapshot(
        scene_id=SCENE_ID,
        frame_id=FRAME_ID,
        out_dir="output/coverage_snapshots",
    )

    def _print_stats(label: str):
        s = tracker.stats()
        print(f"\n  [{label}]")
        for lvl in ("L0", "L1", "L2A", "L2B"):
            v = s[lvl]
            print(f"    {lvl:<4}: {v['gap']:>4} gap / {v['total']:>4} total "
                  f"(covered={v['covered']}, rate={v['rate']:.1f}%)")

    _print_stats("Gap Stats — Initial")

    print(f"  sample_token for JSON: {sample_token[:16]}...")
    scene_dist_global = _scene_type_distribution()
    print(f"  scene type distribution: {scene_dist_global}")

    _tls = threading.local()
    def _get_worker_llm():
        _llm = getattr(_tls, "llm", None)
        if _llm is None:
            from gap_pipeline.llm_client import LLMClient as _LLMClient
            _llm = _LLMClient()
            _tls.llm = _llm
        return _llm

    def _gap_score(cell):
        path = cell.get("path_pattern", "")
        ego_p = 0.5 if "ego" in path else 1.0
        nodes = [cell.get("n1_id",""), cell.get("n2_id",""), cell.get("n3_id","")]
        unc = sum(1 for n in nodes if n and tracker._L0.get(n, CoverageRecord()).hit_count == 0)
        return ego_p * (unc + 1)

    q_type_pool = []
    for qt, wt in _Q_TYPE_WEIGHTS.items():
        q_type_pool.extend([qt] * wt)

    generated = 0
    total_att = 0
    dropped_fast = 0
    dropped_empty = 0
    dropped_non_unique = 0
    dropped_wrong_verify_target = 0
    dropped_low_iter = 0
    dropped_qtype_mismatch = 0
    fail_counts: dict = {}
    used_types: list = []
    nusqa_records: list = []

    def _all_covered() -> bool:
        s = tracker.stats()
        return all(s[lvl]["gap"] == 0 for lvl in ("L0", "L1", "L2A", "L2B"))

    round_idx = 0
    empty_streak = 0
    gap_fail_counts: dict = {}
    while True:
        if _all_covered():
            print(f"\n  [OK] Round {round_idx}: All coverage levels reached 100%.")
            break
        if MAX_ROUNDS > 0 and round_idx >= MAX_ROUNDS:
            print(f"\n  [M0] Reached MAX_ROUNDS={MAX_ROUNDS}, stopping for forensics.")
            break
        round_idx += 1
        raw_l2b = list(tracker.get_gap_cells("L2B"))
        raw_l2a = list(tracker.get_gap_cells("L2A"))

        def _pass_cooldown(c):
            if GAP_FAIL_COOLDOWN <= 0:
                return True
            return gap_fail_counts.get(c.get("_key", ""), 0) < GAP_FAIL_COOLDOWN

        l2b_avail = [c for c in raw_l2b if _pass_cooldown(c)]
        l2a_avail = [c for c in raw_l2a if _pass_cooldown(c)]
        if (
            GAP_FAIL_COOLDOWN > 0
            and not l2b_avail
            and not l2a_avail
            and (raw_l2b or raw_l2a)
        ):
            print(
                f"\n  [INFO] 全部 gap cell 处于失败冷却 (≥{GAP_FAIL_COOLDOWN} 次失败), "
                f"重置本帧失败计数后继续"
            )
            gap_fail_counts.clear()
            l2b_avail = raw_l2b
            l2a_avail = raw_l2a

        sb = sorted(l2b_avail, key=_gap_score, reverse=True)
        sa = sorted(l2a_avail, key=_gap_score, reverse=True)
        n_b = min(4, len(sb))
        n_a = min(4, len(sa))
        round_cells = [("L2B", c) for c in sb[:n_b]] + [("L2A", c) for c in sa[:n_a]]
        if not round_cells:
            print(
                f"\n  [WARN] No selectable gap cells while coverage <100% "
                f"(scene={SCENE_ID}, frame={FRAME_ID}). Stopping this frame."
            )
            break
        round_cells = round_cells[:Q_BATCH_SIZE]

        print(f"\n  [Round {round_idx}] batch={len(round_cells)} "
              f"(L2B={sum(1 for t,_ in round_cells if t=='L2B')}, "
              f"L2A={sum(1 for t,_ in round_cells if t=='L2A')}) "
              f"written={generated}")

        batch_llm_ms = []
        batch_neo_ms = []
        batch_total_ms = []
        round_batch_id = f"{SCENE_ID}_f{FRAME_ID}_r{round_idx}"

        # 先分配q_type，避免线程中竞争
        assigned_cells = []
        for topology, cell in round_cells:
            total_att += 1
            rem = [t for t in q_type_pool if t not in used_types]
            if not rem:
                rem = list(_Q_TYPE_WEIGHTS.keys())
                used_types = []
            q_type = random.choice(rem)
            used_types.append(q_type)
            assigned_cells.append((topology, cell, q_type))

        def _cell_worker(_topology, _cell, _q_type):
            _llm = _get_worker_llm()
            _gap_key = _cell["_key"]
            _path = _cell.get("path_pattern", "?")
            _cell_key = _canonical_cell_key(
                _cell.get("n1_id", "?"),
                _cell.get("n2_id", "?"),
                _cell.get("n3_id", "?"),
            )

            # [物理采样点1]
            _ts_start = _abs_ts()

            # LLM context
            if _topology == "L2A":
                _cypher = _llm.generate_l2a_context_cypher(_cell)
            else:
                _cypher = _llm.generate_l2b_obj_context_cypher(_cell)
            # [物理采样点2]
            _ts_llm = _llm.last_call_meta.get("timestamp_llm") or _abs_ts()
            _llm_ms = float(_llm.last_call_timing.get("total_ms", 0.0))

            _local_chain = CumulativeConstraintChain()
            _qa, _timing = _process_single_cell(
                cell=_cell, topology=_topology, cypher=_cypher,
                driver=driver, chain=_local_chain,
                scene_name=SCENE_ID, frame_idx=FRAME_ID,
                llm_timing=dict(_llm.last_call_timing),
                render_local_question=False,
            )
            _ts_cypher = _timing.get("timestamp_cypher_return", "") or _abs_ts()
            _neo_ms = float(_timing.get("neo4j_ms", 0.0))
            if _qa is None:
                return {"ok": False, "stage": "neo4j", "gap_key": _gap_key, "path": _path}

            _method = _timing.get("method_used", "") or "path"
            _constraint_desc = _CONSTRAINT_DESC_MAP.get(
                _method, f"constraint chain method={_method} identifies target"
            )
            # 约束链兜底方法映射到真实题型（否则上层会把 fallback 结果当成普通 object/status）
            if _method == "count_fallback":
                _q_type_eff = "count"
            elif _method in ("yesno_fallback", "emergency_fallback"):
                _q_type_eff = "exist"
            else:
                _q_type_eff = _q_type

            _verify_cypher = _timing.get("verify_cypher", "") or ""
            if not _verify_cypher.strip() or not _path.strip():
                return {"ok": False, "stage": "empty", "gap_key": _gap_key, "path": _path}
            _verify_text = str(_timing.get("logic_verification", "") or "")
            _verify_n = _parse_verify_n(_verify_text)
            _verified_unique = _is_verified_unique(_verify_text, _verify_n)
            _chain_unique = bool(_timing.get("is_unique", False))
            _target_n3 = str(_cell.get("n3_id") or "").strip()
            _verify_ids = _parse_verify_ids(_verify_text)
            _wrong_verify_target = _is_verify_wrong_target(
                _verify_text, _verify_n, _target_n3
            )
            _is_non_unique = (not _chain_unique or not _verified_unique)
            _fwu({"stage": "upper", "cell_key": _cell_key, "method_name": _method, "method_used": _method, "chain_is_unique": _chain_unique, "verified_unique": _verified_unique, "verify_n": _verify_n, "verify_ids": _verify_ids, "target_n3": _target_n3, "wrong_verify_target": _wrong_verify_target, "verify_cypher": _verify_cypher, "is_non_unique": _is_non_unique, "q_type_eff": _q_type_eff})
            if (
                _wrong_verify_target
                and _q_type_eff not in ("count", "exist")
                and STRICT_UNIQUE_ONLY
            ):
                _fwu({
                    "stage": "verdict",
                    "cell_key": _cell_key,
                    "final_verdict": "drop-wrong-target",
                    "verify_n": _verify_n,
                    "target_n3": _target_n3,
                    "verify_ids": _verify_ids,
                    "method_name": _method,
                    "method_used": _method,
                })
                return {
                    "ok": False,
                    "stage": "wrong_verify_target",
                    "gap_key": _gap_key,
                    "path": _path,
                    "verify": _verify_text,
                    "method": _method,
                    "target_n3": _target_n3,
                    "verify_ids": _verify_ids,
                    "verify_n": _verify_n,
                }
            if _is_non_unique and _q_type_eff not in ("count", "exist") and STRICT_UNIQUE_ONLY:
                _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "drop-non-unique", "verify_n": _verify_n, "method_name": _method, "method_used": _method})
                return {
                    "ok": False,
                    "stage": "non_unique",
                    "gap_key": _gap_key,
                    "path": _path,
                    "verify": _verify_text,
                    "method": _method,
                }

            _work_qt = _q_type_eff
            _answer = _answer_for_q_type(_work_qt, _cell, verify_n=_verify_n)

            def _gen_q_once(qt: str) -> str:
                try:
                    return _llm.generate_question_nlp_strict(
                        path=_path,
                        q_type=qt,
                        n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                        n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                        n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                        n3_status=_cell.get("n3_status",""),
                        r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                        r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                        constraint_desc=_constraint_desc,
                        answer=str(_answer),
                        scene_distribution=scene_dist_global,
                    )
                except Exception as _q_exc:
                    if _is_retryable_llm_exc(_q_exc):
                        time.sleep(0.8)  # 单次快速退避
                        return _llm.generate_question_nlp_strict(
                            path=_path,
                            q_type=qt,
                            n1_id=_cell.get("n1_id",""), n1_type=_cell.get("n1_type",""),
                            n2_id=_cell.get("n2_id",""), n2_type=_cell.get("n2_type",""),
                            n3_id=_cell.get("n3_id",""), n3_type=_cell.get("n3_type",""),
                            n3_status=_cell.get("n3_status",""),
                            r1_dir=_cell.get("r1_dir8") or _cell.get("r1_dir4","front"),
                            r2_dir=_cell.get("r2_dir8") or _cell.get("r2_dir4","front"),
                            constraint_desc=_constraint_desc,
                            answer=str(_answer),
                            scene_distribution=scene_dist_global,
                        )
                    raise

            _question = ""
            _semantic_ok = False
            for _ in range(max(1, MAX_QTYPE_RETRY)):
                _answer = _answer_for_q_type(_work_qt, _cell, verify_n=_verify_n)
                _question = _gen_q_once(_work_qt)
                _eff = _reconcile_q_type_eff(_q_type_eff, _question)
                if _qtype_semantic_ok(_eff, _question, method=_method):
                    _q_type_eff = _eff
                    _semantic_ok = True
                    break
                _work_qt = _eff
            if _semantic_ok:
                _answer = _answer_for_q_type(_q_type_eff, _cell, verify_n=_verify_n)
            if not _semantic_ok:
                _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "drop-qtype", "q_type_eff": _q_type_eff, "rendered_question": _question, "method_name": _method, "method_used": _method})
                return {
                    "ok": False,
                    "stage": "qtype_mismatch",
                    "gap_key": _gap_key,
                    "path": _path,
                    "q_type": _q_type_eff,
                    "question": _question,
                }

            # [物理采样点4]
            _ts_end = _abs_ts()
            _total_ms = _dt_ms(_ts_start, _ts_end)
            if _total_ms >= 0 and _total_ms < MIN_REAL_MS:
                return {"ok": False, "stage": "fast", "gap_key": _gap_key, "path": _path, "total_ms": _total_ms}

            _trace = _timing.get("constraint_trace", "")
            # V24: 优先使用管线内物理计数（CumulativeConstraintChain.trace_log 长度）
            _iter_count = int(_timing.get("constraint_rounds") or 0)
            if _iter_count < 1:
                _parts = [p.strip() for p in _trace.replace("->", "->").split("->")]
                _iter_count = max(1, len([p for p in _parts if p and p != "Path"]))
            if _iter_count < MIN_ITER_COUNT and _method in LOW_ITER_STRICT_METHODS:
                _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "drop-low-iter", "iteration_count": _iter_count, "method_name": _method, "method_used": _method})
                return {
                    "ok": False,
                    "stage": "low_iter",
                    "gap_key": _gap_key,
                    "path": _path,
                    "iteration_count": _iter_count,
                    "trace": _trace,
                }
            _n1 = _cell.get("n1_id",""); _n2 = _cell.get("n2_id",""); _n3 = _cell.get("n3_id","")
            _l0 = [x for x in [_n1, _n2, _n3] if x]
            _l1 = ([{"source": _n1, "target": _n2}] if _n1 and _n2 else []) + \
                  ([{"source": _n2, "target": _n3}] if _n2 and _n3 else [])
            _l2 = [{"o1": _n1, "o2": _n2, "o3": _n3}] if all([_n1, _n2, _n3]) else []
            _gen_id = make_generated_question_id(SCENE_ID, FRAME_ID)
            _fwu({"stage": "verdict", "cell_key": _cell_key, "final_verdict": "kept", "verify_n": _verify_n, "method_name": _method, "method_used": _method, "rendered_question": _question, "q_type_eff": _q_type_eff})
            return {
                "ok": True,
                "gen_id": _gen_id,
                "gap_key": _gap_key,
                "path": _path,
                "topology": _topology,
                "q_type": _q_type_eff,
                "answer": str(_answer),
                "question": _question,
                "verify_cypher": _verify_cypher,
                "verify_text": _verify_text,
                "verify_n": _verify_n,
                "trace": _trace,
                "iteration_count": _iter_count,
                "method": _method,
                "is_unique": _timing.get("is_unique", False),
                "ts_start": _ts_start,
                "ts_llm": _ts_llm,
                "ts_cypher_return": _ts_cypher,
                "ts_end": _ts_end,
                "total_ms": _total_ms,
                "llm_ms": _llm_ms,
                "neo_ms": _neo_ms,
                "l0": _l0, "l1": _l1, "l2": _l2,
            }

        with ThreadPoolExecutor(max_workers=Q_N_WORKERS) as _ex:
            _futs = {}
            for _t, _c, _q in assigned_cells:
                _fu = _ex.submit(_cell_worker, _t, _c, _q)
                _futs[_fu] = (_c.get("_key", ""), _c.get("path_pattern", "?"))
            for _f in as_completed(_futs):
                _gap_key, _path = _futs.get(_f, ("", "?"))
                try:
                    _res = _f.result()
                except Exception as exc:
                    _msg = str(exc)
                    if _is_auth_error(_msg):
                        raise RuntimeError(
                            f"LLM authentication failed ({_msg}). "
                            f"Please set a valid school API key (see VQA_MODEL_NAME, default Qwen3.5-35B-A3B)."
                        ) from exc
                    if _gap_key:
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        if GAP_FAIL_COOLDOWN > 0:
                            gap_fail_counts[_gap_key] = gap_fail_counts.get(_gap_key, 0) + 1
                    print(f"    [FAIL] [worker-exc] {_path} {_console_safe(_msg)}")
                    continue
                _path = _res.get("path", _path)
                _gap_key = _res.get("gap_key", _gap_key)
                if not _res.get("ok"):
                    if GAP_FAIL_COOLDOWN > 0 and _gap_key:
                        gap_fail_counts[_gap_key] = gap_fail_counts.get(_gap_key, 0) + 1
                    _stage = _res.get("stage", "unknown")
                    if _stage == "fast":
                        dropped_fast += 1
                        print(f"    [WARN] [drop-fast] {_path} total={_res.get('total_ms',-1)}ms < {MIN_REAL_MS}ms")
                    elif _stage == "empty":
                        dropped_empty += 1
                        print(f"    [WARN] [drop-empty] path={_path}")
                    elif _stage == "wrong_verify_target":
                        dropped_wrong_verify_target += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(
                            f"    [WARN] [drop-wrong-target] {_path} "
                            f"expected={_res.get('target_n3', '?')} "
                            f"verify_n={_res.get('verify_n', -1)} "
                            f"ids={_res.get('verify_ids', '?')} "
                            f"raw={_console_safe(_res.get('verify', ''))}"
                        )
                    elif _stage == "non_unique":
                        dropped_non_unique += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(f"    [WARN] [drop-non-unique] {_path} verify={_console_safe(_res.get('verify', ''))}")
                    elif _stage == "low_iter":
                        dropped_low_iter += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(f"    [WARN] [drop-low-iter] {_path} iter={_res.get('iteration_count',-1)} < {MIN_ITER_COUNT}")
                    elif _stage == "qtype_mismatch":
                        dropped_qtype_mismatch += 1
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(
                            f"    [WARN] [drop-qtype] {_path} q_type={_res.get('q_type', '?')} "
                            f"q={_console_safe(_res.get('question', ''))}"
                        )
                    else:
                        fail_counts[_gap_key] = fail_counts.get(_gap_key, 0) + 1
                        print(f"    [FAIL] [{_stage}] {_path}")
                    continue

                if GAP_FAIL_COOLDOWN > 0 and _res.get("gap_key"):
                    gap_fail_counts.pop(_res["gap_key"], None)

                ok = write_generated_question(
                    scene_id=SCENE_ID, frame_id=FRAME_ID, question_id=_res["gen_id"],
                    timestamp_start=_res["ts_start"],
                    timestamp_llm=_res["ts_llm"],
                    timestamp_cypher_return=_res["ts_cypher_return"],
                    timestamp_end=_res["ts_end"],
                    iteration_count=_res["iteration_count"],
                    question_type=_res["q_type"],
                    complexity="L2",
                    natural_language_question=_res["question"],
                    cypher_question=_res["verify_cypher"],
                    answer=_res["answer"],
                    l0_nodes=_res["l0"], l1_edges=_res["l1"], l2_paths=_res["l2"],
                    target_gap_cell=_res["path"],
                    batch_id=round_batch_id,
                )
                if ok:
                    generated += 1
                nusqa_records.append({
                    "sample_token": sample_token,
                    "question": _res["question"],
                    "answer": _res["answer"],
                    "template_type": _res["q_type"],
                    "num_hop": 2,
                    "question_id": _res["gen_id"],
                    "scene_name": SCENE_ID,
                    "frame_idx": FRAME_ID,
                    "topology_level": _res["topology"],
                    "path_pattern": _res["path"],
                    "constraint_trace": _res["trace"],
                    "iteration_count": _res["iteration_count"],
                    "is_unique": _res["is_unique"],
                    "method_used": _res["method"],
                    "verify_cypher": _res["verify_cypher"],
                    "verify_result": _res.get("verify_text", ""),
                    "verify_n": _res.get("verify_n", -1),
                    "timestamp_start": _res["ts_start"],
                    "timestamp_llm": _res["ts_llm"],
                    "timestamp_cypher_return": _res["ts_cypher_return"],
                    "timestamp_end": _res["ts_end"],
                    "total_ms": _res["total_ms"],
                    "batch_id": round_batch_id,
                })
                tracker.record_from_qa({
                    "topology_level": _res["topology"],
                    "path_pattern": _res["path"],
                    "template_id": f"v19_{_res['q_type']}_{_res['method']}",
                    "question_id": _res["gen_id"],
                })
                batch_llm_ms.append(_res["llm_ms"])
                batch_neo_ms.append(_res["neo_ms"])
                batch_total_ms.append(_res["total_ms"] if _res["total_ms"] >= 0 else 0.0)
                print(f"    #{generated:03d} [{_res['topology']}] iter={_res['iteration_count']} "
                      f"q={_res['q_type']:<10} v={_res.get('verify_n',-1)} "
                      f"dt={_res['total_ms']}ms batch={round_batch_id}")

        if batch_total_ms:
            avg_llm = int(sum(batch_llm_ms) / len(batch_llm_ms))
            avg_neo = int(sum(batch_neo_ms) / len(batch_neo_ms))
            avg_total = int(sum(batch_total_ms) / len(batch_total_ms))
            status = "REAL" if avg_total >= MIN_REAL_MS else "SUSPECT"
            print(f"[Batch Verify] Round {round_idx} | "
                  f"LLM Latency: {avg_llm}ms | Neo4j Latency: {avg_neo}ms | "
                  f"Total: {avg_total}ms | Status: {status}")
        else:
            print(f"[Batch Verify] Round {round_idx} | Status: EMPTY")

        if batch_total_ms:
            empty_streak = 0
        else:
            empty_streak += 1
            if MAX_EMPTY_STREAK > 0 and empty_streak >= MAX_EMPTY_STREAK:
                print(
                    f"\n  [WARN] 连续 {empty_streak} 轮 Batch EMPTY，"
                    f"达到 VQA_MAX_EMPTY_STREAK={MAX_EMPTY_STREAK}，结束本帧 Gap。"
                )
                break

    qa_dir = pathlib.Path(GEN_QA_DIR)
    qa_dir.mkdir(parents=True, exist_ok=True)
    qa_path = qa_dir / f"{SCENE_ID}_frame{FRAME_ID}_qa.json"
    qa_path.write_text(
        _json.dumps({"questions": nusqa_records}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n  NuScenes-QA JSON -> {qa_path} ({len(nusqa_records)} records)")

    _print_stats("Gap Stats — Final")
    n_l2b_gen = sum(1 for r in nusqa_records if r["topology_level"] == "L2B")
    n_l2a_gen = len(nusqa_records) - n_l2b_gen
    print(f"\n  V18 Generated: {generated}/{total_att} attempts written")
    print(f"  L2B: {n_l2b_gen}  L2A: {n_l2a_gen}  "
          f"L2B share: {n_l2b_gen/max(len(nusqa_records),1)*100:.0f}%")
    print(f"  Dropped (<{MIN_REAL_MS}ms): {dropped_fast}")
    print(f"  Dropped (empty fields): {dropped_empty}")
    print(f"  Dropped (verify n=1 but wrong target id): {dropped_wrong_verify_target}")
    print(f"  Dropped (non-unique / verify fail): {dropped_non_unique}")
    print(f"  Dropped (iteration<{MIN_ITER_COUNT}): {dropped_low_iter}")
    print(f"  Dropped (q_type semantic mismatch): {dropped_qtype_mismatch}")
    print(
        f"  Quality gates: strict_unique={STRICT_UNIQUE_ONLY} "
        f"min_iter={MIN_ITER_COUNT} low_iter_methods={sorted(LOW_ITER_STRICT_METHODS)}"
    )
    print(f"  Question types used: {list(dict.fromkeys(used_types))}")
    return generated


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  Method A — Full Execution Chain (scene-0926 frame-20)")
    print("=" * 65)

    if not maybe_connect_school_vpn():
        print("\n[FAIL] VPN 预连失败或未配置。关闭 VQA_AUTO_CONNECT_VPN 可跳过此步。")
        return

    # Pre-flight
    if not preflight_check():
        print("\n[FAIL] Pre-flight failed. Fix issues above and re-run.")
        return

    if os.getenv("VQA_CLEAR_RUN_DATA", "").lower() in ("1", "true", "yes"):
        from advtest_cleanup import clear_excel_data_rows
        n = clear_excel_data_rows(EXCEL_PATH)
        print(f"\n[VQA_CLEAR_RUN_DATA] Excel cleared (removed ~{n} row-slots across sheets).")

    print("\n[OK] All pre-flight checks passed. Starting execution...\n")

    from neo4j_bootstrap import ensure_neo4j_listening
    if not ensure_neo4j_listening(NEO4J_URI):
        print("\n[FAIL] Neo4j 不可用。请启动数据库或配置 NEO4J_DOCKER_NAMES / Docker。")
        return

    from neo4j import GraphDatabase
    from gap_pipeline.llm_client import LLMClient

    llm    = LLMClient()
    try:
        _ = llm._call(
            "ping",
            system_prompt="Return only: pong",
            max_tokens=8,
            call_tag="audit_healthcheck",
        )
    except Exception as exc:
        print(f"\n[FAIL] LLM API health check failed: {exc}")
        print("   当前是鉴权/账号问题（不是本地流程逻辑问题）。")
        print("   请先配置有效 VQA_API_KEY，并确认 VQA_MODEL_NAME 与网关一致（默认 Qwen3.5-35B-A3B）。")
        return
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PWD))

    try:
        # append-only policy: never wipe historical evidence unless explicitly opted in
        step0_wipe_previous_run()
        step1_cleanup()
        if not step1_5_refresh_filtered_sg():
            print("\n[FAIL] Cannot rebuild filtered SG with official policy.")
            return
        step2_filter_record()
        step3_import_neo4j()
        step4_baseline_audit(driver, llm)
        step5_6_generate(driver, llm)

        print("\n" + "=" * 65)
        print("  [OK] Method A complete")
        print("  Check RQ.xlsx:")
        print("    filter_record          ← upsert 1 row / frame (overwrite if exists)")
        print("    raw_coverage           ← ~29 new rows (baseline)")
        print("    question-answer-our    ← 5 new rows (generated)")
        print("=" * 65)

    finally:
        driver.close()


if __name__ == "__main__":
    main()
