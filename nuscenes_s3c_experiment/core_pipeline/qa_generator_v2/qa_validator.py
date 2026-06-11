"""
QA Validator — 自动化找错题

流程:
  1. 接收生成的 QA 列表
  2. 对每道题发送给 VLM 验证 (可选)
  3. 比较 VLM 回答与标准答案
  4. 不一致 → 重试 (最多 max_retries 次)
  5. 超过重试上限 → 标记为垃圾题
  6. 输出垃圾题 Excel 报告

支持两种模式:
  - 在线模式: 调用 VLM API 实际验证
  - 离线模式: 从已有验证结果文件加载

增强特性:
  - 断点续跑: 每题保存进度，中断后可恢复
  - VLM 答案智能解析: 清理 think 标签、提取核心答案
  - 歧义题检测: VLM 多次回答不一致 → 标记为 ambiguous
  - 丰富合理性检查: 重复题、答案与问题类型一致性
  - Excel 全量导出: 正确题+垃圾题，带颜色标注
  - 精细统计: per-template / per-level / per-type 精确率
"""

import json
import logging
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Callable
from dataclasses import dataclass, field
from datetime import datetime
from collections import Counter, defaultdict

logger = logging.getLogger(__name__)


@dataclass
class ValidationRecord:
    """单道题的验证记录"""
    question_id: str = ""
    template_id: str = ""
    coverage_level: str = ""
    question_type: str = ""
    question: str = ""
    expected_answer: str = ""
    vlm_answers: List[str] = field(default_factory=list)
    vlm_parsed_answers: List[str] = field(default_factory=list)
    retry_count: int = 0
    is_correct: bool = False
    is_garbage: bool = False
    is_ambiguous: bool = False
    error_type: str = ""       # mismatch / timeout / parse_error / ambiguous
                                # / empty_qa / sanity_fail / duplicate
    notes: str = ""
    validation_time: float = 0.0   # 单题验证耗时 (秒)
    covered_elements: List[str] = field(default_factory=list)


class QAValidator:
    """
    QA 验证器

    Usage:
        validator = QAValidator(max_retries=5)
        results = validator.validate_batch(qa_list, verify_fn=my_vlm_call)
        validator.export_garbage_excel("garbage_questions.xlsx")

        # 或直接从生成结果验证
        result = generator.generate_with_tracker()
        validator.validate_generation_result(result, verify_fn=my_vlm)
    """

    def __init__(self, max_retries: int = 5, answer_match_fn: Callable = None,
                 checkpoint_dir: str = None, auto_save_every: int = 10):
        """
        Args:
            max_retries: 最大重试次数，超过则标记为垃圾题
            answer_match_fn: 答案匹配函数 (expected, actual) -> bool
                             默认使用宽松匹配
            checkpoint_dir: 断点保存目录，若指定则每验证 auto_save_every 题自动保存
            auto_save_every: 每多少题自动保存一次
        """
        self.max_retries = max_retries
        self.answer_match_fn = answer_match_fn or self._default_answer_match
        self.checkpoint_dir = checkpoint_dir
        self.auto_save_every = auto_save_every
        self.records: List[ValidationRecord] = []
        self._stats = {
            "total": 0,
            "correct": 0,
            "garbage": 0,
            "retried": 0,
            "ambiguous": 0,
        }
        self._seen_questions: set = set()  # 用于查重

    # ========================================================================
    #  批量验证
    # ========================================================================

    def validate_batch(self, qa_list: List[Dict],
                       verify_fn: Callable = None,
                       progress_callback: Callable = None,
                       resume: bool = True) -> List[ValidationRecord]:
        """
        批量验证 QA 列表

        Args:
            qa_list: QA 字典列表 (含 question, answer, template_id 等)
            verify_fn: VLM 验证函数 (question: str) -> str
                       若为 None 则跳过在线验证，仅做格式检查
            progress_callback: 进度回调 (current, total, record)
            resume: 是否尝试从断点恢复

        Returns:
            ValidationRecord 列表
        """
        # 断点恢复
        start_idx = 0
        if resume and self.checkpoint_dir:
            start_idx = self._try_resume()

        if start_idx == 0:
            self.records = []
            self._stats = {"total": len(qa_list), "correct": 0,
                           "garbage": 0, "retried": 0, "ambiguous": 0}
            self._seen_questions = set()
        else:
            self._stats["total"] = len(qa_list)

        for i in range(start_idx, len(qa_list)):
            qa = qa_list[i]
            record = self._validate_single(qa, verify_fn)
            self.records.append(record)

            if progress_callback:
                progress_callback(i + 1, len(qa_list), record)

            if (i + 1) % self.auto_save_every == 0:
                logger.info(f"验证进度: {i+1}/{len(qa_list)}, "
                            f"正确={self._stats['correct']}, "
                            f"垃圾={self._stats['garbage']}, "
                            f"歧义={self._stats['ambiguous']}")
                if self.checkpoint_dir:
                    self._save_checkpoint(i + 1)

        # 最终保存
        if self.checkpoint_dir:
            self._save_checkpoint(len(qa_list))

        logger.info(f"验证完成: {self._stats}")
        return self.records

    def validate_generation_result(self, gen_result,
                                    verify_fn: Callable = None) -> List[ValidationRecord]:
        """
        直接从 CoverageDrivenTemplateGenerator.generate_with_tracker() 的结果验证

        Args:
            gen_result: GenerationResult (含 .questions 列表)
            verify_fn: VLM 验证函数
        """
        qa_list = gen_result.questions if hasattr(gen_result, 'questions') else gen_result
        return self.validate_batch(qa_list, verify_fn=verify_fn)

    def _validate_single(self, qa: Dict, verify_fn: Callable = None) -> ValidationRecord:
        """验证单道题"""
        t0 = time.time()
        record = ValidationRecord(
            question_id=qa.get("question_id", f"q_{len(self.records)+1:04d}"),
            template_id=qa.get("template_id", ""),
            coverage_level=qa.get("coverage_level", ""),
            question_type=qa.get("question_type", ""),
            question=qa.get("question", ""),
            expected_answer=str(qa.get("answer", "")),
            covered_elements=qa.get("covered_elements", []),
        )

        # --- 检查阶段 1: 空值 ---
        if not record.question or not record.expected_answer:
            record.is_garbage = True
            record.error_type = "empty_qa"
            record.notes = "问题或答案为空"
            self._stats["garbage"] += 1
            record.validation_time = time.time() - t0
            return record

        # --- 检查阶段 2: 合理性 ---
        sanity = self._sanity_check(qa)
        if sanity:
            record.is_garbage = True
            record.error_type = "sanity_fail"
            record.notes = sanity
            self._stats["garbage"] += 1
            record.validation_time = time.time() - t0
            return record

        # --- 检查阶段 3: 查重 ---
        q_norm = record.question.strip().lower()
        if q_norm in self._seen_questions:
            record.is_garbage = True
            record.error_type = "duplicate"
            record.notes = f"与已有问题重复: {record.question[:60]}"
            self._stats["garbage"] += 1
            record.validation_time = time.time() - t0
            return record
        self._seen_questions.add(q_norm)

        # --- 检查阶段 4: VLM 在线验证 ---
        if verify_fn is not None:
            for attempt in range(self.max_retries):
                try:
                    raw_answer = verify_fn(record.question)
                    parsed = self._parse_vlm_answer(raw_answer, record.question_type)
                    record.vlm_answers.append(raw_answer)
                    record.vlm_parsed_answers.append(parsed)
                    record.retry_count = attempt + 1

                    if self.answer_match_fn(record.expected_answer, parsed):
                        record.is_correct = True
                        self._stats["correct"] += 1
                        if attempt > 0:
                            self._stats["retried"] += 1
                        record.validation_time = time.time() - t0
                        return record

                except Exception as e:
                    record.vlm_answers.append(f"ERROR: {e}")
                    record.vlm_parsed_answers.append("")
                    record.retry_count = attempt + 1

            # --- 检查阶段 5: 歧义检测 ---
            valid_parsed = [a for a in record.vlm_parsed_answers if a]
            if len(set(valid_parsed)) >= 3:
                record.is_ambiguous = True
                record.error_type = "ambiguous"
                record.notes = (f"VLM {len(set(valid_parsed))}种不同回答, 问题可能歧义. "
                                f"期望: {record.expected_answer}, "
                                f"解析后: {valid_parsed}")
                self._stats["ambiguous"] += 1
            else:
                record.error_type = "max_retries_exceeded"
                record.notes = (f"VLM {self.max_retries}次回答均不匹配. "
                                f"期望: {record.expected_answer}, "
                                f"解析后: {valid_parsed}")
            record.is_garbage = True
            self._stats["garbage"] += 1
        else:
            # 无 VLM，仅做格式检查 → 视为通过
            record.is_correct = True
            self._stats["correct"] += 1

        record.validation_time = time.time() - t0
        return record

    # ========================================================================
    #  VLM 答案解析
    # ========================================================================

    @staticmethod
    def _parse_vlm_answer(raw: str, question_type: str = "") -> str:
        """
        从LLM/VLM原始输出中提取核心答案

        处理:
          - 移除 <think>...</think> 标签
          - 移除 markdown 代码块
          - 提取“The answer is X” 格式
          - bool 类型只保留 yes/no
          - count 类型只保留数字
        """
        if not raw:
            return ""

        text = raw

        # 清理 <think> 标签
        text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL | re.IGNORECASE)
        # 如果只有 <think> 没有闭合，截断
        idx = text.lower().find('<think>')
        if idx != -1:
            text = text[:idx]

        # 清理 markdown
        text = text.strip()
        if text.startswith('```'):
            text = re.sub(r'```\w*\n?', '', text)
        text = text.strip()

        # 提取 "The answer is X" / "Answer: X"
        m = re.search(r'(?:the\s+answer\s+is|answer\s*[:=])\s*(.+)', text, re.IGNORECASE)
        if m:
            text = m.group(1).strip().rstrip('.')

        # bool 类型: 提取第一个 yes/no
        if question_type in ("exist", "comparison"):
            m2 = re.search(r'\b(yes|no)\b', text, re.IGNORECASE)
            if m2:
                return m2.group(1).lower()

        # count 类型: 提取数字
        if question_type == "count":
            m3 = re.search(r'\b(\d+)\b', text)
            if m3:
                return m3.group(1)

        # 去掉末尾标点
        text = text.strip().rstrip('.!,')
        return text

    # ========================================================================
    #  答案匹配
    # ========================================================================

    @staticmethod
    def _default_answer_match(expected: str, actual: str) -> bool:
        """宽松答案匹配"""
        e = expected.strip().lower()
        a = actual.strip().lower()

        if not a:
            return False

        # 精确匹配
        if e == a:
            return True

        # 数字匹配 (允许小数误差)
        try:
            ef = float(e)
            af = float(a)
            if ef == 0 and af == 0:
                return True
            if abs(ef - af) < 0.5:
                return True
            # 百分比误差
            if ef != 0 and abs(ef - af) / abs(ef) < 0.1:
                return True
        except (ValueError, TypeError):
            pass

        # bool 匹配
        yes_set = {"yes", "true", "1", "correct", "right"}
        no_set = {"no", "false", "0", "incorrect", "wrong", "none"}
        if (e in yes_set and a in yes_set) or (e in no_set and a in no_set):
            return True

        # 包含匹配 (VLM 可能给出完整句子)
        if len(e) >= 2 and e in a:
            return True
        if len(a) >= 2 and a in e:
            return True

        # 对象类型别名 (car vs automobile, pedestrian vs person)
        alias_map = {
            "pedestrian": {"person", "walker", "people"},
            "car": {"automobile", "vehicle"},
            "truck": {"lorry"},
            "bicycle": {"bike", "cycle"},
            "motorcycle": {"motorbike"},
            "traffic_cone": {"cone", "traffic cone"},
            "construction_vehicle": {"construction vehicle"},
        }
        for canonical, aliases in alias_map.items():
            all_names = aliases | {canonical}
            if e in all_names and a in all_names:
                return True

        return False

    # ========================================================================
    #  合理性检查
    # ========================================================================

    @staticmethod
    def _sanity_check(qa: Dict) -> Optional[str]:
        """
        答案合理性检查，返回 None 表示通过，否则返回错误原因

        检查项:
          - count 答案应为非负整数
          - bool 答案应为 yes/no
          - 答案不应过长
          - 问题不应包含未填充的占位符
          - 答案与问题类型一致性
          - 问题长度检查
        """
        answer = str(qa.get("answer", ""))
        question = qa.get("question", "")
        answer_type = qa.get("answer_type", "")
        question_type = qa.get("question_type", "")

        # 检查未填充占位符
        if "{" in question and "}" in question:
            return f"问题包含未填充占位符: {question[:60]}"

        # 问题太短
        if len(question) < 5:
            return f"问题过短 ({len(question)} 字符): {question}"

        # count 类型答案
        if answer_type == "number" or question_type == "count":
            try:
                n = int(answer)
                if n < 0:
                    return f"count 答案为负数: {answer}"
            except ValueError:
                try:
                    f = float(answer)
                    if f < 0:
                        return f"count 答案为负数: {answer}"
                except ValueError:
                    return f"count 答案不是数字: {answer}"

        # bool 类型答案
        if answer_type == "bool" or question_type in ("exist", "comparison"):
            if answer.lower() not in ("yes", "no", "true", "false"):
                # exist/comparison 类型应该返回 yes/no，但允许某些特殊值
                # 如 "approaching" / "moving away" 对于 is_approaching 类型
                if question_type == "comparison" and answer.lower() in (
                        "approaching", "moving away", "same", "different"):
                    pass  # 允许
                elif question_type == "exist" and answer_type != "bool":
                    pass  # answer_type 覆盖
                else:
                    return f"bool 答案格式异常: {answer}"

        # 答案过长
        if len(answer) > 200:
            return f"答案过长 ({len(answer)} 字符)"

        # 问题是否包含无意义重复 (如连续相同单词)
        words = question.lower().split()
        for i in range(len(words) - 2):
            if words[i] == words[i+1] == words[i+2] and words[i] not in ('the', 'a', 'an', 'is', 'of'):
                return f"问题包含无意义重复: '{words[i]}' x3"

        return None

    # ========================================================================
    #  Excel 导出
    # ========================================================================

    def export_garbage_excel(self, path: str):
        """导出垃圾题到 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl 未安装，改用 CSV 导出")
            self.export_garbage_csv(path.replace(".xlsx", ".csv"))
            return

        wb = openpyxl.Workbook()

        # --- Sheet 1: Garbage Questions ---
        ws = wb.active
        ws.title = "Garbage Questions"
        self._write_records_sheet(ws, [r for r in self.records if r.is_garbage],
                                  Font, PatternFill, Alignment, is_garbage=True)

        # --- Sheet 2: All Questions ---
        ws_all = wb.create_sheet("All Questions")
        self._write_records_sheet(ws_all, self.records,
                                  Font, PatternFill, Alignment, is_garbage=False)

        # --- Sheet 3: Summary ---
        ws_sum = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_sum, Font, PatternFill)

        # --- Sheet 4: Per-Template Stats ---
        ws_tmpl = wb.create_sheet("Template Stats")
        self._write_template_stats_sheet(ws_tmpl, Font, PatternFill)

        # 保存
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        garbage_cnt = sum(1 for r in self.records if r.is_garbage)
        logger.info(f"Excel 已保存: {out} (全部 {len(self.records)} 条, 垃圾 {garbage_cnt} 条)")

    def _write_records_sheet(self, ws, records, Font, PatternFill, Alignment, is_garbage=False):
        """写入记录到 worksheet"""
        headers = [
            "序号", "Question ID", "Template ID", "Level", "Type",
            "Question", "Expected Answer", "VLM Parsed Answers",
            "Retry Count", "Status", "Error Type", "Notes"
        ]
        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 颜色定义
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for i, rec in enumerate(records, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=rec.question_id)
            ws.cell(row=row, column=3, value=rec.template_id)
            ws.cell(row=row, column=4, value=rec.coverage_level)
            ws.cell(row=row, column=5, value=rec.question_type)
            ws.cell(row=row, column=6, value=rec.question)
            ws.cell(row=row, column=7, value=rec.expected_answer)
            ws.cell(row=row, column=8, value="; ".join(rec.vlm_parsed_answers or rec.vlm_answers))
            ws.cell(row=row, column=9, value=rec.retry_count)

            if rec.is_correct:
                status = "✓ 正确"
                fill = green_fill
            elif rec.is_ambiguous:
                status = "⚠ 歧义"
                fill = yellow_fill
            else:
                status = "✗ 垃圾"
                fill = red_fill

            ws.cell(row=row, column=10, value=status)
            ws.cell(row=row, column=11, value=rec.error_type)
            ws.cell(row=row, column=12, value=rec.notes)

            # 行着色
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill

        # 列宽
        col_widths = [6, 16, 16, 6, 12, 50, 15, 30, 8, 10, 16, 40]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A'].width = w
        # 冻结首行
        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, ws, Font, PatternFill):
        """写入汇总 sheet"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=1, value="统计项").font = header_font
        ws.cell(row=1, column=2, value="值").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2).fill = header_fill

        total = self._stats.get("total", 0)
        correct = self._stats.get("correct", 0)
        garbage = self._stats.get("garbage", 0)

        summary = [
            ("总题数", total),
            ("正确题数", correct),
            ("垃圾题数", garbage),
            ("歧义题数", self._stats.get("ambiguous", 0)),
            ("重试题数", self._stats.get("retried", 0)),
            ("正确率", f"{correct/max(total,1)*100:.1f}%"),
            ("垃圾率", f"{garbage/max(total,1)*100:.1f}%"),
            ("最大重试次数", self.max_retries),
            ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        # 错误类型分布
        garbage_recs = [r for r in self.records if r.is_garbage]
        error_types = Counter(r.error_type for r in garbage_recs)
        summary.append(("", ""))
        summary.append(("━━━ 错误类型分布 ━━━", ""))
        for etype, cnt in error_types.most_common():
            summary.append((etype, cnt))

        # 按覆盖级别分布
        summary.append(("", ""))
        summary.append(("━━━ 按覆盖级别分布 ━━━", ""))
        for level in ["L0", "L1", "L2"]:
            level_recs = [r for r in self.records if r.coverage_level == level]
            level_ok = sum(1 for r in level_recs if r.is_correct)
            level_total = len(level_recs)
            pct = f"{level_ok/max(level_total,1)*100:.1f}%"
            summary.append((f"{level} 正确率", f"{level_ok}/{level_total} ({pct})"))

        # 按问题类型分布
        summary.append(("", ""))
        summary.append(("━━━ 按问题类型分布 ━━━", ""))
        for qtype in ["exist", "count", "status", "object", "comparison"]:
            t_recs = [r for r in self.records if r.question_type == qtype]
            t_ok = sum(1 for r in t_recs if r.is_correct)
            t_total = len(t_recs)
            pct = f"{t_ok/max(t_total,1)*100:.1f}%"
            summary.append((f"{qtype} 正确率", f"{t_ok}/{t_total} ({pct})"))

        for i, (k, v) in enumerate(summary, 2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25

    def _write_template_stats_sheet(self, ws, Font, PatternFill):
        """写入 per-template 统计"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        headers = ["Template ID", "Level", "总数", "正确", "垃圾", "歧义", "正确率"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # 汇总
        tmpl_data = defaultdict(lambda: {"total": 0, "correct": 0, "garbage": 0, "ambiguous": 0, "level": ""})
        for r in self.records:
            d = tmpl_data[r.template_id]
            d["total"] += 1
            d["level"] = r.coverage_level
            if r.is_correct:
                d["correct"] += 1
            if r.is_garbage:
                d["garbage"] += 1
            if r.is_ambiguous:
                d["ambiguous"] += 1

        # 按垃圾率降序
        sorted_tmpls = sorted(tmpl_data.items(),
                              key=lambda x: x[1]["garbage"]/max(x[1]["total"],1),
                              reverse=True)

        for i, (tid, d) in enumerate(sorted_tmpls, 1):
            row = i + 1
            pct = f"{d['correct']/max(d['total'],1)*100:.1f}%"
            ws.cell(row=row, column=1, value=tid)
            ws.cell(row=row, column=2, value=d["level"])
            ws.cell(row=row, column=3, value=d["total"])
            ws.cell(row=row, column=4, value=d["correct"])
            ws.cell(row=row, column=5, value=d["garbage"])
            ws.cell(row=row, column=6, value=d["ambiguous"])
            ws.cell(row=row, column=7, value=pct)

        ws.column_dimensions['A'].width = 22
        ws.freeze_panes = "A2"

    def export_garbage_csv(self, path: str):
        """CSV 后备导出"""
        import csv
        garbage = [r for r in self.records if r.is_garbage]

        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "序号", "Question ID", "Template ID", "Level", "Type",
                "Question", "Expected Answer", "VLM Answers",
                "Retry Count", "Error Type", "Notes"
            ])
            for i, rec in enumerate(garbage, 1):
                writer.writerow([
                    i, rec.question_id, rec.template_id, rec.coverage_level,
                    rec.question_type, rec.question, rec.expected_answer,
                    "; ".join(rec.vlm_answers), rec.retry_count,
                    rec.error_type, rec.notes,
                ])

        logger.info(f"垃圾题 CSV 已保存: {out} ({len(garbage)} 条)")

    # ========================================================================
    #  断点续跑
    # ========================================================================

    def _save_checkpoint(self, processed_count: int):
        """保存断点"""
        if not self.checkpoint_dir:
            return
        ckpt_path = Path(self.checkpoint_dir) / "validation_checkpoint.json"
        ckpt_path.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "processed_count": processed_count,
            "stats": self._stats,
            "max_retries": self.max_retries,
            "seen_questions": list(self._seen_questions),
            "records": [self._record_to_dict(r) for r in self.records],
            "timestamp": datetime.now().isoformat(),
        }
        with open(ckpt_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.debug(f"断点已保存: {processed_count} 题")

    def _try_resume(self) -> int:
        """尝试从断点恢复，返回已处理数量"""
        if not self.checkpoint_dir:
            return 0
        ckpt_path = Path(self.checkpoint_dir) / "validation_checkpoint.json"
        if not ckpt_path.exists():
            return 0

        try:
            with open(ckpt_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            processed = data.get("processed_count", 0)
            self._stats = data.get("stats", {})
            self._seen_questions = set(data.get("seen_questions", []))
            self.records = [self._dict_to_record(r) for r in data.get("records", [])]
            logger.info(f"从断点恢复: 已处理 {processed} 题")
            return processed
        except Exception as e:
            logger.warning(f"断点恢复失败: {e}，从头开始")
            return 0

    # ========================================================================
    #  JSON 持久化
    # ========================================================================

    @staticmethod
    def _record_to_dict(r: ValidationRecord) -> Dict:
        return {
            "question_id": r.question_id,
            "template_id": r.template_id,
            "coverage_level": r.coverage_level,
            "question_type": r.question_type,
            "question": r.question,
            "expected_answer": r.expected_answer,
            "vlm_answers": r.vlm_answers,
            "vlm_parsed_answers": r.vlm_parsed_answers,
            "retry_count": r.retry_count,
            "is_correct": r.is_correct,
            "is_garbage": r.is_garbage,
            "is_ambiguous": r.is_ambiguous,
            "error_type": r.error_type,
            "notes": r.notes,
            "validation_time": r.validation_time,
            "covered_elements": r.covered_elements,
        }

    @staticmethod
    def _dict_to_record(d: Dict) -> ValidationRecord:
        return ValidationRecord(
            question_id=d.get("question_id", ""),
            template_id=d.get("template_id", ""),
            coverage_level=d.get("coverage_level", ""),
            question_type=d.get("question_type", ""),
            question=d.get("question", ""),
            expected_answer=d.get("expected_answer", ""),
            vlm_answers=d.get("vlm_answers", []),
            vlm_parsed_answers=d.get("vlm_parsed_answers", []),
            retry_count=d.get("retry_count", 0),
            is_correct=d.get("is_correct", False),
            is_garbage=d.get("is_garbage", False),
            is_ambiguous=d.get("is_ambiguous", False),
            error_type=d.get("error_type", ""),
            notes=d.get("notes", ""),
            validation_time=d.get("validation_time", 0.0),
            covered_elements=d.get("covered_elements", []),
        )

    def save_results(self, path: str):
        """保存全部验证结果到 JSON"""
        data = {
            "stats": self._stats,
            "timestamp": datetime.now().isoformat(),
            "max_retries": self.max_retries,
            "records": [self._record_to_dict(r) for r in self.records],
        }
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"验证结果已保存: {out}")

    @classmethod
    def load_results(cls, path: str) -> 'QAValidator':
        """从 JSON 加载验证结果"""
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)

        validator = cls(max_retries=data.get("max_retries", 5))
        validator._stats = data.get("stats", {})
        validator.records = [cls._dict_to_record(r) for r in data.get("records", [])]
        return validator

    # ========================================================================
    #  报告
    # ========================================================================

    def print_report(self):
        """打印验证报告"""
        total = self._stats.get("total", 0)
        correct = self._stats.get("correct", 0)
        garbage = self._stats.get("garbage", 0)
        ambiguous = self._stats.get("ambiguous", 0)
        retried = self._stats.get("retried", 0)

        print("=" * 60)
        print("  QA Validation Report")
        print("=" * 60)
        print(f"  Total:     {total}")
        print(f"  Correct:   {correct}  ({correct/max(total,1)*100:.1f}%)")
        print(f"  Garbage:   {garbage}  ({garbage/max(total,1)*100:.1f}%)")
        print(f"  Ambiguous: {ambiguous}")
        print(f"  Retried:   {retried}")

        # 按级别统计
        print(f"\n  Per-level accuracy:")
        for level in ["L0", "L1", "L2"]:
            lr = [r for r in self.records if r.coverage_level == level]
            lok = sum(1 for r in lr if r.is_correct)
            lt = len(lr)
            if lt > 0:
                print(f"    {level}: {lok}/{lt} ({lok/lt*100:.1f}%)")

        # 按问题类型
        print(f"\n  Per-type accuracy:")
        for qtype in ["exist", "count", "status", "object", "comparison"]:
            tr = [r for r in self.records if r.question_type == qtype]
            tok = sum(1 for r in tr if r.is_correct)
            tt = len(tr)
            if tt > 0:
                print(f"    {qtype}: {tok}/{tt} ({tok/tt*100:.1f}%)")

        if garbage > 0:
            error_types = Counter(r.error_type for r in self.records if r.is_garbage)
            print(f"\n  Error types:")
            for etype, cnt in error_types.most_common():
                print(f"    {etype}: {cnt}")

            template_errors = Counter(r.template_id for r in self.records if r.is_garbage)
            print(f"\n  Top garbage templates:")
            for tid, cnt in template_errors.most_common(5):
                print(f"    {tid}: {cnt}")

    def garbage_template_ids(self) -> List[str]:
        """返回垃圾率超过 50% 的模板 ID 列表，可用于屏蔽"""
        tmpl_data = defaultdict(lambda: {"total": 0, "garbage": 0})
        for r in self.records:
            tmpl_data[r.template_id]["total"] += 1
            if r.is_garbage:
                tmpl_data[r.template_id]["garbage"] += 1
        return [tid for tid, d in tmpl_data.items()
                if d["total"] >= 3 and d["garbage"] / d["total"] > 0.5]


# ============================================================================
#  演示 / 自测
# ============================================================================

if __name__ == "__main__":
    import sys
    sys.path.insert(0, r"E:\Project\ADVTEST\nuscenes_s3c_experiment")
    logging.basicConfig(level=logging.INFO, format="%(name)s: %(message)s")

    # 模拟 QA 列表 — 覆盖各种边界情况
    test_qa_list = [
        # 正常题 (VLM 答对)
        {"question_id": "q001", "template_id": "L0_exist_A1", "coverage_level": "L0",
         "question_type": "exist", "answer_type": "bool",
         "question": "Are there any cars?", "answer": "yes"},
        # VLM 总是答错 → max_retries_exceeded
        {"question_id": "q002", "template_id": "L0_count_A1", "coverage_level": "L0",
         "question_type": "count", "answer_type": "number",
         "question": "How many cars are there?", "answer": "5"},
        # 未填充占位符 → sanity_fail
        {"question_id": "q003", "template_id": "L0_exist_A1", "coverage_level": "L0",
         "question_type": "exist", "answer_type": "bool",
         "question": "Are there any {obj_type}?", "answer": "yes"},
        # 负数答案 → sanity_fail
        {"question_id": "q004", "template_id": "L0_count_A1", "coverage_level": "L0",
         "question_type": "count", "answer_type": "number",
         "question": "How many trucks?", "answer": "-3"},
        # VLM 前两次答错第三次答对 → retried
        {"question_id": "q005", "template_id": "L1_exist_A1", "coverage_level": "L1",
         "question_type": "exist", "answer_type": "bool",
         "question": "Are there any cars to the front?", "answer": "yes"},
        # 重复题 → duplicate
        {"question_id": "q006", "template_id": "L0_exist_A1", "coverage_level": "L0",
         "question_type": "exist", "answer_type": "bool",
         "question": "Are there any cars?", "answer": "yes"},
        # VLM 每次答案都不同 → ambiguous
        {"question_id": "q007", "template_id": "L1_status_A1", "coverage_level": "L1",
         "question_type": "status", "answer_type": "word",
         "question": "What is the status of the car to the left?", "answer": "moving"},
        # VLM 返回带 <think> 标签的回答 → 解析后匹配
        {"question_id": "q008", "template_id": "L0_exist_B1", "coverage_level": "L0",
         "question_type": "exist", "answer_type": "bool",
         "question": "Do you see any pedestrians?", "answer": "yes"},
        # 正常 L2 题
        {"question_id": "q009", "template_id": "L2_count_A1", "coverage_level": "L2",
         "question_type": "count", "answer_type": "number",
         "question": "How many cars have the same status as pedestrian1?", "answer": "3"},
        # 问题太短 → sanity_fail
        {"question_id": "q010", "template_id": "L0_exist_A1", "coverage_level": "L0",
         "question_type": "exist", "answer_type": "bool",
         "question": "Car?", "answer": "yes"},
    ]

    # 模拟 VLM
    call_count = {}
    def mock_vlm(question: str) -> str:
        call_count[question] = call_count.get(question, 0) + 1
        cnt = call_count[question]

        # q002: 总是答错
        if "How many cars are there" in question:
            return "I think there are 3 cars."

        # q005: 前两次答错第三次答对
        if "cars to the front" in question:
            if cnt <= 2:
                return "<think>Let me think...</think>I'm not sure, maybe."
            return "<think>Checking...</think>The answer is yes."

        # q007: 每次不同 (ambiguous)
        if "status of the car to the left" in question:
            return ["stopped", "moving", "parked", "stationary", "idle"][cnt - 1]

        # q008: 带 <think> 标签
        if "pedestrians" in question:
            return "<think>I see some people walking.</think>Yes, there are pedestrians."

        # q009: 答对
        if "same status as pedestrian1" in question:
            return "The answer is 3."

        return "yes"

    print("=" * 60)
    print("  QAValidator 完善版 自测")
    print("=" * 60)

    validator = QAValidator(
        max_retries=5,
        checkpoint_dir="output/validation_checkpoint",
        auto_save_every=5,
    )
    results = validator.validate_batch(test_qa_list, verify_fn=mock_vlm)

    validator.print_report()

    # 保存
    validator.save_results("output/validation_demo_v2.json")
    validator.export_garbage_excel("output/garbage_questions_demo_v2.xlsx")

    # 显示垃圾模板黑名单
    blacklist = validator.garbage_template_ids()
    if blacklist:
        print(f"\nGarbage template blacklist: {blacklist}")

    print(f"\nAll records:")
    for r in results:
        status = "OK" if r.is_correct else ("AMB" if r.is_ambiguous else "BAD")
        print(f"  {r.question_id} [{status}] {r.error_type or '-':20s} "
              f"parsed={r.vlm_parsed_answers[:3] if r.vlm_parsed_answers else '-'}")

    # 测试 JSON 往返
    validator.save_results("output/validation_roundtrip.json")
    loaded = QAValidator.load_results("output/validation_roundtrip.json")
    assert len(loaded.records) == len(results), "Round-trip mismatch!"
    print(f"\nJSON round-trip OK ({len(loaded.records)} records)")
