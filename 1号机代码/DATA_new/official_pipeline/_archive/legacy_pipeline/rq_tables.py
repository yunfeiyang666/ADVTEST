"""
rq_tables.py — V25 CSV/Excel 写入器（物理记录层）

核心纪律：
  1. 动态列名检测：启动时读取 RQ.xlsx Row1，严禁硬编码覆盖用户新增列。
  2. 严格分区路由（铁律）：
       raw_coverage          → 仅 NuScenes-QA baseline 原题及其 L0/L1/L2 足迹
       question-answer-our   → 仅我们系统生成的新题（16列完整指标）
       model_performance_raw → 仅 MUT 视觉模型评测结果
  3. L2 路径持久化：[{"o1":"ego","o2":"car1","o3":"car2"}] JSON 字符串格式完整写入。
  4. 写入原则：每条完成立即 load→append→save，不攒批次。
  5. V24：filter_record 一帧一行 upsert；时间戳一律毫秒级物理 now()；ID 协议见各写入函数。
  6. V25：支持 CSV 模式（环境变量 VQA_USE_CSV=true），避免 Excel 并发损坏问题。

Excel 路径：环境变量 ADVTEST_EXCEL_PATH（见 advtest_paths.EXCEL_PATH）
CSV 模式：环境变量 VQA_USE_CSV=true（推荐，稳定快速）
"""
from __future__ import annotations

import json, logging, re, uuid, time, os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# fcntl 仅在 Unix 系统上可用
try:
    import fcntl
except ImportError:
    fcntl = None  # Windows 系统

# V25: 检查是否使用 CSV 模式
USE_CSV = os.getenv("VQA_USE_CSV", "false").lower() in ("true", "1", "yes")


def _ms_now() -> str:
    """Millisecond-precision timestamp: YYYY-MM-DD HH:MM:SS.mmm"""
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]


def _parse_ms_ts(ts: str):
    try:
        return datetime.strptime(ts, '%Y-%m-%d %H:%M:%S.%f')
    except Exception:
        return None

import openpyxl

from advtest_paths import EXCEL_PATH

# V25: CSV 模式支持
if USE_CSV:
    try:
        import csv_writer
        logger = logging.getLogger(__name__)
        logger.info("Using CSV mode (VQA_USE_CSV=true)")
    except ImportError:
        logger = logging.getLogger(__name__)
        logger.warning("csv_writer not found, falling back to Excel mode")
        USE_CSV = False
else:
    logger = logging.getLogger(__name__)

# Sheet 名称
SHEET_A  = "raw_coverage"
SHEET_B  = "question-answer-our"
SHEET_C  = "model_performance_raw_our"
SHEET_FR = "filter_record"   # V7.8 新增：过滤过程记录


# ─────────────────────────────────────────────────────────────────────────────
# 文件锁机制（防止并发写入损坏 Excel）
# ─────────────────────────────────────────────────────────────────────────────

import threading
import os

_excel_lock = threading.Lock()  # 进程内锁
_lock_file_path = None


def _get_lock_file_path():
    """获取锁文件路径"""
    global _lock_file_path
    if _lock_file_path is None:
        _lock_file_path = str(EXCEL_PATH) + ".lock"
    return _lock_file_path


class ExcelFileLock:
    """Excel 文件锁（支持跨进程）"""
    def __init__(self, timeout=30):
        self.timeout = timeout
        self.lock_file = None
        self.acquired = False

    def __enter__(self):
        # 先获取进程内锁
        _excel_lock.acquire()

        # 再获取跨进程文件锁
        lock_path = _get_lock_file_path()
        start_time = time.time()

        while True:
            try:
                # 尝试创建锁文件
                self.lock_file = open(lock_path, 'w')

                # 尝试获取排他锁（非阻塞）
                if os.name == 'nt':
                    # Windows: 使用 msvcrt
                    import msvcrt
                    msvcrt.locking(self.lock_file.fileno(), msvcrt.LK_NBLCK, 1)
                else:
                    # Linux/Unix: 使用 fcntl
                    if fcntl is None:
                        raise ImportError("fcntl not available on this platform")
                    fcntl.flock(self.lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)

                self.acquired = True
                return self

            except (IOError, OSError):
                # 锁被占用，等待重试
                if time.time() - start_time > self.timeout:
                    _excel_lock.release()
                    raise TimeoutError(f"Failed to acquire Excel lock after {self.timeout}s")

                if self.lock_file:
                    self.lock_file.close()
                    self.lock_file = None

                time.sleep(0.1)  # 等待 100ms 后重试

    def __exit__(self, exc_type, exc_val, exc_tb):
        # 释放文件锁
        if self.lock_file:
            try:
                if os.name == 'nt':
                    import msvcrt
                    msvcrt.locking(self.lock_file.fileno(), msvcrt.LK_UNLCK, 1)
                else:
                    if fcntl is not None:
                        fcntl.flock(self.lock_file.fileno(), fcntl.LOCK_UN)
            except:
                pass

            try:
                self.lock_file.close()
            except:
                pass

            self.lock_file = None

        # 释放进程内锁
        _excel_lock.release()

        return False


# ─────────────────────────────────────────────────────────────────────────────
# 动态列名检测（每次写入前从 Excel 读取真实 Row1，绝不硬编码）
# ─────────────────────────────────────────────────────────────────────────────

def _get_sheet_columns(sheet_name: str) -> List[str]:
    """Read Row1 of sheet_name from Excel and return column list (non-empty only)."""
    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
        ws = wb[sheet_name]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        return [str(v) for v in row1 if v is not None]
    except Exception as exc:
        logger.error("Cannot read columns from sheet '%s': %s", sheet_name, exc)
        return []


def _print_schema_summary():
    """Print current Excel schema for all sheets (debugging aid)."""
    print(f"Excel: {EXCEL_PATH.resolve()}")
    for sh in [SHEET_A, SHEET_B, SHEET_C]:
        cols = _get_sheet_columns(sh)
        print(f"  [{sh}] ({len(cols)} cols): {cols}")


def make_generated_question_id(scene_id: str, frame_id: int, suffix_len: int = 8) -> str:
    """
    Sheet B 专用：gen_{scene}_f{frame}_{hex}
    例 scene-0926, 20 → gen_s0926_f20_a1b2c3d4
    """
    sid = str(scene_id or "").strip().replace("scene-", "s").replace("-", "")
    if not sid.startswith("s"):
        sid = "s" + sid
    suf = uuid.uuid4().hex[: max(4, min(suffix_len, 32))]
    return f"gen_{sid}_f{frame_id}_{suf}"


def _normalize_col_name(name: str) -> str:
    """Normalize a column name to improve matching across localized/annotated headers."""
    s = str(name or "").strip().lower()
    if not s:
        return ""
    s = (
        s.replace("（", "(").replace("）", ")")
         .replace("【", "[").replace("】", "]")
         .replace("｛", "{").replace("｝", "}")
    )
    # Remove annotation blocks, e.g. "timestamp_start（开始时间）" -> "timestamp_start"
    prev = None
    while prev != s:
        prev = s
        s = re.sub(r"\([^)]*\)", "", s)
        s = re.sub(r"\[[^\]]*\]", "", s)
        s = re.sub(r"\{[^}]*\}", "", s)
    # Ignore separators/punctuation to allow robust canonical-key matching.
    s = re.sub(r"[\s_\-:：/\\|·、，,。；;()\[\]{}]+", "", s)
    return s


# ─────────────────────────────────────────────────────────────────────────────
# 核心原子写入（动态对齐版）
# ─────────────────────────────────────────────────────────────────────────────

def _append_row(sheet_name: str, data: Dict[str, Any]) -> bool:
    """
    load → detect columns → map data → append → save.

    Column order is taken from Excel Row1 dynamically.
    Any key in data that doesn't match a column is silently ignored.
    Any column missing from data gets an empty string.
    list/dict values are serialized to JSON strings.
    """
    try:
        with ExcelFileLock(timeout=30):
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            ws = wb[sheet_name]

            # Detect columns from Row1
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            cols = [str(v) for v in row1 if v is not None]

            # Build normalized lookup once: allows canonical keys to match annotated headers.
            norm_data: Dict[str, Any] = {}
            for k, v in data.items():
                nk = _normalize_col_name(k)
                if nk and nk not in norm_data:
                    norm_data[nk] = v

            # Build row aligned to column order
            row = []
            for col in cols:
                if col in data:
                    val = data.get(col, "")
                else:
                    val = norm_data.get(_normalize_col_name(col), "")
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                elif val is None:
                    val = ""
                row.append(val)

            ws.append(row)
            wb.save(str(EXCEL_PATH))
            wb.close()
            return True

    except KeyError:
        logger.error("Sheet '%s' not found in %s", sheet_name, EXCEL_PATH)
    except Exception as exc:
        logger.error("Excel write [%s] failed: %s", sheet_name, exc)
    return False


def _norm_header_cell(v: Any) -> str:
    return _normalize_col_name(str(v or ""))


# filter_record 表头别名 → 规范键（用于一帧一行 upsert）
_FR_CANON = {
    "sceneid": "scene_id",
    "frameid": "frame_id",
    "originalnodescount": "original_nodes_count",
    "filterednodescount": "filtered_nodes_count",
    "originalnum": "original_nodes_count",
    "filterednum": "filtered_nodes_count",
    "beforenodes": "original_nodes_count",
    "afternodes": "filtered_nodes_count",
    "rawnodes": "original_nodes_count",
    "筛选前节点数": "original_nodes_count",
    "筛选后节点数": "filtered_nodes_count",
    "过滤前节点数": "original_nodes_count",
    "过滤后节点数": "filtered_nodes_count",
    "原始总节点": "original_nodes_count",
    "筛选后节点": "filtered_nodes_count",
}


def _fr_header_to_canonical(headers: List[str]) -> Tuple[List[str], Dict[int, str]]:
    """Row1 → 每列索引对应的规范字段名（仅 filter_record 关心的列）。"""
    col2canon: Dict[int, str] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        nk = _norm_header_cell(h)
        canon = _FR_CANON.get(nk, nk)
        if canon in ("scene_id", "frame_id", "original_nodes_count", "filtered_nodes_count", "ratio"):
            col2canon[i] = canon
    return headers, col2canon


def _upsert_filter_record_row(
    scene_id: str,
    frame_id: int,
    original_num: int,
    filtered_num: int,
    ratio: float,
) -> bool:
    """
    filter_record：每个 (scene_id, frame_id) 仅一行；已存在则整行覆盖，否则追加。
    仅写入 5 个元数据字段；表头中其它列留空（不写入 filtered_vex / 时间戳）。
    """
    values: Dict[str, Any] = {
        "scene_id": str(scene_id).strip(),
        "frame_id": int(frame_id),
        "original_nodes_count": int(original_num),
        "filtered_nodes_count": int(filtered_num),
        "ratio": round(float(ratio), 4),
    }
    try:
        with ExcelFileLock(timeout=30):
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            if SHEET_FR not in wb.sheetnames:
                ws = wb.create_sheet(SHEET_FR)
                hdr = [
                    "scene_id", "frame_id",
                    "original_nodes_count", "filtered_nodes_count", "ratio",
                ]
                ws.append(hdr)
            else:
                ws = wb[SHEET_FR]
                if ws.max_row < 1 or all(c.value is None for c in ws[1]):
                    hdr = [
                        "scene_id", "frame_id",
                        "original_nodes_count", "filtered_nodes_count", "ratio",
                    ]
                    for c, name in enumerate(hdr, 1):
                        ws.cell(row=1, column=c, value=name)

            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(v) if v is not None else "" for v in row1]
            # 补齐 openpyxl 中尾部空列未出现在 values_only 的情况
            while len(headers) < (ws.max_column or 0):
                headers.append("")
            _, col2canon = _fr_header_to_canonical(headers)

            if "scene_id" not in col2canon.values() or "frame_id" not in col2canon.values():
                logger.error("filter_record sheet missing scene_id/frame_id columns")
                wb.close()
                return False

            inv_scene = {idx for idx, k in col2canon.items() if k == "scene_id"}
            inv_frame = {idx for idx, k in col2canon.items() if k == "frame_id"}
            si = next(iter(inv_scene))
            fi = next(iter(inv_frame))

            match_row: Optional[int] = None
            for r in range(2, ws.max_row + 1):
                sv = ws.cell(row=r, column=si + 1).value
                fv = ws.cell(row=r, column=fi + 1).value
                if sv is None and fv is None:
                    continue
                try:
                    fvi = int(fv)
                except Exception:
                    continue
                if str(sv).strip() == values["scene_id"] and fvi == values["frame_id"]:
                    match_row = r
                    break

            target_row = match_row if match_row is not None else ws.max_row + 1

            ncols = max(len(headers), ws.max_column or 0, 5)
            for ci in range(ncols):
                ws.cell(row=target_row, column=ci + 1, value=None)

            for ci, h in enumerate(headers):
                canon = col2canon.get(ci)
                if not canon or canon not in values:
                    continue
                ws.cell(row=target_row, column=ci + 1, value=values[canon])

            wb.save(str(EXCEL_PATH))
            wb.close()
            return True
    except Exception as exc:
        logger.error("filter_record upsert failed: %s", exc)
    return False


# ─────────────────────────────────────────────────────────────────────────────
# ┌─────────────────────────────────────────────────────────────────────────┐
# │  SHEET A: raw_coverage — 仅记录 NuScenes-QA Baseline 原题的足迹         │
# │  禁止将我们生成的题写入此 Sheet                                           │
# └─────────────────────────────────────────────────────────────────────────┘
# ─────────────────────────────────────────────────────────────────────────────

def write_baseline_to_coverage(
    scene_id:        str,
    frame_id:        int,
    nuscenes_qa_id:  str,          # 原题在 NuScenes-QA 中的编号
    question:        str,
    answer:          str,
    l0_nodes:        List[str],    # ["ego","car1","truck1"]
    l1_edges:        List[Dict],   # [{"source":"ego","target":"car1"},...]
    l2_paths:        List[Dict],   # [{"o1":"ego","o2":"car1","o3":"car2"}] 或 []
    # 用户新增列（V7.6 动态检测后补充）
    question_type:   str = "",     # NuScenes-QA template_type
    timestamp_start: str = "",     # 审计开始时刻
    timestamp_end:   str = "",     # 审计结束时刻
    audit_cypher:    str = "",     # LLM 生成的审计 Cypher
    global_val_index: Optional[int] = None,  # V24: val 文件全局下标，用于修复丢失的 nuscenes_qa_id
) -> bool:
    """
    将 NuScenes-QA 原题及其语义审计足迹写入 raw_coverage。
    包含用户新增的 4 列：question_type / timestamp_start / timestamp_end /
    question_cypher（llm生成的cypher）。
    """
    # V25: CSV 模式路由
    if USE_CSV:
        data = {
            "qa_unique_id": nuscenes_qa_id,
            "scene_id": scene_id,
            "frame_id": frame_id,
            "question": question,
            "answer": answer,
            "q_type": question_type,
            "num_hop": 0,  # baseline 没有 num_hop
            "l0_nodes": l0_nodes,
            "l1_edges": l1_edges,
            "l2_paths": l2_paths,
            "n_l0": len(l0_nodes),
            "n_l1": len(l1_edges),
            "n_l2": len(l2_paths),
            "llm_ms": 0,  # baseline 没有 llm_ms
            "success": True,
            "timestamp": timestamp_end or _ms_now(),
        }
        return csv_writer.write_baseline_row(data)

    # Excel 模式（原有逻辑）
    from semantic_auditor import make_qa_id as _val_id

    nid = str(nuscenes_qa_id or "").strip()
    if not nid.startswith("val_"):
        if global_val_index is not None:
            nid = _val_id(int(global_val_index), question_type or "qa")
        else:
            nid = _val_id(
                abs(hash((scene_id, frame_id, question))) % 10_000_000,
                question_type or "qa",
            )
            logger.warning(
                "nuscenes_qa_id missing val_ prefix; regenerated as %s (provide global_val_index to avoid hash)",
                nid,
            )
    ts0 = (timestamp_start or "").strip() or _ms_now()
    ts1 = (timestamp_end or "").strip() or ts0
    return _append_row(SHEET_A, {
        "scene_id":       scene_id,
        "frame_id":       frame_id,
        "nuscenes_qa_id": nid,
        "question":       question,
        "answer":         answer,
        "L0":             l0_nodes,
        "L1":             l1_edges,
        "L2":             l2_paths,
        # 用户新增列（动态对齐，_append_row 不会因独包含而出错）
        "question_type":  question_type,
        "timestamp_start": ts0,
        "timestamp_end":  ts1,
        "question_cypher（llm生成的cypher）": audit_cypher,
        # 保险字段
        "source":         "nuscenes_qa_baseline",
    })


# ─────────────────────────────────────────────────────────────────────────────
# ┌─────────────────────────────────────────────────────────────────────────┐
# │  SHEET B: question-answer-our — 仅记录我们系统生成的新题                │
# │  16 列完整指标；禁止将原题写入此 Sheet                                   │
# └─────────────────────────────────────────────────────────────────────────┘
# ─────────────────────────────────────────────────────────────────────────────

def write_generated_question(
    scene_id:              str,
    frame_id:              int,
    question_id:           str,       # 建议 gen_*（见 make_generated_question_id）
    timestamp_start:       str,       # ISO: "2026-04-03T22:30:01"
    timestamp_llm:         str,       # LLM 返回上下文时刻
    timestamp_cypher_return: str,     # Neo4j 初始查询完毕时刻
    timestamp_end:         str,       # 唯一化结束时刻
    iteration_count:       int,       # ConstraintChain 轮数
    question_type:         str,       # exist/status/object/count/comparison
    complexity:            str,       # L2 (topology level)
    natural_language_question: str,   # 自然语言问题
    cypher_question:       str,       # LLM 生成的上下文 Cypher（完整）
    answer:                str,
    l0_nodes:              List[str],
    l1_edges:              List[Dict],
    l2_paths:              List[Dict],
    target_gap_cell:       str = "",  # V11: 被填补的缺口路径 (path_pattern)
    batch_id:              str = "",  # V19: 并发批次标识
) -> bool:
    """
    将我们系统生成的新题写入 question-answer-our。
    16 列全部填入，L2 路径以 JSON 字符串完整写入。
    target_gap_cell 写入同名列（如 Excel 中存在该列）。
    """
    # V25: CSV 模式路由
    if USE_CSV:
        # 计算时间
        _ts_list = [timestamp_start, timestamp_llm, timestamp_cypher_return, timestamp_end]
        _dt = [_parse_ms_ts(str(x).strip()) for x in _ts_list if str(x).strip()]
        if len(_dt) >= 2:
            total_ms = int((_dt[-1] - _dt[0]).total_seconds() * 1000)
        else:
            total_ms = 0

        data = {
            "qa_unique_id": question_id,
            "scene_id": scene_id,
            "frame_id": frame_id,
            "question": natural_language_question,
            "answer": answer,
            "q_type": question_type,
            "num_hop": 2 if complexity in ("L2A", "L2B") else 1,
            "timestamp_start": timestamp_start,
            "timestamp_llm": timestamp_llm,
            "timestamp_cypher_return": timestamp_cypher_return,
            "timestamp_end": timestamp_end,
            "iteration_count": iteration_count,
            "complexity": complexity,
            "cypher_question": cypher_question,
            "l0_nodes": l0_nodes,
            "l1_edges": l1_edges,
            "l2_paths": l2_paths,
            "n_l0": len(l0_nodes),
            "n_l1": len(l1_edges),
            "n_l2": len(l2_paths),
            "llm_ms": total_ms,
            "gap_cell": target_gap_cell,
            "batch_id": batch_id,
            "success": True,
        }
        return csv_writer.write_generated_row(data)

    # Excel 模式（原有逻辑）
    # V19: 多段时间戳硬约束，防止不真实/错序记录进入 Excel
    _ts_list = [timestamp_start, timestamp_llm, timestamp_cypher_return, timestamp_end]
    if any(not str(x).strip() for x in _ts_list):
        logger.error("Reject write_generated_question: empty timestamp segment(s) for question_id=%s", question_id)
        return False
    _dt = [_parse_ms_ts(str(x).strip()) for x in _ts_list]
    if any(x is None for x in _dt):
        logger.error("Reject write_generated_question: invalid timestamp format for question_id=%s", question_id)
        return False
    if not (_dt[0] <= _dt[1] <= _dt[2] <= _dt[3]):
        logger.error(
            "Reject write_generated_question: timestamp order invalid for question_id=%s "
            "(start=%s llm=%s cypher=%s end=%s)",
            question_id, timestamp_start, timestamp_llm, timestamp_cypher_return, timestamp_end
        )
        return False
    qid = str(question_id or "").strip()
    if not qid.startswith("gen_"):
        qid = make_generated_question_id(scene_id, frame_id)
        logger.warning("question_id missing gen_ prefix; auto-assigned %s", qid)
    return _append_row(SHEET_B, {
        "scene_id":                  scene_id,
        "frame_id":                  frame_id,
        "question_id":               qid,
        "timestamp_start":           timestamp_start,
        "timestamp_llm":             timestamp_llm,
        "timestamp_cypher_return":   timestamp_cypher_return,
        "timestamp_end":             timestamp_end,
        "iteration_count":           iteration_count,
        "question_type":             question_type,
        "complexity":                complexity,
        "natural language question": natural_language_question,
        "cypher question":           cypher_question,
        "answer":                    answer,
        "L0":                        l0_nodes,
        "L1":                        l1_edges,
        "L2":                        l2_paths,
        "gap_cell":                  target_gap_cell,  # Excel col: gap_cell
        "target_gap_cell":           target_gap_cell,  # 兼容部分表头命名
        "batch_id":                  batch_id,
        "Batch_ID":                  batch_id,
    })


def write_generated_questions_batch(rows: List[Dict[str, Any]]) -> List[bool]:
    """
    Batch append to question-answer-our with one workbook load/save.
    Input rows should use the same keys as write_generated_question payload mapping.
    Returns per-row success flags aligned to input order.
    """
    if not rows:
        return []

    # V25: CSV 模式路由 - 批量写入
    if USE_CSV:
        status = []
        for row in rows:
            # 提取时间戳并计算 total_ms
            _ts_list = [
                row.get("timestamp_start", ""),
                row.get("timestamp_llm", ""),
                row.get("timestamp_cypher_return", ""),
                row.get("timestamp_end", ""),
            ]
            _dt = [_parse_ms_ts(str(x).strip()) for x in _ts_list if str(x).strip()]
            if len(_dt) >= 2:
                total_ms = int((_dt[-1] - _dt[0]).total_seconds() * 1000)
            else:
                total_ms = 0

            data = {
                "qa_unique_id": row.get("question_id", ""),
                "scene_id": row.get("scene_id", ""),
                "frame_id": row.get("frame_id", 0),
                "question": row.get("natural language question", ""),
                "answer": row.get("answer", ""),
                "q_type": row.get("question_type", ""),
                "num_hop": 2 if row.get("complexity", "") in ("L2A", "L2B") else 1,
                "timestamp_start": row.get("timestamp_start", ""),
                "timestamp_llm": row.get("timestamp_llm", ""),
                "timestamp_cypher_return": row.get("timestamp_cypher_return", ""),
                "timestamp_end": row.get("timestamp_end", ""),
                "iteration_count": row.get("iteration_count", 0),
                "complexity": row.get("complexity", ""),
                "cypher_question": row.get("cypher question", ""),
                "l0_nodes": row.get("L0", []),
                "l1_edges": row.get("L1", []),
                "l2_paths": row.get("L2", []),
                "n_l0": len(row.get("L0", [])),
                "n_l1": len(row.get("L1", [])),
                "n_l2": len(row.get("L2", [])),
                "llm_ms": total_ms,
                "gap_cell": row.get("gap_cell", "") or row.get("target_gap_cell", ""),
                "batch_id": row.get("batch_id", "") or row.get("Batch_ID", ""),
                "success": True,
            }
            success = csv_writer.write_generated_row(data)
            status.append(success)
        return status

    # Excel 模式（原有逻辑）
    status = [False] * len(rows)
    wb = None
    try:
        with ExcelFileLock(timeout=30):
            wb = openpyxl.load_workbook(str(EXCEL_PATH))
            ws = wb[SHEET_B]
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            cols = [str(v) for v in row1 if v is not None]

            def _align(data: Dict[str, Any]) -> List[Any]:
                norm_data: Dict[str, Any] = {}
                for k, v in data.items():
                    nk = _normalize_col_name(k)
                    if nk and nk not in norm_data:
                        norm_data[nk] = v
                out: List[Any] = []
                for col in cols:
                    if col in data:
                        val = data.get(col, "")
                    else:
                        val = norm_data.get(_normalize_col_name(col), "")
                    if isinstance(val, (list, dict)):
                        val = json.dumps(val, ensure_ascii=False)
                    elif val is None:
                        val = ""
                    out.append(val)
                return out

            for i, raw in enumerate(rows):
                try:
                    data = dict(raw or {})
                    _ts = [
                        str(data.get("timestamp_start", "") or "").strip(),
                        str(data.get("timestamp_llm", "") or "").strip(),
                        str(data.get("timestamp_cypher_return", "") or "").strip(),
                        str(data.get("timestamp_end", "") or "").strip(),
                    ]
                    if any(not x for x in _ts):
                        logger.error(
                            "Skip batch row (missing timestamp): idx=%d qid=%s",
                            i,
                            data.get("question_id", ""),
                        )
                        continue
                    _dt = [_parse_ms_ts(x) for x in _ts]
                    if any(x is None for x in _dt):
                        logger.error(
                            "Skip batch row (invalid timestamp format): idx=%d qid=%s",
                            i,
                            data.get("question_id", ""),
                        )
                        continue
                    if not (_dt[0] <= _dt[1] <= _dt[2] <= _dt[3]):
                        logger.error(
                            "Skip batch row (timestamp order invalid): idx=%d qid=%s",
                            i,
                            data.get("question_id", ""),
                        )
                        continue
                    qid = str(data.get("question_id", "") or "").strip()
                    if not qid.startswith("gen_"):
                        qid = make_generated_question_id(
                            str(data.get("scene_id", "") or ""),
                            int(data.get("frame_id", 0) or 0),
                        )
                        data["question_id"] = qid
                    ws.append(_align(data))
                    status[i] = True
                except Exception as exc:
                    logger.error("Skip batch row idx=%d due to exception: %s", i, exc)

            if any(status):
                wb.save(str(EXCEL_PATH))
            wb.close()
            return status
    except KeyError:
        logger.error("Sheet '%s' not found in %s", SHEET_B, EXCEL_PATH)
    except Exception as exc:
        logger.error("Excel batch write [%s] failed: %s", SHEET_B, exc)
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
    return status


# ─────────────────────────────────────────────────────────────────────────────
# ┌─────────────────────────────────────────────────────────────────────────┐
# │  SHEET C: model_performance_raw_our — 仅记录 MUT 评测结果               │
# └─────────────────────────────────────────────────────────────────────────┘
# ─────────────────────────────────────────────────────────────────────────────

def write_mut_result(
    scene_id:            str,
    frame_id:            int,
    question_id:         str,
    model_name:          str,
    model_answer:        str,
    correct_answer:      str,
    question_type:       str,
    question_complexity: str,
    pass_fail:           str,    # "pass" 或 "fail"
    timestamp_start:     str,
    timestamp_end:       str,
) -> bool:
    """MUT 每道题评测完成后调用。"""
    ts0 = (timestamp_start or "").strip() or _ms_now()
    ts1 = (timestamp_end or "").strip() or _ms_now()
    return _append_row(SHEET_C, {
        "scene_id":            scene_id,
        "frame_id":            frame_id,
        "question_id":         question_id,
        "model-name":          model_name,
        "model-answer":        model_answer,
        "correct_answer":      correct_answer,
        "question_type":       question_type,
        "question_complexity": question_complexity,
        "pass":                pass_fail,
        "timestamp_start":     ts0,
        "timestamp_end":       ts1,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline 集成：从 qa_dict + timing_dict 提取后调用上面正确 Sheet
# ─────────────────────────────────────────────────────────────────────────────

def record_generated_qa(
    qa:       Dict[str, Any],
    timing:   Dict[str, Any],
    cypher:   str,
    scene_id: str,
    frame_id: int,
) -> None:
    """
    V6/V7 管线每条生成题完成后调用。
    只写入 SHEET_B（question-answer-our），不写 SHEET_A。
    """
    # 验证 Cypher 和问题的一致性
    from gap_pipeline.answer_logic_cypher import validate_cypher_matches_question
    question = qa.get("question", "")
    answer_logic = qa.get("answer_logic", "")

    is_valid, error_msg = validate_cypher_matches_question(question, cypher, answer_logic)
    if not is_valid:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"[QA Validation Failed] {error_msg}")
        logger.warning(f"  Question: {question}")
        logger.warning(f"  Cypher: {cypher}")
        # 不记录无效的问题
        return

    from gap_pipeline.level_taxonomy import get_meta as _meta

    topo   = qa.get("topology_level", "")
    method = (qa.get("Template_ID", "") or "").split(":")[-1]
    meta   = _meta(method, topology_level=topo)

    t0 = datetime.now()
    llm_ms  = float(timing.get("llm_ms",  0.0))
    neo_ms  = float(timing.get("neo4j_ms", 0.0))
    con_ms  = float(timing.get("constraint_ms", 0.0))
    ver_ms  = float(timing.get("verify_ms", 0.0))

    def _ts_off(ms: float) -> str:
        return (t0 + timedelta(milliseconds=ms)).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    trace = qa.get("Constraint_Trace") or timing.get("constraint_trace", "")
    iteration_count = int(timing.get("constraint_rounds") or 0)
    if iteration_count < 1:
        _parts = [p.strip() for p in str(trace).replace("->", "→").split("→")]
        iteration_count = max(1, len([p for p in _parts if p and p != "Path"]))

    fp_nodes = qa.get("footprint_nodes", []) or []
    path     = qa.get("path_pattern", "")
    l0, l1, l2 = _parse_footprint(fp_nodes, topo, path)

    write_generated_question(
        scene_id=scene_id, frame_id=frame_id,
        question_id=qa.get("question_id", "") or make_generated_question_id(scene_id, frame_id),
        timestamp_start=_ts_off(0),
        timestamp_llm=_ts_off(llm_ms),
        timestamp_cypher_return=_ts_off(llm_ms + neo_ms),
        timestamp_end=_ts_off(llm_ms + neo_ms + con_ms + ver_ms),
        iteration_count=iteration_count,
        question_type=topo or meta.q_type1,
        complexity=meta.difficulty,
        natural_language_question=qa.get("question", ""),
        cypher_question=cypher,
        answer=qa.get("answer", ""),
        l0_nodes=l0, l1_edges=l1, l2_paths=l2,
    )


def record_baseline_audit(
    nuscenes_qa_id: str,
    question:       str,
    answer:         str,
    audit_result:   Dict,
    scene_id:       str,
    frame_id:       int,
) -> None:
    """语义审计完成后，将原题足迹写入 raw_coverage（只写 SHEET_A）。"""
    write_baseline_to_coverage(
        scene_id=scene_id, frame_id=frame_id,
        nuscenes_qa_id=nuscenes_qa_id,
        question=question,
        answer=answer,
        l0_nodes=audit_result.get("l0_nodes", []),
        l1_edges=audit_result.get("l1_edges", []),
        l2_paths=audit_result.get("l2_paths", []),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Footprint parser
# ─────────────────────────────────────────────────────────────────────────────

def _parse_footprint(fp_nodes, topo, path):
    arr = "\u2192"  # →
    l0  = [n for n in fp_nodes if n]
    l1: List[Dict] = []
    l2: List[Dict] = []
    if topo in ("L2A", "L2B") and path.count(arr) == 2:
        parts = path.split(arr)
        if len(parts) == 3:
            n1, n2, n3 = parts
            l1 = [{"source": n1, "target": n2}, {"source": n2, "target": n3}]
            l2 = [{"o1": n1, "o2": n2, "o3": n3}]
    elif topo == "L1" and arr in path:
        parts = path.split(arr)
        if len(parts) == 2:
            l1 = [{"source": parts[0], "target": parts[1]}]
    return l0, l1, l2


# ─────────────────────────────────────────────────────────────────────────────
# ┌─────────────────────────────────────────────────────────────────────────┐
# │  SHEET FR: filter_record — 记录每次核心宇宙过滤的详细过程              │
# └─────────────────────────────────────────────────────────────────────────┘
# ─────────────────────────────────────────────────────────────────────────────

def write_filter_record(
    scene_id:      str,
    frame_id:      int,
    original_num:  int,     # 原始节点总数（含 barrier/trafficcone 等）
    filtered_num:  int,     # 核心宇宙过滤后保留的节点数
    filtered_vex:  str = "",  # 已弃用：filter_record 仅保留 5 列元数据，此参数忽略
    ratio:         float = 0.0,   # filtered_num / original_num；0 则自动计算
    timestamp_start: str = "",  # 已弃用：本 sheet 不再写时间戳
    timestamp_end:   str = "",
) -> bool:
    """
    将核心宇宙过滤统计写入 filter_record Sheet（V24：一帧一行，存在则覆盖）。
    列语义：scene_id, frame_id, original_nodes_count, filtered_nodes_count, ratio
    """
    _ = filtered_vex, timestamp_start, timestamp_end  # API 兼容旧调用
    r = float(ratio) if ratio else (filtered_num / max(original_num, 1))

    # V25: CSV 模式路由
    if USE_CSV:
        return csv_writer.write_filter_record(scene_id, frame_id, original_num, filtered_num, r)

    return _upsert_filter_record_row(scene_id, frame_id, original_num, filtered_num, r)


# ─────────────────────────────────────────────────────────────────────────────
# DIAGNOSTIC TEST（按分区路由测试三张表）
# ─────────────────────────────────────────────────────────────────────────────

def run_diagnostic_test():
    """
    向三张表各写入一行 DIAGNOSTIC_TEST 数据，然后读回验证。
    确认：列名不错位、L2 路径 JSON 完整写入、分区路由正确。
    """
    print("=" * 65)
    print("  DIAGNOSTIC_TEST 分区写入验证")
    print(f"  Excel: {EXCEL_PATH.resolve()}")
    print("=" * 65)

    # 先打印当前 schema
    print("\n[当前 Excel 列名]")
    _print_schema_summary()

    _base = datetime.now()

    def _off(ms: float) -> str:
        return (_base + timedelta(milliseconds=ms)).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

    ts0, ts1, ts2, ts4 = _off(0), _off(8), _off(22), _off(90)

    # ── Sheet A: raw_coverage (baseline only) ─────────────────────────────────
    print("\n[写入 raw_coverage] → baseline 原题足迹")
    ok_a = write_baseline_to_coverage(
        scene_id="DIAG_SCENE", frame_id=99,
        nuscenes_qa_id="val_99999999_diag",
        question="Is there a car to the front of me?",
        answer="yes",
        l0_nodes=["ego", "car1"],
        l1_edges=[{"source": "ego", "target": "car1", "dir": "front"}],
        l2_paths=[],
    )
    print(f"  ok={ok_a}")

    # ── Sheet B: question-answer-our (our generation only) ────────────────────
    print("\n[写入 question-answer-our] → 我们生成的题")
    ok_b = write_generated_question(
        scene_id="DIAG_SCENE", frame_id=99,
        question_id=make_generated_question_id("DIAG_SCENE", 99),
        timestamp_start=ts0,
        timestamp_llm=ts1,
        timestamp_cypher_return=ts2,
        timestamp_end=ts4,
        iteration_count=3,
        question_type="L2A", complexity="Hard",
        natural_language_question="What car is to the front of the truck in front of me?",
        cypher_question=(
            "MATCH (ego:Object {unique_id:'ego'})-[r1:RELATES_TO]->(a:Object)\n"
            "-[r2:RELATES_TO]->(b:Object)\n"
            "WHERE r1.direction_4='front' AND a.type='truck'\n"
            "RETURN collect(b.unique_id) AS l0_nodes, ..."
        ),
        answer="car1",
        l0_nodes=["ego", "truck1", "car1"],
        l1_edges=[{"source":"ego","target":"truck1"},{"source":"truck1","target":"car1"}],
        l2_paths=[{"o1":"ego","o2":"truck1","o3":"car1"}],
    )
    print(f"  ok={ok_b}")

    # ── Sheet C: model_performance_raw_our (MUT only) ─────────────────────────
    print("\n[写入 model_performance_raw_our] → MUT 评测结果")
    ok_c = write_mut_result(
        scene_id="DIAG_SCENE", frame_id=99,
        question_id="gen_sDIAGSCENE_f99_test0001",
        model_name="text_llm_qwen-plus",
        model_answer="car1",
        correct_answer="car1",
        question_type="L2A", question_complexity="Hard",
        pass_fail="pass",
        timestamp_start=ts0,
        timestamp_end=_off(120),
    )
    print(f"  ok={ok_c}")

    # ── Read-back verification ────────────────────────────────────────────────
    print("\n[读回验证]")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    for sh, label in [(SHEET_A,"raw_coverage"), (SHEET_B,"question-answer-our"), (SHEET_C,"model_perf")]:
        ws = wb[sh]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        cols = [str(v) for v in row1 if v is not None]
        data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
                     if any(v is not None for v in r)]
        if data_rows:
            last = data_rows[-1]
            print(f"\n  [{label}] ({len(cols)} cols, {len(data_rows)} data rows)")
            for c, v in zip(cols, last):
                if v is not None:
                    print(f"    '{c}' = {str(v)[:70]}")
    wb.close()

    _ok = ok_a and ok_b and ok_c
    print(f"\n[{'PASS' if _ok else 'FAIL'}] Diagnostic {'PASSED' if _ok else 'FAILED'}")


def scrub_excel_rows_containing(substring: str = "2023-01-01") -> int:
    """
    删除所有工作表中、任意单元格包含 substring 的数据行（自第 2 行起，保留表头）。
    请在关闭 Excel 后调用，否则 save 会失败。
    Returns: 删除行数。
    """
    if not substring:
        return 0
    removed = 0
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    for sh in wb.sheetnames:
        ws = wb[sh]
        mc = ws.max_column or 1
        for r in range(ws.max_row, 1, -1):
            parts: List[str] = []
            for c in range(1, mc + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    parts.append(str(v))
            if substring in "".join(parts):
                ws.delete_rows(r)
                removed += 1
    wb.save(str(EXCEL_PATH))
    wb.close()
    return removed


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--scrub-2023":
        n = scrub_excel_rows_containing("2023-01-01")
        print(f"scrub_excel_rows_containing: removed {n} rows")
    else:
        run_diagnostic_test()
