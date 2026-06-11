"""
analyze_coverage.py
===================
离线覆盖率分析器: 读 universe snapshot + Excel(raw_coverage / question-answer-our),
为每帧生成覆盖率增长曲线 (L0 / L1 / L2 三条线, baseline + our 拼接, X 轴双版本)。

输入:
  - universe snapshot dir : 每帧一个 JSON, 由 CoverageTracker.dump_universe_snapshot() 落盘
  - Excel                 : raw_coverage 表 (baseline) + question-answer-our 表 (our)

输出:
  - output/coverage_curves/
      ├── <scene>_frame<id>_curve_by_qcount.png    每帧两图
      ├── <scene>_frame<id>_curve_by_time.png
      ├── <scene>_frame<id>_curve_data.csv         数据点 (供 Excel/PPT 复用)
      └── _all_frames_summary.csv                  全部帧汇总
  - 写入 Excel RQ2_graph sheet                     供论文图表直接引用

用法:
  python analyze_coverage.py
  python analyze_coverage.py --excel "E:\\Project\\ADVTEST\\RQ(1).xlsx"
  python analyze_coverage.py --frame scene-0926 20      # 仅算单帧
  python analyze_coverage.py --no-excel-write           # 不写回 Excel
"""
from __future__ import annotations

import argparse
import collections
import json
import logging
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("analyze_coverage")

DEFAULT_EXCEL = Path(r"E:\Project\ADVTEST\RQ(1).xlsx")
DEFAULT_SNAPSHOT_DIR = Path(r"output/coverage_snapshots")
DEFAULT_OUT_DIR = Path(r"output/coverage_curves")

SHEET_RAW = "raw_coverage"
SHEET_OUR = "question-answer-our"
SHEET_RQ2 = "RQ2_graph"


# ─────────────────────────────────────────────────────────────────────────────
# Data containers
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Universe:
    """一帧的全集 (分母来源)."""
    scene_id: str
    frame_id: int
    L0_set:  Set[str]
    L1_set:  Set[str]
    L2A_set: Set[str]
    L2B_set: Set[str]

    @property
    def L2_set(self) -> Set[str]:
        return self.L2A_set | self.L2B_set

    @property
    def totals(self) -> Dict[str, int]:
        return {
            "L0":  len(self.L0_set),
            "L1":  len(self.L1_set),
            "L2A": len(self.L2A_set),
            "L2B": len(self.L2B_set),
            "L2":  len(self.L2_set),
        }


@dataclass
class QARow:
    """从 Excel 读出来的一行 QA, 已解析 L0/L1/L2 footprint."""
    source: str          # "baseline" | "our"
    scene_id: str
    frame_id: int
    qid: str             # nuscenes_qa_id 或 question_id
    timestamp: Optional[datetime]
    L0_keys: Set[str]    # node ids
    L1_keys: Set[str]    # "src|tgt"
    L2_keys: Set[str]    # "n1|n2|n3"  (统一格式, 不区分 A/B)


@dataclass
class CoverageCurvePoint:
    """曲线上的一个点."""
    x_qcount: int        # 累计题号 (1, 2, 3, ...)
    x_seconds: float     # 距第一题的秒数
    timestamp: str       # 可读时间戳
    source: str          # "baseline" | "our"
    cov_L0: float        # 0-100
    cov_L1: float
    cov_L2: float
    n_L0_hit: int
    n_L1_hit: int
    n_L2_hit: int


# ─────────────────────────────────────────────────────────────────────────────
# Universe loader
# ─────────────────────────────────────────────────────────────────────────────

def load_universes(snapshot_dir: Path) -> Dict[Tuple[str, int], Universe]:
    """加载所有 universe snapshot, 按 (scene_id, frame_id) 索引."""
    if not snapshot_dir.exists():
        logger.error("snapshot 目录不存在: %s", snapshot_dir)
        return {}

    universes: Dict[Tuple[str, int], Universe] = {}
    files = sorted(snapshot_dir.glob("*_universe.json"))
    logger.info("加载 universe snapshots: %d 个", len(files))

    for f in files:
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            u = Universe(
                scene_id=data["scene_id"],
                frame_id=int(data["frame_id"]),
                L0_set=set(data.get("L0", [])),
                L1_set=set(data.get("L1", [])),
                L2A_set=set(data.get("L2A", [])),
                L2B_set=set(data.get("L2B", [])),
            )
            universes[(u.scene_id, u.frame_id)] = u
        except Exception as exc:
            logger.warning("跳过损坏的 snapshot %s: %s", f.name, exc)

    logger.info("✓ 加载 %d 帧 universe", len(universes))
    return universes


# ─────────────────────────────────────────────────────────────────────────────
# Excel reader
# ─────────────────────────────────────────────────────────────────────────────

def _parse_l0(cell_value: Any) -> Set[str]:
    """L0 列存的是 JSON list 字符串, e.g. '[\"ego\",\"car1\"]'."""
    if not cell_value:
        return set()
    try:
        data = json.loads(str(cell_value))
        if isinstance(data, list):
            return {str(x) for x in data if x}
    except Exception:
        pass
    return set()


def _parse_l1(cell_value: Any) -> Set[str]:
    """L1 列存的是 JSON list of {source, target}."""
    if not cell_value:
        return set()
    try:
        data = json.loads(str(cell_value))
        if isinstance(data, list):
            keys = set()
            for item in data:
                if isinstance(item, dict):
                    s, t = item.get("source", ""), item.get("target", "")
                    if s and t:
                        keys.add(f"{s}|{t}")
                elif isinstance(item, (list, tuple)) and len(item) >= 2:
                    keys.add(f"{item[0]}|{item[1]}")
            return keys
    except Exception:
        pass
    return set()


def _parse_l2(cell_value: Any) -> Set[str]:
    """L2 列存的是 JSON list of {o1, o2, o3}."""
    if not cell_value:
        return set()
    try:
        data = json.loads(str(cell_value))
        if isinstance(data, list):
            keys = set()
            for item in data:
                if isinstance(item, dict):
                    o1, o2, o3 = item.get("o1", ""), item.get("o2", ""), item.get("o3", "")
                    if o1 and o2 and o3:
                        keys.add(f"{o1}|{o2}|{o3}")
                elif isinstance(item, (list, tuple)) and len(item) >= 3:
                    keys.add(f"{item[0]}|{item[1]}|{item[2]}")
            return keys
    except Exception:
        pass
    return set()


def _parse_ts(s: Any) -> Optional[datetime]:
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def read_qa_rows(excel_path: Path) -> List[QARow]:
    """从 Excel 的 raw_coverage + question-answer-our 表读 QA, 解析成统一结构."""
    import openpyxl

    if not excel_path.exists():
        logger.error("Excel 不存在: %s", excel_path)
        return []

    logger.info("打开 Excel: %s", excel_path)
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    rows: List[QARow] = []

    def _read_sheet(sheet_name: str, source_label: str, qid_col: str):
        if sheet_name not in wb.sheetnames:
            logger.warning("  sheet '%s' 不存在, 跳过", sheet_name)
            return 0
        ws = wb[sheet_name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            return 0
        headers = [str(h) if h else "" for h in first]
        col = {h: i for i, h in enumerate(headers)}

        def _g(row, name, default=""):
            i = col.get(name, -1)
            return row[i] if 0 <= i < len(row) else default

        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            scene_id = _g(row, "scene_id")
            if not scene_id:
                continue
            fid_raw = _g(row, "frame_id")
            if fid_raw is None or fid_raw == "":
                continue
            try:
                frame_id = int(fid_raw)
            except Exception:
                continue
            qid = _g(row, qid_col) or ""
            ts_end = _parse_ts(_g(row, "timestamp_end"))
            ts_start = _parse_ts(_g(row, "timestamp_start"))
            ts = ts_end or ts_start
            l0 = _parse_l0(_g(row, "L0"))
            l1 = _parse_l1(_g(row, "L1"))
            l2 = _parse_l2(_g(row, "L2"))
            rows.append(QARow(
                source=source_label, scene_id=str(scene_id), frame_id=frame_id,
                qid=str(qid), timestamp=ts,
                L0_keys=l0, L1_keys=l1, L2_keys=l2,
            ))
            n += 1
        logger.info("  %-22s %d rows", sheet_name + ":", n)
        return n

    _read_sheet(SHEET_RAW, "baseline", "nuscenes_qa_id")
    _read_sheet(SHEET_OUR, "our",      "question_id")

    wb.close()
    logger.info("✓ 共加载 %d 行 QA", len(rows))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Curve computation
# ─────────────────────────────────────────────────────────────────────────────

def compute_curve_for_frame(
    universe: Universe,
    frame_rows: List[QARow],
) -> List[CoverageCurvePoint]:
    """
    对一帧的所有 QA 按时间顺序累积计算覆盖率曲线。
    Baseline 题先被处理 (按时间), 然后 our 题接着累积。
    """
    def _sort_key(r: QARow):
        src_priority = 0 if r.source == "baseline" else 1
        ts = r.timestamp or datetime.min
        return (src_priority, ts, r.qid)

    sorted_rows = sorted(frame_rows, key=_sort_key)
    if not sorted_rows:
        return []

    # 找第一个有 timestamp 的题作为 t0
    t0 = next((r.timestamp for r in sorted_rows if r.timestamp), None) or datetime.min

    hit_L0: Set[str] = set()
    hit_L1: Set[str] = set()
    hit_L2: Set[str] = set()

    U_L0 = universe.L0_set
    U_L1 = universe.L1_set
    U_L2 = universe.L2_set
    n_L0 = max(len(U_L0), 1)
    n_L1 = max(len(U_L1), 1)
    n_L2 = max(len(U_L2), 1)

    points: List[CoverageCurvePoint] = []
    for i, r in enumerate(sorted_rows, 1):
        # 与全集求交: 题里有但不在全集里的 keys 不算入分子
        hit_L0 |= (r.L0_keys & U_L0)
        hit_L1 |= (r.L1_keys & U_L1)
        hit_L2 |= (r.L2_keys & U_L2)

        secs = (r.timestamp - t0).total_seconds() if r.timestamp else 0.0

        points.append(CoverageCurvePoint(
            x_qcount=i,
            x_seconds=secs,
            timestamp=(r.timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                       if r.timestamp else ""),
            source=r.source,
            cov_L0=round(100 * len(hit_L0) / n_L0, 4),
            cov_L1=round(100 * len(hit_L1) / n_L1, 4),
            cov_L2=round(100 * len(hit_L2) / n_L2, 4),
            n_L0_hit=len(hit_L0),
            n_L1_hit=len(hit_L1),
            n_L2_hit=len(hit_L2),
        ))
    return points


# ─────────────────────────────────────────────────────────────────────────────
# Plotting
# ─────────────────────────────────────────────────────────────────────────────

def plot_curve(
    universe: Universe,
    points: List[CoverageCurvePoint],
    out_dir: Path,
):
    """画两张图: 按题号 X 轴 + 按时间 X 轴。"""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        logger.warning("matplotlib 未安装, 跳过画图 (pip install matplotlib)")
        return

    if not points:
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    base = f"{universe.scene_id}_frame{universe.frame_id}"

    # 找 baseline → our 切换点
    switch_idx = next((i for i, p in enumerate(points) if p.source == "our"), -1)

    for x_mode, x_label, x_attr in [
        ("qcount", "Question count",  "x_qcount"),
        ("time",   "Elapsed seconds", "x_seconds"),
    ]:
        fig, ax = plt.subplots(figsize=(9, 5.5))
        xs = [getattr(p, x_attr) for p in points]
        ax.plot(xs, [p.cov_L0 for p in points], "-", label="L0 (nodes)", color="#1f77b4", linewidth=1.8)
        ax.plot(xs, [p.cov_L1 for p in points], "-", label="L1 (edges)", color="#ff7f0e", linewidth=1.8)
        ax.plot(xs, [p.cov_L2 for p in points], "-", label="L2 (paths)", color="#2ca02c", linewidth=1.8)

        if 0 < switch_idx < len(points):
            split_x = xs[switch_idx]
            ax.axvline(split_x, color="gray", linestyle="--", alpha=0.6, linewidth=1)
            ax.text(split_x, 5, " baseline → our", fontsize=8, color="gray")

        ax.set_xlabel(x_label)
        ax.set_ylabel("Coverage rate (%)")
        ax.set_title(f"{universe.scene_id} frame{universe.frame_id}  "
                     f"(L0 total={universe.totals['L0']}, "
                     f"L1={universe.totals['L1']}, L2={universe.totals['L2']})")
        ax.set_ylim(0, 105)
        ax.grid(True, alpha=0.3)
        ax.legend(loc="lower right")
        fig.tight_layout()
        out_path = out_dir / f"{base}_curve_by_{x_mode}.png"
        fig.savefig(out_path, dpi=150)
        plt.close(fig)
        logger.info("  ✓ %s", out_path.name)


def write_curve_csv(universe: Universe, points: List[CoverageCurvePoint], out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{universe.scene_id}_frame{universe.frame_id}_curve_data.csv"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("scene_id,frame_id,x_qcount,x_seconds,timestamp,source,"
                "cov_L0,cov_L1,cov_L2,n_L0_hit,n_L1_hit,n_L2_hit,"
                "total_L0,total_L1,total_L2\n")
        t = universe.totals
        for p in points:
            f.write(f"{universe.scene_id},{universe.frame_id},"
                    f"{p.x_qcount},{p.x_seconds},{p.timestamp},{p.source},"
                    f"{p.cov_L0},{p.cov_L1},{p.cov_L2},"
                    f"{p.n_L0_hit},{p.n_L1_hit},{p.n_L2_hit},"
                    f"{t['L0']},{t['L1']},{t['L2']}\n")
    logger.info("  ✓ %s", out_path.name)


# ─────────────────────────────────────────────────────────────────────────────
# Excel write-back to RQ2_graph
# ─────────────────────────────────────────────────────────────────────────────

RQ2_HEADERS = [
    "scene_id", "frame_id", "x_qcount", "x_seconds", "timestamp", "source",
    "cov_L0", "cov_L1", "cov_L2", "n_L0_hit", "n_L1_hit", "n_L2_hit",
    "total_L0", "total_L1", "total_L2",
]


def write_rq2_graph(
    excel_path: Path,
    all_curves: List[Tuple[Universe, List[CoverageCurvePoint]]],
):
    """把所有曲线数据点写进 RQ2_graph sheet (覆盖式: 先清后写)."""
    import openpyxl
    if not excel_path.exists():
        logger.warning("Excel 不存在, 跳过 RQ2_graph 写入")
        return

    logger.info("写回 Excel %s -> sheet %s", excel_path.name, SHEET_RQ2)
    wb = openpyxl.load_workbook(str(excel_path))
    if SHEET_RQ2 not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_RQ2)
    else:
        ws = wb[SHEET_RQ2]
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)

    ws.append(RQ2_HEADERS)
    n_rows = 0
    for u, points in all_curves:
        t = u.totals
        for p in points:
            ws.append([
                u.scene_id, u.frame_id, p.x_qcount, p.x_seconds, p.timestamp, p.source,
                p.cov_L0, p.cov_L1, p.cov_L2,
                p.n_L0_hit, p.n_L1_hit, p.n_L2_hit,
                t["L0"], t["L1"], t["L2"],
            ])
            n_rows += 1
    wb.save(str(excel_path))
    wb.close()
    logger.info("  ✓ %d 行写入 RQ2_graph", n_rows)


# ─────────────────────────────────────────────────────────────────────────────
# Summary
# ─────────────────────────────────────────────────────────────────────────────

def write_summary_csv(
    out_dir: Path,
    all_curves: List[Tuple[Universe, List[CoverageCurvePoint]]],
):
    out_path = out_dir / "_all_frames_summary.csv"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("scene_id,frame_id,total_L0,total_L1,total_L2,"
                "n_baseline_q,n_our_q,n_total_q,"
                "final_cov_L0,final_cov_L1,final_cov_L2,"
                "baseline_end_cov_L0,baseline_end_cov_L1,baseline_end_cov_L2\n")
        for u, points in all_curves:
            if not points:
                continue
            n_base = sum(1 for p in points if p.source == "baseline")
            base_end = points[n_base - 1] if n_base > 0 else points[0]
            final = points[-1]
            t = u.totals
            f.write(f"{u.scene_id},{u.frame_id},{t['L0']},{t['L1']},{t['L2']},"
                    f"{n_base},{len(points) - n_base},{len(points)},"
                    f"{final.cov_L0},{final.cov_L1},{final.cov_L2},"
                    f"{base_end.cov_L0},{base_end.cov_L1},{base_end.cov_L2}\n")
    logger.info("✓ 全帧汇总: %s", out_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--excel",        type=Path, default=DEFAULT_EXCEL)
    ap.add_argument("--snapshot-dir", type=Path, default=DEFAULT_SNAPSHOT_DIR)
    ap.add_argument("--out-dir",      type=Path, default=DEFAULT_OUT_DIR)
    ap.add_argument("--frame", nargs=2, metavar=("SCENE", "FRAME"), default=None,
                    help="只算单帧, e.g. --frame scene-0926 20")
    ap.add_argument("--no-excel-write", action="store_true",
                    help="不写回 Excel RQ2_graph")
    args = ap.parse_args()

    universes = load_universes(args.snapshot_dir)
    if not universes:
        logger.error("没有 universe snapshot, 请先在 run_*.py 里调用 dump_universe_snapshot()")
        sys.exit(1)

    qa_rows = read_qa_rows(args.excel)
    if not qa_rows:
        logger.error("Excel 没读到任何 QA, 请检查路径和 sheet 名")
        sys.exit(1)

    qa_by_frame: Dict[Tuple[str, int], List[QARow]] = collections.defaultdict(list)
    for r in qa_rows:
        qa_by_frame[(r.scene_id, r.frame_id)].append(r)

    if args.frame:
        target = (args.frame[0], int(args.frame[1]))
        if target not in universes:
            logger.error("未找到 universe snapshot for %s", target)
            sys.exit(1)
        targets = [target]
    else:
        targets = sorted(universes.keys())

    args.out_dir.mkdir(parents=True, exist_ok=True)
    all_curves: List[Tuple[Universe, List[CoverageCurvePoint]]] = []

    for key in targets:
        scene_id, frame_id = key
        u = universes[key]
        rows = qa_by_frame.get(key, [])
        logger.info("=" * 60)
        logger.info("分析 %s frame%d  (%d QA rows, totals L0=%d L1=%d L2=%d)",
                    scene_id, frame_id, len(rows),
                    u.totals["L0"], u.totals["L1"], u.totals["L2"])

        if not rows:
            logger.warning("  ⚠ 该帧在 Excel 里没有任何 QA, 跳过")
            continue

        points = compute_curve_for_frame(u, rows)
        if not points:
            continue

        write_curve_csv(u, points, args.out_dir)
        plot_curve(u, points, args.out_dir)
        all_curves.append((u, points))

        n_base = sum(1 for p in points if p.source == "baseline")
        final = points[-1]
        logger.info("  → baseline=%d our=%d total=%d  最终覆盖率: L0=%.1f%% L1=%.1f%% L2=%.1f%%",
                    n_base, len(points) - n_base, len(points),
                    final.cov_L0, final.cov_L1, final.cov_L2)

    if all_curves:
        write_summary_csv(args.out_dir, all_curves)
        if not args.no_excel_write:
            write_rq2_graph(args.excel, all_curves)

    logger.info("=" * 60)
    logger.info("✓ 完成. 输出目录: %s", args.out_dir)


if __name__ == "__main__":
    main()
