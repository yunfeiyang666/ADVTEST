#!/usr/bin/env python3
"""
重建服务器 Excel 文件
根据 ADVTEST_EXCEL_PATH 环境变量或默认路径创建 Excel
"""
import openpyxl
import os
from pathlib import Path

# 从环境变量读取或使用默认路径
EXCEL_PATH_STR = os.getenv("ADVTEST_EXCEL_PATH", "")
if EXCEL_PATH_STR:
    EXCEL_PATH = Path(EXCEL_PATH_STR).expanduser()
else:
    # 默认路径：~/ADVTEST/data/RQ_nuscenesqa_val_full.xlsx
    EXCEL_PATH = Path.home() / "ADVTEST" / "data" / "RQ_nuscenesqa_val_full.xlsx"

# Sheet 名称
SHEET_A = "raw_coverage"
SHEET_B = "question-answer-our"
SHEET_C = "model_performance_raw_our"
SHEET_FR = "filter_record"

# 各 sheet 的列定义
COLUMNS = {
    SHEET_A: [
        "qa_unique_id",
        "scene_id",
        "frame_id",
        "question",
        "answer",
        "q_type",
        "num_hop",
        "l0_nodes",
        "l1_edges",
        "l2_paths",
        "n_l0",
        "n_l1",
        "n_l2",
        "llm_ms",
        "success",
        "timestamp",
    ],
    SHEET_B: [
        "qa_unique_id",
        "scene_id",
        "frame_id",
        "question",
        "answer",
        "q_type",
        "num_hop",
        "l0_nodes",
        "l1_edges",
        "l2_paths",
        "n_l0",
        "n_l1",
        "n_l2",
        "llm_ms",
        "success",
        "timestamp",
    ],
    SHEET_C: [
        "qa_unique_id",
        "scene_id",
        "frame_id",
        "question",
        "answer",
        "model_name",
        "model_answer",
        "is_correct",
        "eval_time_ms",
        "timestamp",
    ],
    SHEET_FR: [
        "scene_id",
        "frame_id",
        "original_nodes_count",
        "filtered_nodes_count",
        "ratio",
    ],
}


def rebuild_excel():
    """重建 Excel 文件"""
    print("="*80)
    print("Rebuilding Excel for Server")
    print("="*80)
    print(f"\nTarget: {EXCEL_PATH}")

    # 确保父目录存在
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"Parent dir: {EXCEL_PATH.parent}")

    # 备份旧文件（如果存在）
    if EXCEL_PATH.exists():
        backup_path = EXCEL_PATH.with_suffix('.xlsx.backup')
        print(f"\n[Backup] Moving old file to: {backup_path}")
        try:
            if backup_path.exists():
                backup_path.unlink()
            EXCEL_PATH.rename(backup_path)
            print("  [OK] Backup created")
        except Exception as e:
            print(f"  [ERROR] Backup failed: {e}")
            print("  Continuing anyway...")

    # 创建新的 workbook
    print("\n[Create] Creating new workbook...")
    wb = openpyxl.Workbook()

    # 删除默认的 Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 创建所有 sheet
    for sheet_name, columns in COLUMNS.items():
        print(f"\n[Sheet] Creating '{sheet_name}'...")
        ws = wb.create_sheet(sheet_name)

        # 写入表头
        ws.append(columns)
        print(f"  [OK] {len(columns)} columns: {columns[:3]}...")

    # 保存文件
    print(f"\n[Save] Saving to {EXCEL_PATH}...")
    try:
        wb.save(str(EXCEL_PATH))
        wb.close()
        print("  [OK] File saved successfully")
    except Exception as e:
        print(f"  [ERROR] Save failed: {e}")
        return False

    # 验证文件
    print("\n[Verify] Verifying file structure...")
    try:
        wb_verify = openpyxl.load_workbook(str(EXCEL_PATH))
        print(f"  [OK] File can be opened")
        print(f"  [OK] Sheets: {wb_verify.sheetnames}")

        for sheet_name in COLUMNS.keys():
            ws = wb_verify[sheet_name]
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            cols = [str(v) for v in row1 if v is not None]
            print(f"  [OK] [{sheet_name}] {len(cols)} columns")

        wb_verify.close()
        print("\n" + "="*80)
        print("[OK] Excel rebuilt successfully")
        print("="*80)
        return True

    except Exception as e:
        print(f"  [ERROR] Verification failed: {e}")
        return False


if __name__ == "__main__":
    success = rebuild_excel()

    if success:
        print("\nNext steps:")
        print("  1. Restart V19 production")
        print("  2. Monitor log for Excel write success")
        print("  3. Check Excel file for data")
    else:
        print("\n[ERROR] Rebuild failed - please check errors above")
